"""config — FR-002: Configuration loading and validation.

Loads and validates all configuration files (YAML patterns, Excel lookup tables,
output template). Compiles regex patterns, builds normalized lookup dictionaries,
and returns an ``AppConfig`` instance used by all downstream modules.
"""

import logging
import re
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook

from .errors import ConfigError, ErrorCode
from .models import AppConfig, FieldPattern
from .utils import normalize_lookup_key

logger = logging.getLogger(__name__)

_REQUIRED_TOP_KEYS: list[str] = [
    "invoice_sheet",
    "packing_sheet",
    "invoice_columns",
    "packing_columns",
    "inv_no_cell",
]

_REQUIRED_INVOICE_FIELDS: list[str] = [
    "part_no",
    "po_no",
    "qty",
    "price",
    "amount",
    "currency",
    "coo",
    "brand",
    "brand_type",
    "model",
    "cod",
    "weight",
    "inv_no",
    "serial",
]

_REQUIRED_PACKING_FIELDS: list[str] = [
    "part_no",
    "po_no",
    "qty",
    "nw",
    "gw",
    "pack",
]


def load_config(config_dir: Path) -> AppConfig:
    """Load and validate all configuration files from config_dir.

    Compiles regex patterns, builds normalized lookup tables, and validates
    the output template structure. Raises ConfigError on any failure.

    Args:
        config_dir: Path to the directory containing field_patterns.yaml,
            currency_rules.xlsx, country_rules.xlsx, and output_template.xlsx.

    Returns:
        AppConfig with compiled patterns, lookup tables, and template path.

    Raises:
        ConfigError: ERR_001 through ERR_005 on validation failures.
    """
    # Step 1: Load and validate field_patterns.yaml
    yaml_path = config_dir / "field_patterns.yaml"
    yaml_data = _load_yaml(yaml_path)
    _validate_yaml_keys(yaml_data, yaml_path)
    logger.debug("Loaded and validated field_patterns.yaml")

    # Step 2: Compile all regex patterns from YAML
    invoice_sheet_patterns = _compile_patterns(
        yaml_data["invoice_sheet"]["patterns"],
        "invoice_sheet",
        str(yaml_path),
    )
    packing_sheet_patterns = _compile_patterns(
        yaml_data["packing_sheet"]["patterns"],
        "packing_sheet",
        str(yaml_path),
    )
    inv_no_cell = yaml_data["inv_no_cell"]
    inv_no_patterns = _compile_patterns(
        inv_no_cell["patterns"],
        "inv_no_cell",
        str(yaml_path),
    )
    inv_no_label_patterns = _compile_patterns(
        inv_no_cell["label_patterns"],
        "inv_no_cell.label_patterns",
        str(yaml_path),
    )
    inv_no_exclude_patterns = _compile_patterns(
        inv_no_cell["exclude_patterns"],
        "inv_no_cell.exclude_patterns",
        str(yaml_path),
    )
    logger.debug("Compiled all regex patterns")

    # Step 3: Build FieldPattern objects for invoice and packing columns
    invoice_columns = _build_field_patterns(
        yaml_data["invoice_columns"],
        "invoice_columns",
        str(yaml_path),
    )
    packing_columns = _build_field_patterns(
        yaml_data["packing_columns"],
        "packing_columns",
        str(yaml_path),
    )
    logger.debug("Built FieldPattern objects for invoice and packing columns")

    # Step 4: Load and validate currency_rules.xlsx
    currency_path = config_dir / "currency_rules.xlsx"
    currency_lookup = _load_lookup_table(
        currency_path,
        sheet_name="Currency_Rules",
        source_col="Source_Value",
        target_col="Target_Code",
    )
    logger.debug("Loaded currency lookup table (%d entries)", len(currency_lookup))

    # Step 5: Load and validate country_rules.xlsx
    country_path = config_dir / "country_rules.xlsx"
    country_lookup = _load_lookup_table(
        country_path,
        sheet_name="Country_Rules",
        source_col="Source_Value",
        target_col="Target_Code",
    )
    logger.debug("Loaded country lookup table (%d entries)", len(country_lookup))

    # Step 6: Load and validate output_template.xlsx structure
    template_path = config_dir / "output_template.xlsx"
    _validate_template(template_path)
    logger.debug("Validated output template structure")

    # Step 7: Construct and return AppConfig
    return AppConfig(
        invoice_sheet_patterns=invoice_sheet_patterns,
        packing_sheet_patterns=packing_sheet_patterns,
        invoice_columns=invoice_columns,
        packing_columns=packing_columns,
        inv_no_patterns=inv_no_patterns,
        inv_no_label_patterns=inv_no_label_patterns,
        inv_no_exclude_patterns=inv_no_exclude_patterns,
        currency_lookup=currency_lookup,
        country_lookup=country_lookup,
        template_path=template_path,
    )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _load_yaml(yaml_path: Path) -> dict[str, Any]:
    """Load YAML file. Raises ERR_001 if missing, ERR_004 if not a dict.

    Args:
        yaml_path: Path to the YAML file.

    Returns:
        Parsed YAML content as a dictionary.
    """
    if not yaml_path.exists():
        raise ConfigError(
            code=ErrorCode.ERR_001,
            message=f"Required configuration file not found: {yaml_path}",
            path=str(yaml_path),
        )
    with open(yaml_path, encoding="utf-8") as fh:
        data: Any = yaml.safe_load(fh)
    # Reason: yaml.safe_load returns Any; must verify top-level is dict.
    if not isinstance(data, dict):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=f"YAML file is malformed (expected dict, got {type(data).__name__}): {yaml_path}",
            path=str(yaml_path),
        )
    return data


def _validate_yaml_keys(data: dict[str, Any], yaml_path: Path) -> None:
    """Validate all required YAML keys exist. Raises ERR_004 on missing key.

    Args:
        data: Parsed YAML content.
        yaml_path: Path for error reporting.
    """
    path_str = str(yaml_path)

    for key in _REQUIRED_TOP_KEYS:
        if key not in data:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: {key}",
                path=path_str,
            )

    # Validate sheet pattern sections
    for section in ("invoice_sheet", "packing_sheet"):
        val = data[section]
        if not isinstance(val, dict) or "patterns" not in val:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: {section}.patterns",
                path=path_str,
            )

    # Validate invoice_columns subkeys (14 fields)
    inv_cols = data["invoice_columns"]
    if not isinstance(inv_cols, dict):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message="Missing required YAML key: invoice_columns",
            path=path_str,
        )
    for field in _REQUIRED_INVOICE_FIELDS:
        if field not in inv_cols:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: invoice_columns.{field}",
                path=path_str,
            )

    # Validate packing_columns subkeys (6 fields)
    pack_cols = data["packing_columns"]
    if not isinstance(pack_cols, dict):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message="Missing required YAML key: packing_columns",
            path=path_str,
        )
    for field in _REQUIRED_PACKING_FIELDS:
        if field not in pack_cols:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: packing_columns.{field}",
                path=path_str,
            )

    # Validate inv_no_cell subkeys
    inv_no = data["inv_no_cell"]
    if not isinstance(inv_no, dict):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message="Missing required YAML key: inv_no_cell",
            path=path_str,
        )
    for subkey in ("patterns", "label_patterns", "exclude_patterns"):
        if subkey not in inv_no:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: inv_no_cell.{subkey}",
                path=path_str,
            )


def _compile_patterns(
    raw_patterns: Any,
    field_name: str,
    file_path: str,
) -> list[re.Pattern[str]]:
    """Compile regex pattern strings with IGNORECASE. Raises ERR_002 on failure.

    Args:
        raw_patterns: List of regex strings from YAML.
        field_name: Dotted field name for error context.
        file_path: YAML file path for error reporting.

    Returns:
        List of compiled re.Pattern objects.
    """
    if not isinstance(raw_patterns, list):
        raise ConfigError(
            code=ErrorCode.ERR_004,
            message=f"Expected list of patterns for {field_name}, got {type(raw_patterns).__name__}",
            path=file_path,
        )
    compiled: list[re.Pattern[str]] = []
    for idx, pattern in enumerate(raw_patterns):
        try:
            compiled.append(re.compile(str(pattern), re.IGNORECASE))
        except re.error as exc:
            raise ConfigError(
                code=ErrorCode.ERR_002,
                message=f"Invalid regex pattern {field_name}.pattern[{idx}]: {pattern!r} — {exc}",
                path=file_path,
            ) from exc
    return compiled


def _build_field_patterns(
    columns_data: dict[str, Any],
    section_name: str,
    file_path: str,
) -> dict[str, FieldPattern]:
    """Build FieldPattern objects from YAML column definitions.

    Validates pattern compilation (ERR_002) but stores raw strings in FieldPattern.

    Args:
        columns_data: Dict of field_name -> {patterns, type, required} from YAML.
        section_name: Section name for error context (e.g. "invoice_columns").
        file_path: YAML file path for error reporting.

    Returns:
        Dict mapping field name to FieldPattern.
    """
    result: dict[str, FieldPattern] = {}
    for field_name, field_def in columns_data.items():
        if not isinstance(field_def, dict):
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Expected dict for {section_name}.{field_name}, got {type(field_def).__name__}",
                path=file_path,
            )
        if "patterns" not in field_def:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required YAML key: {section_name}.{field_name}.patterns",
                path=file_path,
            )
        raw_patterns = field_def["patterns"]
        if not isinstance(raw_patterns, list):
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=(
                    f"Expected list of patterns for {section_name}.{field_name}.patterns, "
                    f"got {type(raw_patterns).__name__}"
                ),
                path=file_path,
            )
        # Validate compilation (ERR_002) but store raw strings in FieldPattern
        for idx, pattern in enumerate(raw_patterns):
            try:
                re.compile(str(pattern), re.IGNORECASE)
            except re.error as exc:
                raise ConfigError(
                    code=ErrorCode.ERR_002,
                    message=f"Invalid regex pattern {section_name}.{field_name}.pattern[{idx}]: {pattern!r} — {exc}",
                    path=file_path,
                ) from exc
        result[field_name] = FieldPattern(
            patterns=[str(p) for p in raw_patterns],
            type=str(field_def.get("type", "string")),
            required=bool(field_def.get("required", False)),
        )
    return result


def _load_lookup_table(
    xlsx_path: Path,
    *,
    sheet_name: str,
    source_col: str,
    target_col: str,
) -> dict[str, str]:
    """Load a lookup table from Excel. Keys normalized, Target_Code cast to str.

    Args:
        xlsx_path: Path to the Excel file.
        sheet_name: Required sheet name.
        source_col: Column header for lookup keys.
        target_col: Column header for lookup values.

    Returns:
        Dict mapping normalized source keys to string target codes.

    Raises:
        ConfigError: ERR_001 if missing, ERR_003 for duplicates, ERR_004 if malformed.
    """
    if not xlsx_path.exists():
        raise ConfigError(
            code=ErrorCode.ERR_001,
            message=f"Required configuration file not found: {xlsx_path}",
            path=str(xlsx_path),
        )
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ConfigError(
                code=ErrorCode.ERR_004,
                message=f"Missing required sheet '{sheet_name}' in {xlsx_path}",
                path=str(xlsx_path),
            )
        ws = wb[sheet_name]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        for col_name in (source_col, target_col):
            if col_name not in headers:
                raise ConfigError(
                    code=ErrorCode.ERR_004,
                    message=f"Missing required column '{col_name}' in sheet '{sheet_name}' of {xlsx_path}",
                    path=str(xlsx_path),
                )
        src_idx = headers.index(source_col)
        tgt_idx = headers.index(target_col)
        lookup: dict[str, str] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            src_val = row[src_idx] if src_idx < len(row) else None
            tgt_val = row[tgt_idx] if tgt_idx < len(row) else None
            if src_val is None:
                continue
            normalized_key = normalize_lookup_key(str(src_val))
            if normalized_key in lookup:
                raise ConfigError(
                    code=ErrorCode.ERR_003,
                    message=f"Duplicate Source_Value '{src_val}' (normalized: '{normalized_key}') in {xlsx_path}",
                    path=str(xlsx_path),
                )
            # Reason: Numeric Target_Code like 502.0 must become "502", not "502.0".
            lookup[normalized_key] = _normalize_target_code(tgt_val)
        return lookup
    finally:
        wb.close()


def _normalize_target_code(value: Any) -> str:
    """Normalize a Target_Code value to str via str(int(float(v))) for numerics.

    Args:
        value: Raw Target_Code cell value.

    Returns:
        Normalized string representation.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return str(int(float(value)))
    str_val = str(value).strip()
    try:
        return str(int(float(str_val)))
    except (ValueError, OverflowError):
        return str_val


def _validate_template(template_path: Path) -> None:
    """Validate output template structure (sheet, columns, rows). Does NOT keep wb open.

    Args:
        template_path: Path to the output template Excel file.

    Raises:
        ConfigError: ERR_001 if missing, ERR_005 if structure is invalid.
    """
    if not template_path.exists():
        raise ConfigError(
            code=ErrorCode.ERR_001,
            message=f"Required configuration file not found: {template_path}",
            path=str(template_path),
        )
    wb = load_workbook(template_path, data_only=True, read_only=True)
    try:
        required_sheet = "\u5de5\u4f5c\u88681"  # 工作表1
        if required_sheet not in wb.sheetnames:
            raise ConfigError(
                code=ErrorCode.ERR_005,
                message=f"Missing required sheet '{required_sheet}' in template {template_path}",
                path=str(template_path),
            )
        ws = wb[required_sheet]
        max_col = ws.max_column
        if max_col is None or max_col < 40:
            raise ConfigError(
                code=ErrorCode.ERR_005,
                message=f"Template has insufficient columns ({max_col}), expected >= 40 (A-AN) in {template_path}",
                path=str(template_path),
            )
        max_row = ws.max_row
        if max_row is None or max_row < 4:
            raise ConfigError(
                code=ErrorCode.ERR_005,
                message=f"Template has insufficient rows ({max_row}), expected >= 4 header rows in {template_path}",
                path=str(template_path),
            )
    finally:
        wb.close()
