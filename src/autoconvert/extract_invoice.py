"""extract_invoice --- FR-011: Invoice item extraction with 13 per-item fields.

Extracts per-item fields from the invoice sheet. Weight is NOT extracted here
(calculated by weight allocation in FR-025).
"""

import logging
import re
from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet

from .errors import ErrorCode, ProcessingError
from .merge_tracker import MergeTracker
from .models import ColumnMapping, InvoiceItem
from .utils import (
    FOOTER_KEYWORDS,
    STOP_KEYWORD_COL_COUNT,
    detect_cell_precision,
    is_stop_keyword,
    safe_decimal,
    strip_unit_suffix,
)

logger = logging.getLogger(__name__)

# Regex for cleaning invoice number prefixes (case-insensitive).
_INV_PREFIX_RE = re.compile(r"^(INV#|NO\.)\s*", re.IGNORECASE)


def extract_invoice_items(
    sheet: Worksheet,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
    inv_no: str | None,
) -> list[InvoiceItem]:
    """Extract 13 per-item fields from the invoice sheet.

    Args:
        sheet: Invoice worksheet (already unmerged by MergeTracker).
        column_map: Column index mapping with header_row and effective_header_row.
        merge_tracker: Pre-initialized MergeTracker for merged cell propagation.
        inv_no: Invoice number from batch orchestration (column or header fallback);
                used to populate inv_no field when inv_no column is absent from
                field_map.

    Returns:
        List of InvoiceItem records in extraction order.

    Raises:
        ProcessingError: ERR_030 for empty required field; ERR_031 for invalid
            numeric.
    """
    field_map = column_map.field_map
    header_row = column_map.header_row
    start_row = column_map.effective_header_row + 1

    items: list[InvoiceItem] = []
    found_first_data = False

    # Determine whether inv_no comes from a column or the parameter.
    has_inv_no_column = "inv_no" in field_map

    for row in range(start_row, sheet.max_row + 1):
        # --- Step 1: Read part_no and qty raw values for blank / stop checks ---
        part_no_col = field_map.get("part_no")
        qty_col = field_map.get("qty")

        raw_part_no = _get_string(sheet, merge_tracker, row, part_no_col, header_row)
        raw_qty_str = _raw_cell_str(sheet, row, qty_col)

        # --- Leading blank row skip ---
        # Reason: A "blank" row has both part_no empty AND qty cell empty.
        # Before the first data row, we skip these without triggering stop.
        if raw_part_no == "" and raw_qty_str == "":
            if not found_first_data:
                logger.debug("Row %d: leading blank row, skipping.", row)
                continue
            # Reason: Even blank-looking rows may contain a stop keyword in a
            # non-part_no column (e.g., TOTAL in the PO No. column). Scan cols
            # A-J before skipping, so TOTAL rows are caught even when
            # part_no/qty are both empty.
            for col_idx in range(1, STOP_KEYWORD_COL_COUNT + 1):
                cell_val = sheet.cell(row=row, column=col_idx).value
                if cell_val is not None and is_stop_keyword(str(cell_val)):
                    logger.debug("Row %d: blank row with stop keyword in col %d, ending extraction.", row, col_idx)
                    break
            else:
                # No stop keyword found — skip blank row normally.
                logger.debug("Row %d: blank row after data, skipping.", row)
                continue
            break  # Stop keyword found — end extraction

        # --- Header continuation filtering ---
        if "part no" in raw_part_no.lower():
            logger.debug("Row %d: header continuation row, skipping.", row)
            continue

        # --- Step 2: Stop conditions ---
        if _should_stop(sheet, row, raw_part_no, raw_qty_str, found_first_data):
            logger.debug("Row %d: stop condition met, ending extraction.", row)
            break

        # --- Step 3: Process data row ---
        found_first_data = True
        item = _extract_row(
            sheet,
            merge_tracker,
            row,
            field_map,
            header_row,
            has_inv_no_column,
            inv_no,
        )
        items.append(item)
        logger.debug(
            "Row %d: extracted part_no=%s, qty=%s, price=%s, amount=%s",
            row,
            item.part_no,
            item.qty,
            item.price,
            item.amount,
        )

    logger.debug("Invoice extraction complete: %d items.", len(items))
    return items


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _should_stop(
    sheet: Worksheet,
    row: int,
    part_no: str,
    qty_str: str,
    found_first_data: bool,
) -> bool:
    """Evaluate stop conditions for the current row.

    Args:
        sheet: The worksheet.
        row: Current 1-based row number.
        part_no: Stripped part_no string for this row.
        qty_str: Raw qty string (stripped, unit-suffix removed) for this row.
        found_first_data: Whether we have already extracted at least one data row.

    Returns:
        True if extraction should stop at this row.
    """
    # Condition 1: empty part_no AND qty == 0 (after first data row).
    if found_first_data and part_no == "":
        try:
            qty_val = float(strip_unit_suffix(qty_str)) if qty_str else None
        except (ValueError, TypeError):
            qty_val = None
        if qty_val is not None and qty_val == 0:
            logger.debug("Row %d: stop — empty part_no + qty=0.", row)
            return True

    # Condition 2: part_no contains "total" (case-insensitive).
    if "total" in part_no.lower():
        logger.debug("Row %d: stop — part_no contains 'total'.", row)
        return True

    # Condition 3: part_no contains a footer keyword.
    part_lower = part_no.lower()
    for kw in FOOTER_KEYWORDS:
        if kw in part_lower:
            logger.debug("Row %d: stop — part_no contains footer keyword '%s'.", row, kw)
            return True

    # Condition 4: any cell in columns A-J contains stop keyword.
    for col_idx in range(1, STOP_KEYWORD_COL_COUNT + 1):
        cell_val = sheet.cell(row=row, column=col_idx).value
        if cell_val is not None and is_stop_keyword(str(cell_val)):
            logger.debug(
                "Row %d, Col %d: stop — cell contains stop keyword '%s'.",
                row,
                col_idx,
                cell_val,
            )
            return True

    return False


def _extract_row(
    sheet: Worksheet,
    merge_tracker: MergeTracker,
    row: int,
    field_map: dict[str, int],
    header_row: int,
    has_inv_no_column: bool,
    inv_no_param: str | None,
) -> InvoiceItem:
    """Extract a single data row into an InvoiceItem.

    Args:
        sheet: The worksheet.
        merge_tracker: MergeTracker for merged cell propagation.
        row: 1-based row number.
        field_map: Column index mapping.
        header_row: 1-based header row for merge_tracker.
        has_inv_no_column: Whether inv_no is in field_map.
        inv_no_param: Fallback inv_no from batch orchestration.

    Returns:
        An InvoiceItem with all 13 fields populated.

    Raises:
        ProcessingError: ERR_030 or ERR_031.
    """
    # --- String fields (via merge_tracker for horizontal merge propagation) ---
    part_no = _get_string(sheet, merge_tracker, row, field_map.get("part_no"), header_row)
    po_no = _get_string(sheet, merge_tracker, row, field_map.get("po_no"), header_row)
    currency = _get_string(sheet, merge_tracker, row, field_map.get("currency"), header_row)
    coo = _get_string(sheet, merge_tracker, row, field_map.get("coo"), header_row)
    brand = _get_string(sheet, merge_tracker, row, field_map.get("brand"), header_row)
    brand_type = _get_string(sheet, merge_tracker, row, field_map.get("brand_type"), header_row)
    model_no = _get_string(sheet, merge_tracker, row, field_map.get("model"), header_row)
    serial = _get_string(sheet, merge_tracker, row, field_map.get("serial"), header_row)

    # COD field (optional).
    cod = _get_string(sheet, merge_tracker, row, field_map.get("cod"), header_row)

    # --- COO/COD fallback (per row) ---
    # Reason: When COO is empty but COD has a value, use COD as COO.
    if coo == "" and cod != "":
        coo = cod

    # --- Invoice number ---
    if has_inv_no_column:
        raw_inv = _get_string(sheet, merge_tracker, row, field_map.get("inv_no"), header_row)
        inv_no_val = _clean_inv_no(raw_inv)
    else:
        inv_no_val = inv_no_param or ""

    # --- Numeric fields ---
    qty = _parse_numeric(sheet, row, field_map.get("qty"), "qty", precision_from_cell=True)
    price = _parse_numeric(sheet, row, field_map.get("price"), "price", fixed_precision=5)
    amount = _parse_numeric(sheet, row, field_map.get("amount"), "amount", fixed_precision=2)

    # --- Required field validation (ERR_030) ---
    # Reason: ERR_030 fires for required fields that are empty AFTER extraction
    # and fallbacks. COO is checked after COD fallback.
    required_fields = {
        "part_no": part_no,
        "po_no": po_no,
        "currency": currency,
        "coo": coo,
        "brand": brand,
        "brand_type": brand_type,
        "model_no": model_no,
    }
    for field_name, value in required_fields.items():
        if value == "":
            raise ProcessingError(
                code=ErrorCode.ERR_030,
                message=f"Required field '{field_name}' is empty at row {row}.",
                context={"row": row, "field_name": field_name},
            )

    return InvoiceItem(
        part_no=part_no,
        po_no=po_no,
        qty=qty,
        price=price,
        amount=amount,
        currency=currency,
        coo=coo,
        cod=cod,
        brand=brand,
        brand_type=brand_type,
        model_no=model_no,
        inv_no=inv_no_val,
        serial=serial,
        allocated_weight=None,
    )


def _get_string(
    sheet: Worksheet,
    merge_tracker: MergeTracker,
    row: int,
    col: int | None,
    header_row: int,
) -> str:
    """Get a string field value via merge_tracker propagation.

    Args:
        sheet: The worksheet.
        merge_tracker: MergeTracker instance.
        row: 1-based row number.
        col: 1-based column index, or None if field not in field_map.
        header_row: 1-based header row for merge_tracker.

    Returns:
        Stripped string value, or empty string if column is absent or value is None.
    """
    if col is None:
        return ""
    raw = merge_tracker.get_string_value(row, col, header_row)
    if raw is None:
        return ""
    try:
        return str(raw).strip()
    except (TypeError, AttributeError):
        return ""


def _raw_cell_str(sheet: Worksheet, row: int, col: int | None) -> str:
    """Get raw cell value as stripped string for blank/stop checks.

    Args:
        sheet: The worksheet.
        row: 1-based row number.
        col: 1-based column index, or None.

    Returns:
        Stripped string of cell value, or empty string.
    """
    if col is None:
        return ""
    val = sheet.cell(row=row, column=col).value
    if val is None:
        return ""
    try:
        return str(val).strip()
    except (TypeError, AttributeError):
        return ""


def _parse_numeric(
    sheet: Worksheet,
    row: int,
    col: int | None,
    field_name: str,
    *,
    precision_from_cell: bool = False,
    fixed_precision: int = 0,
) -> Decimal:
    """Parse a numeric field from a cell, applying unit suffix stripping and safe_decimal.

    Args:
        sheet: The worksheet.
        row: 1-based row number.
        col: 1-based column index.
        field_name: Field name for error reporting.
        precision_from_cell: If True, detect precision from cell format (for qty).
        fixed_precision: Fixed decimal precision (for price, amount).

    Returns:
        Parsed Decimal value.

    Raises:
        ProcessingError: ERR_031 if the value cannot be parsed as numeric.
    """
    if col is None:
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=f"Numeric field '{field_name}' column not mapped at row {row}.",
            context={"row": row, "column": col, "field_name": field_name},
        )

    cell = sheet.cell(row=row, column=col)
    raw_value = cell.value

    if raw_value is None:
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=f"Numeric field '{field_name}' is empty at row {row}, column {col}.",
            context={"row": row, "column": col, "field_name": field_name},
        )

    # Convert to string and strip unit suffixes.
    raw_str = strip_unit_suffix(str(raw_value))

    # Determine precision.
    if precision_from_cell:
        decimals = detect_cell_precision(cell)
    else:
        decimals = fixed_precision

    try:
        result = safe_decimal(raw_str, decimals)
    except (ValueError, TypeError) as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=(f"Invalid numeric value '{raw_value}' for field '{field_name}' at row {row}, column {col}."),
            context={"row": row, "column": col, "field_name": field_name},
        ) from exc

    logger.debug(
        "Row %d, Col %d (%s): raw=%r, stripped=%r, decimals=%d, parsed=%s",
        row,
        col,
        field_name,
        raw_value,
        raw_str,
        decimals,
        result,
    )
    return result


def _clean_inv_no(value: str) -> str:
    """Remove INV# and NO. prefixes from an invoice number string.

    Args:
        value: Raw invoice number string (already stripped).

    Returns:
        Cleaned invoice number string.
    """
    # Reason: Apply prefix removal iteratively so "INV# NO. 12345" becomes "12345".
    cleaned = value
    cleaned = _INV_PREFIX_RE.sub("", cleaned).strip()
    # Apply again in case both prefixes are stacked.
    cleaned = _INV_PREFIX_RE.sub("", cleaned).strip()
    return cleaned
