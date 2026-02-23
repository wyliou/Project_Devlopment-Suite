"""merge_tracker — FR-010: Merged cell capture, unmerge, and value propagation.

Captures merged cell ranges before unmerging, provides value propagation for
string and weight fields per FR-010.  One ``MergeTracker`` instance is created
per sheet, BEFORE any extraction begins.
"""

import logging
from decimal import Decimal
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

from .models import MergeRange

logger = logging.getLogger(__name__)


class MergeTracker:
    """Captures merged cell ranges before unmerging, provides value propagation
    for string and weight fields per FR-010.

    The constructor performs two actions in strict order:
    1. Capture every merged range (bounds + anchor value) into an internal list.
    2. Unmerge all cells in the sheet so downstream code sees flat cells.

    Public query methods then use the captured data to answer merge-related
    questions without touching the (now unmerged) sheet's merge metadata.
    """

    def __init__(self, sheet: Worksheet) -> None:
        """Capture all merge ranges (with anchor values), then unmerge all
        cells in the sheet.

        Args:
            sheet: The openpyxl Worksheet to process.  The sheet is mutated
                in-place: all merged regions are removed.
        """
        self._sheet = sheet
        self._ranges: list[MergeRange] = []

        # Reason: Must snapshot merged_cells.ranges into a plain list BEFORE
        # calling unmerge_cells, because unmerge mutates the underlying set
        # and iterating while mutating raises RuntimeError / skips entries.
        raw_ranges = list(sheet.merged_cells.ranges)  # type: ignore[attr-defined]

        # Phase 1 — capture
        for cell_range in raw_ranges:
            min_row: int = cell_range.min_row
            max_row: int = cell_range.max_row
            min_col: int = cell_range.min_col
            max_col: int = cell_range.max_col
            anchor_value: Any = sheet.cell(row=min_row, column=min_col).value

            mr = MergeRange(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                value=anchor_value,
            )
            self._ranges.append(mr)
            logger.debug(
                "Captured merge range: rows %d-%d, cols %d-%d, anchor=%r",
                min_row,
                max_row,
                min_col,
                max_col,
                anchor_value,
            )

        # Phase 2 — unmerge (using the snapshot so we don't modify while iterating)
        for cell_range in raw_ranges:
            sheet.unmerge_cells(str(cell_range))

        logger.debug(
            "MergeTracker initialised: %d range(s) captured and unmerged.",
            len(self._ranges),
        )

    # ------------------------------------------------------------------
    # Public query helpers
    # ------------------------------------------------------------------

    def is_merge_anchor(self, row: int, col: int) -> bool:
        """Return True if (row, col) is the top-left anchor of any merge range.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            True when the cell is the anchor (min_row, min_col) of a captured
            merge range.
        """
        for mr in self._ranges:
            if mr.min_row == row and mr.min_col == col:
                return True
        return False

    def is_in_merge(self, row: int, col: int) -> bool:
        """Return True if (row, col) falls within any merge range (anchor or
        non-anchor).

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            True when the cell is inside any captured merge range.
        """
        return self._find_range(row, col) is not None

    def get_merge_range(self, row: int, col: int) -> MergeRange | None:
        """Return the MergeRange containing this cell, or None if not merged.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The matching ``MergeRange``, or ``None``.
        """
        return self._find_range(row, col)

    def get_anchor_value(self, row: int, col: int) -> Any:
        """Return the anchor cell's value for a cell in a merge range
        (pre-unmerge).  Works for both anchor and non-anchor cells.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The captured anchor value, or ``None`` if the cell is not in any
            merge range.
        """
        mr = self._find_range(row, col)
        if mr is None:
            return None
        return mr.value

    def is_data_area_merge(self, row: int, col: int, header_row: int) -> bool:
        """Return True if the merge range for this cell starts after
        ``header_row`` (data-area merge).

        A merge that starts at ``header_row`` or before is a header-area merge
        (formatting only) and returns ``False``.

        Args:
            row: 1-based row index.
            col: 1-based column index.
            header_row: 1-based header row number.

        Returns:
            True for data-area merges; False otherwise (including when the cell
            is not in any merge range).
        """
        mr = self._find_range(row, col)
        if mr is None:
            return False
        return mr.min_row > header_row

    def get_string_value(self, row: int, col: int, header_row: int) -> Any:
        """For data-area merges: return anchor value (string propagation).
        For all other cells: return current sheet cell value.

        Args:
            row: 1-based row index.
            col: 1-based column index.
            header_row: 1-based header row number.

        Returns:
            The propagated string value for data-area merges, or the raw
            sheet cell value otherwise.
        """
        if self.is_data_area_merge(row, col, header_row):
            # Reason: For data-area merges we propagate the anchor value so
            # that non-anchor cells (which are None after unmerge) receive the
            # correct string content.
            return self.get_anchor_value(row, col)
        return self._sheet.cell(row=row, column=col).value

    def get_weight_value(self, row: int, col: int, header_row: int) -> Any:
        """For data-area merges: anchor row returns captured numeric value;
        non-anchor rows return ``Decimal('0.0')``.  For non-merged data cells:
        return sheet cell value (raw, not parsed to Decimal).  Returns
        ``Decimal('0.0')`` for non-data-area merges.

        Args:
            row: 1-based row index.
            col: 1-based column index.
            header_row: 1-based header row number.

        Returns:
            ``Decimal('0.0')`` for non-anchor merged rows and header-area
            merges; the captured anchor value for anchor rows; or the raw
            sheet cell value for non-merged cells.
        """
        mr = self._find_range(row, col)
        if mr is None:
            # Reason: Cell is not part of any merge — return raw sheet value
            # and let the caller decide how to parse it.
            return self._sheet.cell(row=row, column=col).value

        # Cell is inside a merge range.
        if mr.min_row <= header_row:
            # Reason: Header-area merge — weight values here are formatting
            # artifacts, not real data.  Return zero.
            return Decimal("0.0")

        # Data-area merge: only the anchor row retains the weight value.
        if row == mr.min_row and col == mr.min_col:
            return mr.value
        return Decimal("0.0")

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _find_range(self, row: int, col: int) -> MergeRange | None:
        """Return the first MergeRange that contains ``(row, col)``.

        Args:
            row: 1-based row index.
            col: 1-based column index.

        Returns:
            The matching ``MergeRange``, or ``None``.
        """
        for mr in self._ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                return mr
        return None
