"""extract_packing — FR-012, FR-013: Packing item extraction and merged weight validation."""

import logging
from decimal import Decimal, InvalidOperation

from openpyxl.worksheet.worksheet import Worksheet

from .errors import ErrorCode, ProcessingError
from .merge_tracker import MergeTracker
from .models import ColumnMapping, PackingItem
from .utils import (
    DITTO_MARKS,
    STOP_KEYWORD_COL_COUNT,
    detect_cell_precision,
    is_stop_keyword,
    safe_decimal,
    strip_unit_suffix,
    try_float,
)

logger = logging.getLogger(__name__)

_NW_DECIMALS = 5  # NW always stored at 5-decimal precision (FR-012)
_PALLET_KEYWORDS = ("plt.", "pallet")
_HEADER_CONTINUATION = "part no"


def extract_packing_items(
    sheet: Worksheet,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
) -> tuple[list[PackingItem], int]:
    """Extract packing items (part_no, qty, nw) from the packing sheet.

    Args:
        sheet: Packing worksheet (already unmerged by MergeTracker).
        column_map: Column index mapping with header_row and effective_header_row.
        merge_tracker: Pre-initialized MergeTracker for merged cell propagation.

    Returns:
        Tuple of (list of PackingItem, last_data_row). last_data_row is the 1-based
        row number of the last extracted item, used by extract_totals.detect_total_row().

    Raises:
        ProcessingError: ERR_030 for empty required field (non-continuation rows);
                         ERR_031 for invalid numeric value.
    """
    part_no_col = column_map.field_map["part_no"]
    qty_col = column_map.field_map["qty"]
    nw_col = column_map.field_map["nw"]
    # Reason: GW column is needed for implicit total row detection (stop condition 3).
    gw_col = column_map.field_map.get("gw")
    header_row = column_map.header_row
    start_row = column_map.effective_header_row + 1

    items: list[PackingItem] = []
    last_data_row = start_row  # fallback if no items extracted
    found_first_data = False

    for row in range(start_row, sheet.max_row + 1):  # type: ignore[operator]
        # Stop condition check — BEFORE blank/filter checks (FR-012 CRITICAL)
        if _check_stop_keyword(sheet, row):
            logger.debug("Row %d: stop keyword detected in columns A–J.", row)
            break

        raw_part_no = sheet.cell(row=row, column=part_no_col).value
        raw_nw = merge_tracker.get_weight_value(row, nw_col, header_row)
        raw_gw_value = sheet.cell(row=row, column=gw_col).value if gw_col else None

        # Merged part_no propagation (FR-012)
        part_no_str = _resolve_part_no(raw_part_no, row, part_no_col, merge_tracker, header_row)

        # Stop condition 2: truly blank row (after first data row)
        raw_qty = sheet.cell(row=row, column=qty_col).value
        if found_first_data and _is_truly_blank(part_no_str, raw_qty, raw_nw):
            logger.debug("Row %d: truly blank row after first data — stopping.", row)
            break

        # Stop condition 3: implicit total row — empty part_no + NW>0 + GW>0,
        # excluding rows where part_no is empty due to vertical merge continuation
        if found_first_data and _is_implicit_total_row(
            part_no_str, raw_nw, raw_gw_value, row, part_no_col, merge_tracker
        ):
            logger.debug("Row %d: implicit total row detected — stopping.", row)
            break

        # Row filtering (skip but don't stop)
        if not found_first_data and not part_no_str.strip():
            continue
        if _HEADER_CONTINUATION in part_no_str.lower():
            continue
        if any(kw in part_no_str.lower() for kw in _PALLET_KEYWORDS):
            logger.debug("Row %d: pallet summary row — skipping.", row)
            continue
        if not part_no_str.strip():
            continue

        found_first_data = True

        nw_value, is_first = _parse_nw(raw_nw, row, nw_col, part_no_str, items, merge_tracker, header_row)
        qty_value = _parse_qty(raw_qty, row, qty_col, part_no_str, items, merge_tracker, header_row)

        # Skip rows where qty=0 AND nw=0 (PO-reference rows)
        if qty_value == Decimal(0) and nw_value == Decimal("0.00000"):
            logger.debug("Row %d: qty=0 and nw=0 — skipping PO-reference row.", row)
            continue

        item = PackingItem(part_no=part_no_str, qty=qty_value, nw=nw_value, is_first_row_of_merge=is_first)
        items.append(item)
        last_data_row = row
        logger.debug(
            "Row %d: extracted PackingItem(part_no=%r, qty=%s, nw=%s, first=%s)",
            row,
            part_no_str,
            qty_value,
            nw_value,
            is_first,
        )

    logger.debug("Packing extraction complete: %d items, last_data_row=%d", len(items), last_data_row)
    return items, last_data_row


def validate_merged_weights(
    packing_items: list[PackingItem],
    merge_tracker: MergeTracker,
    column_map: ColumnMapping,
) -> None:
    """Validate that no different part_nos share a merged NW cell.

    Args:
        packing_items: Extracted packing items from extract_packing_items().
        merge_tracker: MergeTracker with merge range metadata.
        column_map: Column mapping to locate the NW column index.

    Returns:
        None if all merged weight cells are valid.

    Raises:
        ProcessingError: ERR_046 if different part_nos share a merged NW cell.
    """
    nw_col = column_map.field_map["nw"]
    part_no_col = column_map.field_map["part_no"]
    header_row = column_map.header_row

    # Reason: Iterate all captured merge ranges. For each NW-column data-area
    # merge, read part_no values from the sheet for all rows in the range.
    for mr in merge_tracker._ranges:  # noqa: SLF001
        if not (mr.min_col <= nw_col <= mr.max_col):
            continue
        # Only data-area merges (starting after header_row)
        if mr.min_row <= header_row:
            continue

        part_nos_in_range: set[str] = set()
        for row in range(mr.min_row, mr.max_row + 1):
            propagated = merge_tracker.get_string_value(row, part_no_col, header_row)
            part_no_str = _to_str(propagated)
            if part_no_str:
                part_nos_in_range.add(part_no_str)

        if len(part_nos_in_range) > 1:
            raise ProcessingError(
                code=ErrorCode.ERR_046,
                message=(
                    f"Different part numbers {sorted(part_nos_in_range)} share "
                    f"a merged NW cell (rows {mr.min_row}-{mr.max_row})."
                ),
                context={
                    "merge_rows": f"{mr.min_row}-{mr.max_row}",
                    "part_nos": sorted(part_nos_in_range),
                    "nw_col": nw_col,
                },
            )


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _check_stop_keyword(sheet: Worksheet, row: int) -> bool:
    """Return True if any cell in columns A-J contains a stop keyword."""
    for col in range(1, STOP_KEYWORD_COL_COUNT + 1):
        cell_value = sheet.cell(row=row, column=col).value
        if cell_value is not None and is_stop_keyword(str(cell_value)):
            return True
    return False


def _resolve_part_no(
    raw_value: object,
    row: int,
    part_no_col: int,
    merge_tracker: MergeTracker,
    header_row: int,
) -> str:
    """Resolve part_no with merge propagation for vertical merges."""
    # Reason: If part_no is empty and the cell is a non-anchor row of a vertical
    # merge, propagate the anchor value so the row is processed as data continuation.
    if _is_blank_value(raw_value):
        if merge_tracker.is_in_merge(row, part_no_col) and not merge_tracker.is_merge_anchor(row, part_no_col):
            propagated = merge_tracker.get_string_value(row, part_no_col, header_row)
            logger.debug("Row %d: part_no propagated from merge anchor: %r", row, propagated)
            return _to_str(propagated)
        return ""
    return _to_str(raw_value)


def _is_blank_value(value: object) -> bool:
    """Return True if value is None or an empty/whitespace-only string."""
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def _to_str(value: object) -> str:
    """Convert value to a stripped string; None becomes empty string."""
    if value is None:
        return ""
    return str(value).strip()


def _is_truly_blank(part_no_str: str, raw_qty: object, raw_nw: object) -> bool:
    """Return True if all key columns (part_no, qty, nw) are empty."""
    return not part_no_str.strip() and _is_blank_value(raw_qty) and _is_blank_value(raw_nw)


def _is_implicit_total_row(
    part_no_str: str,
    raw_nw: object,
    raw_gw: object,
    row: int,
    part_no_col: int,
    merge_tracker: MergeTracker,
) -> bool:
    """Check for implicit total row: empty part_no + NW>0 + GW>0.

    Excludes rows where part_no is empty due to vertical merge continuation.
    """
    if part_no_str.strip():
        return False
    # Reason: Exclude rows where part_no is empty due to vertical merge —
    # these are data continuations, not total rows (FR-012 gotcha #2).
    if merge_tracker.is_in_merge(row, part_no_col):
        return False
    nw_numeric = try_float(raw_nw)
    gw_numeric = try_float(raw_gw)
    return nw_numeric is not None and nw_numeric > 0 and gw_numeric is not None and gw_numeric > 0


def _parse_nw(
    raw_nw: object,
    row: int,
    nw_col: int,
    part_no_str: str,
    items: list[PackingItem],
    merge_tracker: MergeTracker,
    header_row: int,
) -> tuple[Decimal, bool]:
    """Parse NW value with merge/continuation/ditto handling.

    Returns:
        Tuple of (nw Decimal at 5 decimals, is_first_row_of_merge bool).

    Raises:
        ProcessingError: ERR_030 for empty NW (non-continuation);
                         ERR_031 for invalid numeric NW value.
    """
    # Ditto mark check — text in the NW cell
    raw_nw_str = _to_str(raw_nw)
    if raw_nw_str in DITTO_MARKS:
        logger.debug("Row %d: ditto mark %r in NW — setting nw=0.", row, raw_nw_str)
        return Decimal("0.00000"), False

    # Non-anchor merged row — weight already zero from get_weight_value
    if merge_tracker.is_in_merge(row, nw_col) and not merge_tracker.is_merge_anchor(row, nw_col):
        logger.debug("Row %d: non-anchor NW merge row — nw=0.", row)
        return Decimal("0.00000"), False

    # Implicit NW continuation (no merge, same part_no, empty NW)
    if _is_blank_value(raw_nw):
        if items and items[-1].part_no == part_no_str:
            logger.debug("Row %d: implicit NW continuation (same part_no=%r) — nw=0.", row, part_no_str)
            return Decimal("0.00000"), False
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field 'nw' at row {row}.",
            context={"row": row, "column": nw_col, "field_name": "nw"},
        )

    # Normal numeric NW parsing — strip unit suffix before parsing
    nw_str = strip_unit_suffix(str(raw_nw))
    try:
        nw_value = safe_decimal(nw_str, _NW_DECIMALS)
    except (ValueError, InvalidOperation) as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=f"Invalid numeric value '{raw_nw}' for NW at row {row}.",
            context={"row": row, "column": nw_col, "field_name": "nw", "raw_value": str(raw_nw)},
        ) from exc

    # Reason: Row with genuine NW or merge anchor row is is_first_row_of_merge=True.
    return nw_value, True


def _parse_qty(
    raw_qty: object,
    row: int,
    qty_col: int,
    part_no_str: str,
    items: list[PackingItem],
    merge_tracker: MergeTracker,
    header_row: int,
) -> Decimal:
    """Parse QTY value with merge/continuation handling.

    Returns:
        Decimal QTY value at cell display precision.

    Raises:
        ProcessingError: ERR_030 for empty QTY (non-continuation);
                         ERR_031 for invalid numeric QTY.
    """
    is_in_qty_merge = merge_tracker.is_in_merge(row, qty_col)
    is_anchor = merge_tracker.is_merge_anchor(row, qty_col)

    # Non-anchor merged row: set QTY=0
    if is_in_qty_merge and not is_anchor:
        logger.debug("Row %d: non-anchor QTY merge row — qty=0.", row)
        return Decimal(0)

    # Empty QTY with continuation logic
    if _is_blank_value(raw_qty):
        if items and items[-1].part_no == part_no_str:
            logger.debug("Row %d: implicit QTY continuation (same part_no=%r) — qty=0.", row, part_no_str)
            return Decimal(0)
        raise ProcessingError(
            code=ErrorCode.ERR_030,
            message=f"Empty required field 'qty' at row {row}.",
            context={"row": row, "column": qty_col, "field_name": "qty"},
        )

    # Normal numeric QTY parsing
    qty_str = strip_unit_suffix(str(raw_qty))
    cell = merge_tracker._sheet.cell(row=row, column=qty_col)  # noqa: SLF001
    precision = detect_cell_precision(cell)
    try:
        qty_value = safe_decimal(qty_str, precision)
    except (ValueError, InvalidOperation) as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_031,
            message=f"Invalid numeric value '{raw_qty}' for QTY at row {row}.",
            context={"row": row, "column": qty_col, "field_name": "qty", "raw_value": str(raw_qty)},
        ) from exc

    return qty_value
