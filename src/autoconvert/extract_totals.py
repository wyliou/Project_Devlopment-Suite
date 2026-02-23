"""extract_totals — FR-014, FR-015, FR-016, FR-017: Total row detection and totals extraction."""

import logging
import re
from decimal import Decimal

from openpyxl.worksheet.worksheet import Worksheet

from .errors import ErrorCode, ProcessingError
from .merge_tracker import MergeTracker
from .models import ColumnMapping, PackingTotals
from .utils import (
    STOP_KEYWORD_COL_COUNT,
    WEIGHT_PRECISION_MAX,
    WEIGHT_PRECISION_MIN,
    detect_cell_precision,
    is_stop_keyword,
    safe_decimal,
    strip_unit_suffix,
    try_float,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_MAX_SEARCH_OFFSET = 15
"""Maximum row offset beyond last_data_row to search for total row (FR-014)."""

_PACKETS_MIN = 1
"""Minimum valid total_packets value (FR-017)."""

_PACKETS_MAX = 1000
"""Maximum valid total_packets value (FR-017)."""

# Compiled regex patterns for FR-017 total_packets extraction
_JIAN_SHU_RE = re.compile(r"件[数數]")
_PLT_NUMBER_BEFORE_RE = re.compile(r"(\d+)\s*PLT", re.IGNORECASE)
_PLT_NUMBER_AFTER_RE = re.compile(r"PLT", re.IGNORECASE)
_BREAKDOWN_RE = re.compile(r"^(\d+)\s*[（(]")
_UNIT_SUFFIX_PACKETS_RE = re.compile(r"(\d+)\s*(托|箱|件|CTNS)\b", re.IGNORECASE)
_EMBEDDED_CHINESE_RE = re.compile(r"共(\d+)(托|箱)")
_PALLET_RANGE_RE = re.compile(r"PLT#\d+\(\d+~(\d+)\)", re.IGNORECASE)
# Reason: Pallet-vs-box priority — extract pallet count (托) when both 托 and 件 appear
_PALLET_VS_BOX_RE = re.compile(r"共(\d+)托")


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def detect_total_row(
    sheet: Worksheet,
    last_data_row: int,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
) -> int:
    """Detect total row using two-strategy approach.

    Args:
        sheet: Packing worksheet (already unmerged).
        last_data_row: 1-based row number of last extracted packing item.
        column_map: Column index mapping including nw, gw, part_no indices.
        merge_tracker: For excluding merged continuation rows in Strategy 2.

    Returns:
        1-based row number of the detected total row.

    Raises:
        ProcessingError: ERR_032 if both strategies fail to find a total row.
    """
    start_row = last_data_row + 1
    end_row = last_data_row + _MAX_SEARCH_OFFSET

    # Strategy 1 — keyword search
    result = _strategy1_keyword(sheet, start_row, end_row)
    if result is not None:
        logger.info("Total row detected via Strategy 1 (keyword) at row %d", result)
        return result

    # Strategy 2 — implicit detection (only runs when Strategy 1 returns None)
    result = _strategy2_implicit(sheet, start_row, end_row, column_map, merge_tracker)
    if result is not None:
        logger.info("Total row detected via Strategy 2 (implicit) at row %d", result)
        return result

    # Both strategies failed — raise ERR_032
    raise ProcessingError(
        code=ErrorCode.ERR_032,
        message=(f"Total row not found: searched rows {start_row}-{end_row} using keyword and implicit strategies"),
        context={"last_data_row": last_data_row},
    )


def extract_totals(
    sheet: Worksheet,
    total_row: int,
    column_map: ColumnMapping,
) -> PackingTotals:
    """Extract total_nw (FR-015), total_gw (FR-016), total_packets (FR-017).

    Args:
        sheet: Packing worksheet (already unmerged).
        total_row: 1-based row number of the total row from detect_total_row().
        column_map: Column index mapping.

    Returns:
        PackingTotals with total_nw, total_nw_precision, total_gw, total_gw_precision,
        total_packets (None if not found). ATT_002 warning is logged via logger if
        total_packets not found; caller (batch.py) may also collect it as a warning
        via returned PackingTotals with total_packets=None.

    Raises:
        ProcessingError: ERR_033 if total_nw is non-numeric or missing;
                         ERR_034 if total_gw is non-numeric or missing.
    """
    # FR-015 — total_nw
    nw_col = column_map.field_map["nw"]
    total_nw, total_nw_precision = _extract_weight(sheet, total_row, nw_col, "total_nw", ErrorCode.ERR_033)

    # FR-016 — total_gw (with packaging weight addition check)
    gw_col = column_map.field_map["gw"]
    total_gw, total_gw_precision = _extract_gw_with_packaging_check(sheet, total_row, gw_col)

    # FR-017 — total_packets
    total_packets = _extract_total_packets(sheet, total_row, column_map)

    logger.info(
        "Packing total row at row %d, NW= %s, GW= %s, Packets= %s",
        total_row,
        total_nw,
        total_gw,
        total_packets,
    )

    return PackingTotals(
        total_nw=total_nw,
        total_nw_precision=total_nw_precision,
        total_gw=total_gw,
        total_gw_precision=total_gw_precision,
        total_packets=total_packets,
    )


# ---------------------------------------------------------------------------
# Strategy 1 — Keyword search (FR-014)
# ---------------------------------------------------------------------------


def _strategy1_keyword(sheet: Worksheet, start_row: int, end_row: int) -> int | None:
    """Search for total row by keyword match in columns A-J.

    Args:
        sheet: Packing worksheet.
        start_row: First row to search (1-based).
        end_row: Last row to search (1-based).

    Returns:
        1-based row number of the first row containing a stop keyword, or None.
    """
    for row in range(start_row, end_row + 1):
        for col in range(1, STOP_KEYWORD_COL_COUNT + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is not None and is_stop_keyword(str(cell_value)):
                return row
    return None


# ---------------------------------------------------------------------------
# Strategy 2 — Implicit detection (FR-014)
# ---------------------------------------------------------------------------


def _strategy2_implicit(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    column_map: ColumnMapping,
    merge_tracker: MergeTracker,
) -> int | None:
    """Search for total row by implicit pattern (empty part_no + numeric NW/GW).

    Args:
        sheet: Packing worksheet.
        start_row: First row to search (1-based).
        end_row: Last row to search (1-based).
        column_map: Column index mapping.
        merge_tracker: For excluding merged continuation rows.

    Returns:
        1-based row number of the first matching row, or None.
    """
    part_no_col = column_map.field_map["part_no"]
    nw_col = column_map.field_map["nw"]
    gw_col = column_map.field_map["gw"]

    for row in range(start_row, end_row + 1):
        # Check part_no is empty
        part_no_value = sheet.cell(row=row, column=part_no_col).value
        if part_no_value is not None and str(part_no_value).strip() != "":
            continue

        # Exclude merged cell continuation rows
        if merge_tracker.is_in_merge(row, part_no_col):
            continue

        # Check NW > 0
        nw_value = try_float(sheet.cell(row=row, column=nw_col).value)
        if nw_value is None or nw_value <= 0:
            continue

        # Check GW > 0
        gw_value = try_float(sheet.cell(row=row, column=gw_col).value)
        if gw_value is None or gw_value <= 0:
            continue

        return row

    return None


# ---------------------------------------------------------------------------
# FR-015 / FR-016 — Weight extraction helpers
# ---------------------------------------------------------------------------


def _extract_weight(
    sheet: Worksheet,
    row: int,
    col: int,
    field_name: str,
    error_code: ErrorCode,
) -> tuple[Decimal, int]:
    """Extract a weight value from a cell with precision detection.

    Args:
        sheet: Packing worksheet.
        row: 1-based row number.
        col: 1-based column number.
        field_name: Field name for error messages (e.g. "total_nw").
        error_code: Error code to raise if value is invalid.

    Returns:
        Tuple of (Decimal value, precision).

    Raises:
        ProcessingError: With the given error_code if cell value is non-numeric
            or missing.
    """
    cell = sheet.cell(row=row, column=col)
    raw_value = cell.value

    if raw_value is None:
        raise ProcessingError(
            code=error_code,
            message=f"{field_name} is missing at row {row}, column {col}",
            context={"row": row, "column": col, "field": field_name},
        )

    # Strip unit suffix if value is a string
    if isinstance(raw_value, str):
        stripped = strip_unit_suffix(raw_value)
        if stripped == "":
            raise ProcessingError(
                code=error_code,
                message=f"{field_name} is non-numeric at row {row}, column {col}: {raw_value!r}",
                context={"row": row, "column": col, "field": field_name, "raw_value": raw_value},
            )
        raw_value = stripped

    # Detect precision from cell format
    precision = detect_cell_precision(cell)

    # Normalize General format precision: strip trailing zeros, min 2
    precision = _normalize_precision(precision, cell, raw_value)

    try:
        result = safe_decimal(raw_value, precision)
    except (ValueError, TypeError) as exc:
        raise ProcessingError(
            code=error_code,
            message=f"{field_name} is non-numeric at row {row}, column {col}: {raw_value!r}",
            context={"row": row, "column": col, "field": field_name, "raw_value": raw_value},
        ) from exc

    return result, precision


def _normalize_precision(precision: int, cell: object, raw_value: object) -> int:
    """Normalize precision with bounds and General-format trailing zero removal.

    For General format (detect_cell_precision returns the raw trailing-zero count),
    this function applies the min 2, max 5 bounds.

    For explicit formats, it also clamps to [2, 5].

    Args:
        precision: Raw precision from detect_cell_precision().
        cell: The openpyxl cell object.
        raw_value: The processed value (after unit stripping).

    Returns:
        Clamped precision value (min 2, max 5).
    """
    # Clamp precision to bounds
    if precision < WEIGHT_PRECISION_MIN:
        precision = WEIGHT_PRECISION_MIN
    if precision > WEIGHT_PRECISION_MAX:
        precision = WEIGHT_PRECISION_MAX

    return precision


def _extract_gw_with_packaging_check(
    sheet: Worksheet,
    total_row: int,
    gw_col: int,
) -> tuple[Decimal, int]:
    """Extract total_gw with packaging weight addition check (FR-016).

    Checks +1 and +2 rows in GW column. If both are numeric, uses +2 row
    as final total_gw (handles pallet weight addition).

    Args:
        sheet: Packing worksheet.
        total_row: 1-based total row number.
        gw_col: 1-based GW column index.

    Returns:
        Tuple of (Decimal value, precision).

    Raises:
        ProcessingError: ERR_034 if total_row GW cell is non-numeric or missing.
    """
    # First extract from total_row (validates and raises ERR_034 if invalid)
    total_gw, total_gw_precision = _extract_weight(sheet, total_row, gw_col, "total_gw", ErrorCode.ERR_034)

    # Check +1 row
    plus1_value = try_float(sheet.cell(row=total_row + 1, column=gw_col).value)
    if plus1_value is None:
        # No additional rows — use total_row value
        return total_gw, total_gw_precision

    # Check +2 row
    plus2_value = try_float(sheet.cell(row=total_row + 2, column=gw_col).value)
    if plus2_value is None:
        # Only +1 is numeric (but not +2) — use total_row value
        return total_gw, total_gw_precision

    # Both +1 and +2 are numeric — use +2 row as final total_gw
    # Reason: Handles pallet weight addition pattern where row+1 = pallet weight,
    # row+2 = final total including pallet weight.
    plus2_cell = sheet.cell(row=total_row + 2, column=gw_col)
    plus2_raw = plus2_cell.value

    # Strip unit suffix if string
    if isinstance(plus2_raw, str):
        plus2_raw = strip_unit_suffix(plus2_raw)

    plus2_precision = detect_cell_precision(plus2_cell)
    plus2_precision = _normalize_precision(plus2_precision, plus2_cell, plus2_raw)

    try:
        final_gw = safe_decimal(plus2_raw, plus2_precision)
    except (ValueError, TypeError) as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_034,
            message=(
                f"total_gw packaging weight row+2 is non-numeric at row {total_row + 2}, column {gw_col}: {plus2_raw!r}"
            ),
            context={
                "row": total_row + 2,
                "column": gw_col,
                "field": "total_gw",
                "raw_value": plus2_raw,
            },
        ) from exc

    logger.info(
        "Packaging weight addition: using row %d GW=%s (precision=%d) instead of row %d",
        total_row + 2,
        final_gw,
        plus2_precision,
        total_row,
    )
    return final_gw, plus2_precision


# ---------------------------------------------------------------------------
# FR-017 — Total packets extraction
# ---------------------------------------------------------------------------


def _extract_total_packets(
    sheet: Worksheet,
    total_row: int,
    column_map: ColumnMapping,
) -> int | None:
    """Extract total_packets using multi-priority search (FR-017).

    Args:
        sheet: Packing worksheet.
        total_row: 1-based total row number.
        column_map: Column index mapping.

    Returns:
        Total packets as integer, or None if not found.
    """
    nw_col = column_map.field_map["nw"]
    max_col = max(nw_col + 2, 11)

    # Priority 1 — 件数/件數 label
    result = _priority1_jian_shu(sheet, total_row, max_col)
    if result is not None:
        logger.info("Total packets found via Priority 1 (件数 label): %d", result)
        return result

    # Priority 2 — PLT.G indicator
    result = _priority2_plt_g(sheet, total_row, max_col)
    if result is not None:
        logger.info("Total packets found via Priority 2 (PLT.G indicator): %d", result)
        return result

    # Priority 3 — Below-total patterns
    result = _priority3_below_total(sheet, total_row, max_col)
    if result is not None:
        logger.info("Total packets found via Priority 3 (below-total pattern): %d", result)
        return result

    # All priorities failed — return None; batch.py creates ATT_002 ProcessingError
    logger.debug("Total packets not found after all 3 priority searches at total row %d", total_row)
    return None


def _validate_packets(value: int | None) -> int | None:
    """Validate that a packets value is a positive integer in range 1-1000.

    Args:
        value: Candidate packets value.

    Returns:
        The value if valid, None otherwise.
    """
    if value is None:
        return None
    if not isinstance(value, int):
        return None
    if value < _PACKETS_MIN or value > _PACKETS_MAX:
        return None
    return value


def _parse_int_from_match(text: str) -> int | None:
    """Try to parse an integer from a string, stripping unit suffixes first.

    Args:
        text: String potentially containing a number with unit suffix.

    Returns:
        Integer value or None if parsing fails.
    """
    stripped = strip_unit_suffix(str(text).strip())
    try:
        float_val = float(stripped)
        int_val = int(float_val)
        # Reason: Only accept values that are essentially integers
        if abs(float_val - int_val) < 0.01:
            return int_val
    except (ValueError, TypeError):
        pass
    return None


def _priority1_jian_shu(sheet: Worksheet, total_row: int, max_col: int) -> int | None:
    """Priority 1: Search for 件数/件數 label and extract adjacent value.

    Args:
        sheet: Packing worksheet.
        total_row: 1-based total row number.
        max_col: Maximum column to search.

    Returns:
        Packets count or None.
    """
    for row in range(total_row + 1, total_row + 4):
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                continue
            cell_str = str(cell_value)
            if not _JIAN_SHU_RE.search(cell_str):
                continue

            # Reason: Check if value is embedded in the label cell itself
            # e.g. "件数: 55" or "件数55"
            embedded = re.search(r"(\d+)", cell_str)
            # Check adjacent right cells first (up to +3 columns from label)
            for adj_offset in range(1, 4):
                adj_col = col + adj_offset
                if adj_col > max_col:
                    break
                adj_value = sheet.cell(row=row, column=adj_col).value
                if adj_value is not None:
                    parsed = _parse_int_from_match(str(adj_value))
                    result = _validate_packets(parsed)
                    if result is not None:
                        return result

            # Fall back to embedded numeric in label cell
            if embedded:
                parsed = int(embedded.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

    return None


def _priority2_plt_g(sheet: Worksheet, total_row: int, max_col: int) -> int | None:
    """Priority 2: Search for PLT.G indicator above total row.

    Args:
        sheet: Packing worksheet.
        total_row: 1-based total row number.
        max_col: Maximum column to search.

    Returns:
        Packets count or None.
    """
    for row in [total_row - 1, total_row - 2]:
        if row < 1:
            continue
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip()

            # Format 1: Number-before-PLT — e.g. "7PLT.G", "7 PLT"
            match = _PLT_NUMBER_BEFORE_RE.search(cell_str)
            if match:
                parsed = int(match.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

            # Format 2: PLT-before-number — e.g. "PLT" with number in next cell
            if _PLT_NUMBER_AFTER_RE.search(cell_str) and not _PLT_NUMBER_BEFORE_RE.search(cell_str):
                # Check cell immediately to the right
                if col + 1 <= max_col:
                    right_value = sheet.cell(row=row, column=col + 1).value
                    if right_value is not None:
                        parsed = _parse_int_from_match(str(right_value))
                        result = _validate_packets(parsed)
                        if result is not None:
                            return result

    return None


def _priority3_below_total(sheet: Worksheet, total_row: int, max_col: int) -> int | None:
    """Priority 3: Search below-total rows for various patterns.

    Patterns searched in order per row:
    (a) Total-with-breakdown: "348（256胶框+92纸箱）" -> 348
    (b) Unit-suffix patterns: "7托", "30箱", "55 CTNS"
    (c) Embedded Chinese: "共7托"
    (d) Pallet range: "PLT#1(1~34)" -> 34

    Args:
        sheet: Packing worksheet.
        total_row: 1-based total row number.
        max_col: Maximum column to search.

    Returns:
        Packets count or None.
    """
    for row in range(total_row + 1, total_row + 4):
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip()
            if cell_str == "":
                continue

            # (a) Total-with-breakdown: leading number before （ or (
            # Reason: Pallet vs box priority — check for "共N托" first when both appear
            pallet_match = _PALLET_VS_BOX_RE.search(cell_str)
            if pallet_match:
                parsed = int(pallet_match.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

            match_a = _BREAKDOWN_RE.search(cell_str)
            if match_a:
                parsed = int(match_a.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

            # (b) Unit-suffix patterns: 托/箱/件/CTNS
            match_b = _UNIT_SUFFIX_PACKETS_RE.search(cell_str)
            if match_b:
                parsed = int(match_b.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

            # (c) Embedded Chinese: 共N托
            match_c = _EMBEDDED_CHINESE_RE.search(cell_str)
            if match_c:
                parsed = int(match_c.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

            # (d) Pallet range: PLT#1(1~34)
            match_d = _PALLET_RANGE_RE.search(cell_str)
            if match_d:
                parsed = int(match_d.group(1))
                result = _validate_packets(parsed)
                if result is not None:
                    return result

    return None
