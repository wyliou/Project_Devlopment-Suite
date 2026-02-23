"""sheet_detect — FR-004, FR-005, FR-006: Invoice and Packing sheet identification."""

import logging

from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from .errors import ErrorCode, ProcessingError
from .models import AppConfig

logger = logging.getLogger(__name__)


def detect_sheets(workbook: Workbook, config: AppConfig) -> tuple[Worksheet, Worksheet]:
    """Identify Invoice and Packing sheets by matching sheet names against regex patterns.

    Scans all worksheets in the workbook, stripping whitespace from sheet names,
    and matches against patterns from config. Pattern matching is case-insensitive
    (re.IGNORECASE). First matching sheet wins for each type.

    Args:
        workbook: The openpyxl Workbook to scan.
        config: Application configuration with compiled sheet patterns.

    Returns:
        A tuple of (invoice_sheet, packing_sheet) as Worksheet objects.

    Raises:
        ProcessingError: ERR_012 if no invoice sheet found.
        ProcessingError: ERR_013 if no packing sheet found.
    """
    invoice_sheet: Worksheet | None = None
    packing_sheet: Worksheet | None = None

    # Scan all sheets in the workbook
    for worksheet in workbook.sheetnames:
        sheet = workbook[worksheet]

        # Strip whitespace from sheet name for pattern matching
        normalized_name = worksheet.strip()

        # Check invoice patterns first
        if invoice_sheet is None:
            for pattern in config.invoice_sheet_patterns:
                if pattern.match(normalized_name):
                    invoice_sheet = sheet
                    logger.debug(
                        "Detected invoice sheet: '%s' matched pattern %s",
                        worksheet,
                        pattern.pattern,
                    )
                    break

        # Check packing patterns
        if packing_sheet is None:
            for pattern in config.packing_sheet_patterns:
                if pattern.match(normalized_name):
                    packing_sheet = sheet
                    logger.debug(
                        "Detected packing sheet: '%s' matched pattern %s",
                        worksheet,
                        pattern.pattern,
                    )
                    break

        # Early exit if both found
        if invoice_sheet is not None and packing_sheet is not None:
            break

    # Validate that both sheets were found
    if invoice_sheet is None:
        logger.error("ERR_012: Invoice sheet not found in workbook")
        raise ProcessingError(
            code=ErrorCode.ERR_012,
            message="Invoice sheet not found in workbook",
            context={"sheet_names": workbook.sheetnames},
        )

    if packing_sheet is None:
        logger.error("ERR_013: Packing sheet not found in workbook")
        raise ProcessingError(
            code=ErrorCode.ERR_013,
            message="Packing sheet not found in workbook",
            context={"sheet_names": workbook.sheetnames},
        )

    return invoice_sheet, packing_sheet
