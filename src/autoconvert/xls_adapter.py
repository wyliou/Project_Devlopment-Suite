"""xls_adapter — FR-003: Convert .xls files to in-memory openpyxl Workbook via xlrd."""

import logging
from datetime import datetime
from pathlib import Path

import xlrd
from openpyxl import Workbook
from xlrd import (
    XL_CELL_BOOLEAN,
    XL_CELL_DATE,
    XL_CELL_EMPTY,
    XL_CELL_ERROR,
    XL_CELL_NUMBER,
    XL_CELL_TEXT,
    open_workbook,
    xldate_as_datetime,
)

logger = logging.getLogger(__name__)


def convert_xls_to_xlsx(filepath: Path) -> Workbook:
    """Convert a legacy .xls file to an in-memory openpyxl Workbook.

    Reads the .xls file using xlrd with formatting_info=True (required for
    merged-cell information), then reconstructs every sheet, cell value, and
    merge range inside a freshly created openpyxl Workbook.  The original
    .xls file is never modified and no temporary files are written.

    Args:
        filepath: Absolute or relative path to the .xls file to convert.

    Returns:
        An in-memory openpyxl Workbook containing all sheets from the source
        file with cell values and merge ranges preserved.  Number formats are
        not preserved — all cells carry the default "General" format.

    Raises:
        Exception: Any exception raised by xlrd (e.g. xlrd.XLRDError for a
            corrupted file, FileNotFoundError for a missing path) propagates
            unchanged so that the caller (batch.py) can map it to ERR_011.
    """
    # Open the .xls workbook.  formatting_info=True is required for xlrd 2.x
    # to populate sheet.merged_cells; without it the list is always empty.
    book: xlrd.Book = open_workbook(str(filepath), formatting_info=True)
    datemode: int = book.datemode

    wb = Workbook()
    # Remove the default empty sheet that openpyxl creates.
    wb.remove(wb.active)  # type: ignore[arg-type]

    sheet_count = book.nsheets
    logger.debug("Converting %d sheet(s) from %s", sheet_count, filepath.name)

    for sheet_index in range(sheet_count):
        xlrd_sheet = book.sheet_by_index(sheet_index)
        ws = wb.create_sheet(title=xlrd_sheet.name)

        # --- Transfer cell values -------------------------------------------
        for xlrd_row in range(xlrd_sheet.nrows):
            for xlrd_col in range(xlrd_sheet.ncols):
                cell = xlrd_sheet.cell(xlrd_row, xlrd_col)

                # Convert xlrd 0-based indices to openpyxl 1-based indices.
                openpyxl_row = xlrd_row + 1
                openpyxl_col = xlrd_col + 1

                value = _convert_cell_value(cell, datemode)
                ws.cell(row=openpyxl_row, column=openpyxl_col, value=value)

        # --- Transfer merged cell ranges ------------------------------------
        merge_count = len(xlrd_sheet.merged_cells)
        logger.debug(
            "Sheet '%s': %d merge range(s) to transfer",
            xlrd_sheet.name,
            merge_count,
        )

        for rlo, rhi, clo, chi in xlrd_sheet.merged_cells:
            # xlrd tuple: (rlo, rhi, clo, chi)
            #   rlo, clo — 0-based inclusive start
            #   rhi, chi — 0-based exclusive end
            # openpyxl merge_cells: start_row/start_column are 1-based
            #                       end_row/end_column are 1-based inclusive
            # Conversion:
            #   start_row   = rlo + 1
            #   end_row     = rhi      (rhi is exclusive in xlrd, so the last
            #                           included 0-based row is rhi-1, which
            #                           maps to 1-based rhi)
            #   start_column = clo + 1
            #   end_column  = chi      (same logic as end_row)

            # Anchor cell value must be written BEFORE merge_cells() because
            # openpyxl clears non-anchor cells during the merge operation.
            anchor_value = _convert_cell_value(xlrd_sheet.cell(rlo, clo), datemode)
            ws.cell(row=rlo + 1, column=clo + 1, value=anchor_value)

            ws.merge_cells(
                start_row=rlo + 1,
                start_column=clo + 1,
                end_row=rhi,
                end_column=chi,
            )

    logger.debug("Conversion complete: %d sheet(s) from '%s'", sheet_count, filepath.name)
    return wb


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _convert_cell_value(cell: xlrd.sheet.Cell, datemode: int) -> str | float | bool | datetime | None:
    """Map an xlrd Cell to the Python value that openpyxl should store.

    Args:
        cell: The xlrd Cell object to convert.
        datemode: The datemode of the xlrd Book (0 = 1900-based, 1 = 1904-based).

    Returns:
        A Python value suitable for assignment to an openpyxl cell:
        - XL_CELL_TEXT   -> str
        - XL_CELL_NUMBER -> float  (xlrd always stores numbers as float)
        - XL_CELL_DATE   -> datetime
        - XL_CELL_BOOLEAN -> bool
        - XL_CELL_ERROR  -> None   (treat as empty, mirrors data_only=True)
        - XL_CELL_EMPTY  -> None
    """
    ctype = cell.ctype

    if ctype == XL_CELL_EMPTY:
        return None
    if ctype == XL_CELL_TEXT:
        return str(cell.value)
    if ctype == XL_CELL_NUMBER:
        # xlrd always returns numbers as float; downstream modules handle
        # float-to-Decimal conversion (safe_decimal in utils.py).
        return float(cell.value)
    if ctype == XL_CELL_DATE:
        # Reason: datemode is workbook-level (0 for 1900, 1 for 1904) and
        # must come from the Book object, not hardcoded, to correctly parse
        # dates from both Excel date systems.
        return xldate_as_datetime(cell.value, datemode)
    if ctype == XL_CELL_BOOLEAN:
        # xlrd returns 0 or 1 for booleans; convert to Python bool.
        return bool(cell.value)
    if ctype == XL_CELL_ERROR:
        # Error cells (e.g. #DIV/0!, #VALUE!) are stored as None to mirror
        # openpyxl data_only=True semantics for formula-error cells.
        return None

    # Fallback for any unknown future cell type — treat as None.
    return None
