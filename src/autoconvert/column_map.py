"""column_map — FR-007, FR-008, FR-009: Header detection, column mapping, and invoice number extraction."""

import logging
import re

from openpyxl.worksheet.worksheet import Worksheet

from .errors import ErrorCode, ProcessingError
from .models import AppConfig, ColumnMapping, FieldPattern
from .utils import HEADER_KEYWORDS, normalize_header

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_SCAN_ROW_START = 7
_SCAN_ROW_END = 30
_SCAN_COL_END = 13

_INVOICE_THRESHOLD = 7
_PACKING_THRESHOLD = 4

_METADATA_MARKERS: tuple[str, ...] = ("tel:", "fax:", "cust id:", "contact:", "address:")

# Reason: Detect numeric/data-like cells for tier classification. Alphanumeric
# codes must contain at least one digit AND at least one letter to avoid
# matching normal header words like "brand", "model", "price".
_NUMERIC_DATA_RE = re.compile(
    r"^[\d]+\.?[\d]*$"  # Pure numbers / decimals
    r"|^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9]{4,}$"  # Alphanumeric codes (4+ chars, mixed)
)

# Currency code patterns for data-row fallback
_CURRENCY_CODE_RE = re.compile(r"^(USD|CNY|EUR|RMB|JPY|GBP|TWD|HKD|SGD|KRW)$", re.IGNORECASE)

# Reason: Strip "INV#", "INV." prefixes and "NO." prefix from extracted values.
# Also strip bare "INV" when directly followed by a digit (e.g. "INV2025-001" -> "2025-001")
# but NOT when followed by a dash (e.g. "INV-001" stays as "INV-001").
_INV_PREFIX_RE = re.compile(r"^(?:INV[#.]\s*|INV(?=\d)|NO\.?\s*)", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def detect_header_row(sheet: Worksheet, sheet_type: str, config: AppConfig) -> int:
    """Detect the header row by scanning rows 7-30 with three-tier priority.

    Args:
        sheet: Worksheet to scan (already unmerged).
        sheet_type: "invoice" or "packing".
        config: Application configuration.

    Returns:
        1-based row number of the detected header row.

    Raises:
        ProcessingError: ERR_014 if no row meets the cell-count threshold.
    """
    threshold = _INVOICE_THRESHOLD if sheet_type == "invoice" else _PACKING_THRESHOLD

    # Reason: We collect all qualifying rows with their tier, then pick
    # the best one (lowest tier, then earliest row on tie).
    candidates: list[tuple[int, int]] = []  # (tier, row_number)

    for row_idx in range(_SCAN_ROW_START, _SCAN_ROW_END + 1):
        cell_values: list[str] = []
        for col_idx in range(1, _SCAN_COL_END + 1):
            raw = sheet.cell(row=row_idx, column=col_idx).value
            if raw is None:
                continue
            text = str(raw).strip()
            if not text:
                continue
            # Reason: Exclude "Unnamed:" prefix cells (library artifacts from pandas-exported sheets).
            if text.startswith("Unnamed:"):
                continue
            cell_values.append(text)

        non_empty_count = len(cell_values)
        logger.debug(
            "Row %d: %d non-empty cells (threshold=%d)",
            row_idx,
            non_empty_count,
            threshold,
        )

        if non_empty_count < threshold:
            continue

        # Classify tier
        tier = _classify_tier(cell_values)
        logger.debug("Row %d classified as Tier %d", row_idx, tier)
        candidates.append((tier, row_idx))

    if not candidates:
        raise ProcessingError(
            code=ErrorCode.ERR_014,
            message=f"No header row found for {sheet_type} sheet (rows {_SCAN_ROW_START}-{_SCAN_ROW_END})",
            context={"sheet_type": sheet_type},
        )

    # Reason: Sort by (tier, row_number) — lowest tier wins, then earliest row.
    candidates.sort(key=lambda t: (t[0], t[1]))
    best_tier, best_row = candidates[0]
    logger.debug("Selected header row %d (Tier %d) for %s sheet", best_row, best_tier, sheet_type)
    return best_row


def map_columns(
    sheet: Worksheet,
    header_row: int,
    sheet_type: str,
    config: AppConfig,
) -> ColumnMapping:
    """Map columns via regex with sub-header and currency fallback.

    Args:
        sheet: Worksheet (already unmerged).
        header_row: 1-based header row from detect_header_row().
        sheet_type: "invoice" or "packing".
        config: Application configuration.

    Returns:
        ColumnMapping with field_map, header_row, effective_header_row.

    Raises:
        ProcessingError: ERR_020 listing all missing required column names.
    """
    columns_config = config.invoice_columns if sheet_type == "invoice" else config.packing_columns

    # Step 1: Scan the primary header row
    field_map = _scan_row_for_fields(sheet, header_row, columns_config)
    effective_header_row = header_row
    logger.debug("Primary header scan at row %d: mapped %s", header_row, list(field_map.keys()))

    # Step 2: Sub-header fallback — scan header_row+1 for unmapped REQUIRED fields
    required_fields = {name for name, fp in columns_config.items() if fp.required}
    unmapped_required = required_fields - set(field_map.keys())

    if unmapped_required:
        sub_row = header_row + 1
        logger.debug(
            "Sub-header scan at row %d for unmapped required: %s",
            sub_row,
            unmapped_required,
        )
        # Guard: check sub-header row is not data-like
        if not _is_data_like_row(sheet, sub_row):
            # Only try to match the unmapped required fields
            sub_columns = {name: columns_config[name] for name in unmapped_required if name in columns_config}
            sub_map = _scan_row_for_fields(sheet, sub_row, sub_columns)
            if sub_map:
                field_map.update(sub_map)
                effective_header_row = sub_row
                logger.debug(
                    "Sub-header matched %s; effective_header_row advanced to %d",
                    list(sub_map.keys()),
                    effective_header_row,
                )
        else:
            logger.debug("Sub-header row %d rejected (data-like)", sub_row)

    # Step 3: Currency data-row fallback (invoice only)
    if sheet_type == "invoice" and "currency" not in field_map:
        _currency_data_row_fallback(sheet, header_row, field_map)

    # Step 4: Check all required fields are mapped (collect-then-report)
    still_unmapped = sorted(required_fields - set(field_map.keys()))
    if still_unmapped:
        raise ProcessingError(
            code=ErrorCode.ERR_020,
            message=f"Required columns missing for {sheet_type}: {', '.join(still_unmapped)}",
            context={"sheet_type": sheet_type, "missing_fields": still_unmapped},
        )

    return ColumnMapping(
        sheet_type=sheet_type,
        field_map=field_map,
        header_row=header_row,
        effective_header_row=effective_header_row,
    )


def extract_inv_no_from_header(sheet: Worksheet, config: AppConfig) -> str | None:
    """Scan invoice sheet rows 1-15 for invoice number via capture/label patterns.

    Args:
        sheet: Invoice worksheet (already unmerged).
        config: Application configuration with inv_no patterns.

    Returns:
        Cleaned invoice number string, or None if not found.
        Does NOT raise ERR_021 -- caller (batch.py) decides if error applies.
    """
    max_col = sheet.max_column or 20
    # Reason: Limit scan columns to a reasonable upper bound to avoid
    # scanning phantom columns in sheets with excessive max_column.
    scan_max_col = min(max_col, 20)

    for row_idx in range(1, 16):
        for col_idx in range(1, scan_max_col + 1):
            raw = sheet.cell(row=row_idx, column=col_idx).value
            if raw is None:
                continue
            text = str(raw).strip()
            if not text:
                continue

            # Try capture-group patterns (value in same cell)
            for pattern in config.inv_no_patterns:
                match = pattern.search(text)
                if match:
                    # Use the first capture group if present, else whole match
                    value = match.group(1) if match.lastindex else match.group(0)
                    cleaned = _clean_inv_no(value)
                    if cleaned and not _is_excluded_inv_no(cleaned, config):
                        logger.debug(
                            "inv_no capture match at row %d, col %d: %r -> %r",
                            row_idx,
                            col_idx,
                            text,
                            cleaned,
                        )
                        return cleaned

            # Try label patterns (value in adjacent cell)
            for pattern in config.inv_no_label_patterns:
                if pattern.search(text):
                    logger.debug(
                        "inv_no label match at row %d, col %d: %r",
                        row_idx,
                        col_idx,
                        text,
                    )
                    # Search right: up to +3 columns
                    result = _search_adjacent_right(sheet, row_idx, col_idx, scan_max_col, config)
                    if result:
                        return result

                    # Search below: row+1 AND row+2
                    result = _search_adjacent_below(sheet, row_idx, col_idx, config)
                    if result:
                        return result

    logger.debug("No invoice number found in header area (rows 1-15)")
    return None


# ---------------------------------------------------------------------------
# Private helpers
# ---------------------------------------------------------------------------


def _classify_tier(cell_values: list[str]) -> int:
    """Classify a row into tier 0, 1, or 2 based on content.

    Args:
        cell_values: Non-empty, stripped cell text values from the row.

    Returns:
        Tier number: 0 (best), 1, or 2 (worst).
    """
    lower_values = [v.lower() for v in cell_values]

    # Check for metadata markers -> Tier 2
    for val in lower_values:
        for marker in _METADATA_MARKERS:
            if marker in val:
                return 2

    # Count numeric/data-like cells
    numeric_count = sum(1 for v in cell_values if _NUMERIC_DATA_RE.match(v))
    if numeric_count >= 3:
        return 2

    # Check for header keywords -> Tier 0
    has_keyword = False
    for val in lower_values:
        for kw in HEADER_KEYWORDS:
            if kw in val:
                has_keyword = True
                break
        if has_keyword:
            break

    if has_keyword and numeric_count < 2:
        return 0

    return 1


def _scan_row_for_fields(
    sheet: Worksheet,
    row_idx: int,
    columns_config: dict[str, FieldPattern],
) -> dict[str, int]:
    """Scan a single row and match cells against field regex patterns.

    Args:
        sheet: Worksheet to scan.
        row_idx: 1-based row number.
        columns_config: Dict of field_name -> FieldPattern to match against.

    Returns:
        Dict mapping matched field_name to 1-based column index. First match per field wins.
    """
    field_map: dict[str, int] = {}
    max_col = sheet.max_column or _SCAN_COL_END
    # Reason: Limit scan to avoid phantom columns.
    scan_cols = min(max_col, 50)

    for col_idx in range(1, scan_cols + 1):
        raw = sheet.cell(row=row_idx, column=col_idx).value
        if raw is None:
            continue
        text = normalize_header(str(raw))
        if not text:
            continue

        for field_name, field_pattern in columns_config.items():
            if field_name in field_map:
                continue  # Already mapped
            for pattern_str in field_pattern.patterns:
                compiled = re.compile(pattern_str, re.IGNORECASE)
                if compiled.search(text):
                    field_map[field_name] = col_idx
                    logger.debug(
                        "Matched field %r at row %d, col %d: header=%r, pattern=%r",
                        field_name,
                        row_idx,
                        col_idx,
                        text,
                        pattern_str,
                    )
                    break

    return field_map


def _is_data_like_row(sheet: Worksheet, row_idx: int) -> bool:
    """Check if a row is data-like (3+ numeric/code cells).

    Args:
        sheet: Worksheet to check.
        row_idx: 1-based row number.

    Returns:
        True if the row has 3+ numeric/data-like cells.
    """
    max_col = sheet.max_column or _SCAN_COL_END
    scan_cols = min(max_col, 50)
    numeric_count = 0

    for col_idx in range(1, scan_cols + 1):
        raw = sheet.cell(row=row_idx, column=col_idx).value
        if raw is None:
            continue
        text = str(raw).strip()
        if text and _NUMERIC_DATA_RE.match(text):
            numeric_count += 1
            if numeric_count >= 3:
                return True

    return False


def _currency_data_row_fallback(
    sheet: Worksheet,
    header_row: int,
    field_map: dict[str, int],
) -> None:
    """Scan data rows for embedded currency codes when no currency column detected.

    If a currency code is found in a column already mapped to price or amount,
    shift that field's mapping to col+1 (actual numeric value is one column right).

    Args:
        sheet: Worksheet to scan.
        header_row: 1-based header row number.
        field_map: Current field_map (mutated in place).
    """
    max_col = sheet.max_column or _SCAN_COL_END
    scan_cols = min(max_col, 50)
    currency_col: int | None = None

    for data_row in range(header_row + 1, header_row + 5):
        for col_idx in range(1, scan_cols + 1):
            raw = sheet.cell(row=data_row, column=col_idx).value
            if raw is None:
                continue
            text = str(raw).strip()
            if not _CURRENCY_CODE_RE.match(text):
                continue

            logger.debug(
                "Currency code %r found at row %d, col %d",
                text,
                data_row,
                col_idx,
            )

            # Check if this column is mapped to price or amount
            fields_to_shift: list[str] = []
            for field_name in ("price", "amount"):
                if field_name in field_map and field_map[field_name] == col_idx:
                    fields_to_shift.append(field_name)

            # Shift affected fields +1 column
            for field_name in fields_to_shift:
                old_col = field_map[field_name]
                field_map[field_name] = old_col + 1
                logger.debug(
                    "Currency fallback: shifted %s from col %d to col %d",
                    field_name,
                    old_col,
                    old_col + 1,
                )

            # Record the currency column if we shifted anything or if
            # the column is adjacent to a mapped numeric field
            if fields_to_shift and currency_col is None:
                currency_col = col_idx

        # Reason: Once we find currency in any data row, we have
        # the information we need. Continue scanning remaining rows
        # to catch all affected fields (e.g., two currency columns).

    if currency_col is not None:
        field_map["currency"] = currency_col
        logger.debug("Currency column set to %d via data-row fallback", currency_col)


def _search_adjacent_right(
    sheet: Worksheet,
    row_idx: int,
    col_idx: int,
    max_col: int,
    config: AppConfig,
) -> str | None:
    """Search right of a label cell for invoice number (up to +3 columns).

    Args:
        sheet: Worksheet to search.
        row_idx: Row of the label cell.
        col_idx: Column of the label cell.
        max_col: Maximum column to scan.
        config: AppConfig with exclude patterns.

    Returns:
        Cleaned invoice number, or None if not found.
    """
    for offset in range(1, 4):
        adj_col = col_idx + offset
        if adj_col > max_col:
            break
        raw = sheet.cell(row=row_idx, column=adj_col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        cleaned = _clean_inv_no(text)
        if cleaned and not _is_excluded_inv_no(cleaned, config):
            logger.debug(
                "inv_no found right-adjacent at row %d, col %d: %r",
                row_idx,
                adj_col,
                cleaned,
            )
            return cleaned
    return None


def _search_adjacent_below(
    sheet: Worksheet,
    row_idx: int,
    col_idx: int,
    config: AppConfig,
) -> str | None:
    """Search below a label cell for invoice number (row+1 and row+2).

    Args:
        sheet: Worksheet to search.
        row_idx: Row of the label cell.
        col_idx: Column of the label cell.
        config: AppConfig with exclude patterns.

    Returns:
        Cleaned invoice number, or None if not found.
    """
    # Reason: Must check row+1 AND row+2 because some vendors place
    # a date at row+1 and the actual invoice number at row+2.
    for row_offset in range(1, 3):
        below_row = row_idx + row_offset
        raw = sheet.cell(row=below_row, column=col_idx).value
        if raw is None:
            continue
        text = str(raw).strip()
        if not text:
            continue
        cleaned = _clean_inv_no(text)
        if cleaned and not _is_excluded_inv_no(cleaned, config):
            logger.debug(
                "inv_no found below-adjacent at row %d, col %d: %r",
                below_row,
                col_idx,
                cleaned,
            )
            return cleaned
    return None


def _clean_inv_no(value: str) -> str:
    """Remove INV# and NO. prefixes from extracted invoice number values.

    Args:
        value: Raw extracted invoice number string.

    Returns:
        Cleaned invoice number string.
    """
    cleaned = _INV_PREFIX_RE.sub("", value).strip()
    return cleaned


def _is_excluded_inv_no(value: str, config: AppConfig) -> bool:
    """Check if an extracted invoice number matches any exclude pattern.

    Args:
        value: Cleaned invoice number candidate.
        config: AppConfig with inv_no_exclude_patterns.

    Returns:
        True if the value should be excluded (false positive).
    """
    for pattern in config.inv_no_exclude_patterns:
        if pattern.search(value):
            logger.debug("inv_no value %r excluded by pattern %r", value, pattern.pattern)
            return True
    return False
