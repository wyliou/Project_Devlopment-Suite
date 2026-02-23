"""output — FR-029, FR-030: Template population and output file generation."""

import logging
from pathlib import Path

from openpyxl import load_workbook

from .errors import ErrorCode, ProcessingError
from .models import AppConfig, InvoiceItem, PackingTotals

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column index constants (1-based, A=1 through AN=40)
# ---------------------------------------------------------------------------
_COL_A = 1  # 企业料号 (part_no)
_COL_B = 2  # 采购订单号 (po_no)
_COL_C = 3  # 征免方式 — fixed "3"
_COL_D = 4  # 币制 (currency)
_COL_E = 5  # 申报数量 (qty)
_COL_F = 6  # 申报单价 (price)
_COL_G = 7  # 申报总价 (amount)
_COL_H = 8  # 原产国 (coo)
# Cols I(9), J(10), K(11) — reserved/empty
_COL_L = 12  # 报关单商品序号 (serial)
_COL_M = 13  # 净重 (allocated_weight)
_COL_N = 14  # 发票号码 (inv_no)
# Col O(15) — reserved/empty
_COL_P = 16  # 毛重 (total_gw) — row 5 ONLY
# Col Q(17) — reserved/empty
_COL_R = 18  # 境内目的地代码 — fixed "32052"
_COL_S = 19  # 行政区划代码 — fixed "320506"
_COL_T = 20  # 最终目的国 — fixed "142"
# Cols U(21)–AJ(36) — reserved/empty
_COL_AK = 37  # 件数 (total_packets) — row 5 ONLY
_COL_AL = 38  # 品牌 (brand)
_COL_AM = 39  # 品牌类型 (brand_type)
_COL_AN = 40  # 型号 (model_no)

# Row offset: template rows 1–4 are metadata; data starts at row 5
_DATA_START_ROW = 5

# Fixed values written to every data row
_FIXED_C = "3"
_FIXED_R = "32052"
_FIXED_S = "320506"
_FIXED_T = "142"

# Template sheet name (Chinese characters are part of the name)
_SHEET_NAME = "工作表1"


def write_template(
    invoice_items: list[InvoiceItem],
    packing_totals: PackingTotals,
    config: AppConfig,
    output_path: Path,
) -> None:
    """Populate 40-column template and write to output_path.

    Loads the template from config.template_path fresh for each call (no
    caching) to avoid state contamination between files in a batch.  Rows 1–4
    are preserved as-is from the template.  One data row is written per item in
    invoice_items, starting at row 5.

    Args:
        invoice_items: Fully processed invoice items with allocated_weight populated.
        packing_totals: Packing totals for total_gw (col P, row 5) and
                        total_packets (col AK, row 5).
        config: Application configuration with template_path.
        output_path: Full path to write output file (data/finished/{stem}_template.xlsx).

    Returns:
        None.

    Raises:
        ProcessingError: ERR_051 if template cannot be loaded at write time;
                         ERR_052 if the output file cannot be written.
    """
    # ------------------------------------------------------------------
    # Load the template workbook (ERR_051 on failure)
    # ------------------------------------------------------------------
    try:
        wb = load_workbook(config.template_path)
    except Exception as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_051,
            message=f"Failed to load output template '{config.template_path}': {exc}",
            context={"template_path": str(config.template_path)},
        ) from exc

    ws = wb[_SHEET_NAME]

    # ------------------------------------------------------------------
    # Write one data row per InvoiceItem (rows 1–4 preserved)
    # ------------------------------------------------------------------
    for item_index, item in enumerate(invoice_items):
        row = _DATA_START_ROW + item_index
        _write_item_row(ws, row, item)

    # ------------------------------------------------------------------
    # Row-5-only fields: total_gw (col P) and total_packets (col AK)
    # ------------------------------------------------------------------
    if invoice_items:
        # total_gw is always present on PackingTotals
        ws.cell(row=_DATA_START_ROW, column=_COL_P).value = packing_totals.total_gw

        # total_packets may be None (ATT_002 case) — write nothing if absent
        if packing_totals.total_packets is not None:
            ws.cell(row=_DATA_START_ROW, column=_COL_AK).value = packing_totals.total_packets

    # ------------------------------------------------------------------
    # Save the workbook (ERR_052 on failure)
    # ------------------------------------------------------------------
    try:
        wb.save(output_path)
    except Exception as exc:
        raise ProcessingError(
            code=ErrorCode.ERR_052,
            message=f"Failed to write output file '{output_path}': {exc}",
            context={"output_path": str(output_path)},
        ) from exc

    logger.info("Output successfully written to: %s", output_path.name)


def _write_item_row(ws, row: int, item: InvoiceItem) -> None:  # type: ignore[no-untyped-def]
    """Write a single invoice item into the given worksheet row.

    Columns I, J, K, O, Q, and U–AJ are intentionally left empty (not written).
    Columns P and AK are written separately (row-5-only logic) and must NOT be
    written here.

    Args:
        ws: openpyxl Worksheet (output sheet).
        row: 1-based row number for this item.
        item: The InvoiceItem to write.
    """
    # Column A — part_no
    ws.cell(row=row, column=_COL_A).value = item.part_no

    # Column B — po_no
    ws.cell(row=row, column=_COL_B).value = item.po_no

    # Column C — fixed "3" (征免方式)
    ws.cell(row=row, column=_COL_C).value = _FIXED_C

    # Column D — currency (raw string; ATT_003 passthrough requires no special logic)
    ws.cell(row=row, column=_COL_D).value = item.currency

    # Column E — qty
    ws.cell(row=row, column=_COL_E).value = item.qty

    # Column F — price
    ws.cell(row=row, column=_COL_F).value = item.price

    # Column G — amount
    ws.cell(row=row, column=_COL_G).value = item.amount

    # Column H — coo (raw string; ATT_004 passthrough requires no special logic)
    ws.cell(row=row, column=_COL_H).value = item.coo

    # Columns I (9), J (10), K (11) — reserved; intentionally left empty

    # Column L — serial (报关单商品序号); format 0.00000_ set in template
    ws.cell(row=row, column=_COL_L).value = item.serial

    # Column M — allocated_weight (净重); format 0.00000_ set in template
    ws.cell(row=row, column=_COL_M).value = item.allocated_weight

    # Column N — inv_no
    ws.cell(row=row, column=_COL_N).value = item.inv_no

    # Column O (15) — reserved; intentionally left empty

    # Column P (16) — total_gw; written ONLY at row 5 (handled by caller)

    # Column Q (17) — reserved; intentionally left empty

    # Column R — fixed "32052"
    ws.cell(row=row, column=_COL_R).value = _FIXED_R

    # Column S — fixed "320506"
    ws.cell(row=row, column=_COL_S).value = _FIXED_S

    # Column T — fixed "142"
    ws.cell(row=row, column=_COL_T).value = _FIXED_T

    # Columns U (21) through AJ (36) — reserved; intentionally left empty

    # Column AK (37) — total_packets; written ONLY at row 5 (handled by caller)

    # Column AL — brand
    ws.cell(row=row, column=_COL_AL).value = item.brand

    # Column AM — brand_type
    ws.cell(row=row, column=_COL_AM).value = item.brand_type

    # Column AN — model_no (PRD name: 型号; Python attr: model_no)
    ws.cell(row=row, column=_COL_AN).value = item.model_no
