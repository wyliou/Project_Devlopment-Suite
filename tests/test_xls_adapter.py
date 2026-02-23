"""Tests for xls_adapter.convert_xls_to_xlsx.

All tests that can use the real corpus file (茂綸股份有限公司.xls) do so.
Tests for cell types not present in the corpus (XL_CELL_DATE, XL_CELL_BOOLEAN,
XL_CELL_ERROR) use xlrd's public API surface through unittest.mock targeting
the private helper _convert_cell_value.
"""

from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
import xlrd
from openpyxl import Workbook

from autoconvert.xls_adapter import _convert_cell_value, convert_xls_to_xlsx

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

CORPUS_XLS = Path(__file__).parent.parent / "data" / "茂綸股份有限公司.xls"


@pytest.fixture
def corpus_path() -> Path:
    """Return the path to the real .xls corpus file."""
    if not CORPUS_XLS.exists():
        pytest.skip(f"Corpus file not found: {CORPUS_XLS}")
    return CORPUS_XLS


# ---------------------------------------------------------------------------
# Happy-path tests using the real corpus file
# ---------------------------------------------------------------------------


def test_convert_xls_basic_cell_values(corpus_path: Path) -> None:
    """Test convert_xls_to_xlsx when given the corpus file returns a Workbook
    with the expected sheet names and known text cell values.
    """
    wb = convert_xls_to_xlsx(corpus_path)

    assert isinstance(wb, Workbook)
    assert "INVOICE" in wb.sheetnames
    assert "PACKING" in wb.sheetnames

    # The first cell (row 0, col 0 in xlrd → row 1, col 1 in openpyxl) of
    # the INVOICE sheet is the heading 'COMMERCIAL INVOICE'.
    invoice_ws = wb["INVOICE"]
    assert invoice_ws.cell(row=1, column=1).value == "COMMERCIAL INVOICE"


def test_convert_xls_index_conversion_1based(corpus_path: Path) -> None:
    """Test that xlrd row 0 / col 0 maps to openpyxl row 1 / col 1."""
    book = xlrd.open_workbook(str(corpus_path), formatting_info=True)
    xlrd_value_0_0 = book.sheet_by_index(0).cell_value(0, 0)

    wb = convert_xls_to_xlsx(corpus_path)
    openpyxl_value = wb.active.cell(row=1, column=1).value  # type: ignore[union-attr]

    assert openpyxl_value == xlrd_value_0_0


def test_convert_xls_multiple_sheets_preserved(corpus_path: Path) -> None:
    """Test that all sheets from the source .xls are present in the Workbook."""
    book = xlrd.open_workbook(str(corpus_path), formatting_info=True)
    expected_names = book.sheet_names()

    wb = convert_xls_to_xlsx(corpus_path)

    assert len(wb.sheetnames) == len(expected_names)
    for name in expected_names:
        assert name in wb.sheetnames


def test_convert_xls_merge_ranges_preserved(corpus_path: Path) -> None:
    """Test that merge ranges from the source .xls appear in the openpyxl Workbook.

    The corpus INVOICE sheet has 6 merged cell ranges.  We verify at least one
    is present in the converted workbook and that its anchor cell retains its value.
    """
    wb = convert_xls_to_xlsx(corpus_path)
    invoice_ws = wb["INVOICE"]

    # Confirm some merged ranges exist in the converted sheet.
    merge_ranges = list(invoice_ws.merged_cells.ranges)
    assert len(merge_ranges) > 0

    # The xlrd tuple (0, 1, 0, 15) → openpyxl A1:O1 (row 1, cols 1-15).
    # Anchor at (0,0) xlrd == (1,1) openpyxl has value 'COMMERCIAL INVOICE'.
    anchor_value = invoice_ws.cell(row=1, column=1).value
    assert anchor_value == "COMMERCIAL INVOICE"

    # Verify the specific merge A1:O1 is present (rlo=0,rhi=1,clo=0,chi=15
    # converts to start_row=1, end_row=1, start_col=1, end_col=15).
    range_strs = {str(r) for r in merge_ranges}
    assert "A1:O1" in range_strs


def test_convert_xls_numeric_values_as_float(corpus_path: Path) -> None:
    """Test that integer-like numeric cells are stored as float in the Workbook.

    The INVOICE TOTAL row has QTY=91600 (stored by xlrd as 91600.0).
    In the corpus this is at xlrd row 33, col 4 → openpyxl row 34, col 5.
    """
    wb = convert_xls_to_xlsx(corpus_path)
    invoice_ws = wb["INVOICE"]

    # xlrd row 33, col 4 → openpyxl row 34, col 5
    cell_value = invoice_ws.cell(row=34, column=5).value

    assert cell_value == 91600.0
    assert isinstance(cell_value, float)


def test_convert_xls_text_cells(corpus_path: Path) -> None:
    """Test that text cells (including Chinese characters) are preserved as str."""
    wb = convert_xls_to_xlsx(corpus_path)
    invoice_ws = wb["INVOICE"]

    # xlrd row 3, col 0 → openpyxl row 4, col 1 has '茂綸股份有限公司'
    shipper_value = invoice_ws.cell(row=4, column=1).value

    assert isinstance(shipper_value, str)
    assert shipper_value == "茂綸股份有限公司"


def test_convert_xls_empty_cells_are_none(corpus_path: Path) -> None:
    """Test that XL_CELL_EMPTY cells produce None in the converted Workbook."""
    wb = convert_xls_to_xlsx(corpus_path)
    invoice_ws = wb["INVOICE"]

    # xlrd row 0, col 1 is XL_CELL_EMPTY (type=6) → openpyxl row 1, col 2
    empty_value = invoice_ws.cell(row=1, column=2).value

    assert empty_value is None


# ---------------------------------------------------------------------------
# Cell-type conversion tests via _convert_cell_value helper
# ---------------------------------------------------------------------------


def _make_cell(ctype: int, value: object) -> xlrd.sheet.Cell:
    """Create a minimal xlrd Cell-like object for testing."""
    cell = MagicMock(spec=xlrd.sheet.Cell)
    cell.ctype = ctype
    cell.value = value
    return cell


def test_convert_cell_value_date() -> None:
    """Test that XL_CELL_DATE cells are converted to datetime objects."""
    # xlrd date serial for 2025-09-11 with datemode=0 (1900 epoch).
    # Use xlrd to compute the serial so the test is not hard-coded to a magic number.
    import xlrd as _xlrd

    target = datetime(2025, 9, 11)
    # xlrd.xldate.xldate_from_datetime_tuple is not public; use date arithmetic.
    # Excel serial for 2025-09-11 (1900 datemode):  45911
    serial = 45911.0
    cell = _make_cell(_xlrd.XL_CELL_DATE, serial)
    result = _convert_cell_value(cell, datemode=0)

    assert isinstance(result, datetime)
    assert result.year == 2025
    assert result.month == 9
    assert result.day == 11


def test_convert_cell_value_boolean_true() -> None:
    """Test that XL_CELL_BOOLEAN with value=1 converts to Python True."""
    cell = _make_cell(xlrd.XL_CELL_BOOLEAN, 1)
    result = _convert_cell_value(cell, datemode=0)

    assert result is True
    assert isinstance(result, bool)


def test_convert_cell_value_boolean_false() -> None:
    """Test that XL_CELL_BOOLEAN with value=0 converts to Python False."""
    cell = _make_cell(xlrd.XL_CELL_BOOLEAN, 0)
    result = _convert_cell_value(cell, datemode=0)

    assert result is False
    assert isinstance(result, bool)


def test_convert_cell_value_error_is_none() -> None:
    """Test that XL_CELL_ERROR cells produce None (mirrors data_only=True)."""
    cell = _make_cell(xlrd.XL_CELL_ERROR, 7)  # 7 = #DIV/0! in xlrd
    result = _convert_cell_value(cell, datemode=0)

    assert result is None


def test_convert_cell_value_text() -> None:
    """Test that XL_CELL_TEXT cells are returned as str."""
    cell = _make_cell(xlrd.XL_CELL_TEXT, "Hello, World!")
    result = _convert_cell_value(cell, datemode=0)

    assert result == "Hello, World!"
    assert isinstance(result, str)


def test_convert_cell_value_number() -> None:
    """Test that XL_CELL_NUMBER cells are returned as float."""
    cell = _make_cell(xlrd.XL_CELL_NUMBER, 42.0)
    result = _convert_cell_value(cell, datemode=0)

    assert result == 42.0
    assert isinstance(result, float)


def test_convert_cell_value_empty() -> None:
    """Test that XL_CELL_EMPTY cells produce None."""
    cell = _make_cell(xlrd.XL_CELL_EMPTY, "")
    result = _convert_cell_value(cell, datemode=0)

    assert result is None


# ---------------------------------------------------------------------------
# Return-type and no-side-effect assertions
# ---------------------------------------------------------------------------


def test_convert_xls_returns_openpyxl_workbook(corpus_path: Path) -> None:
    """Test that convert_xls_to_xlsx always returns an openpyxl Workbook."""
    wb = convert_xls_to_xlsx(corpus_path)

    assert isinstance(wb, Workbook)


def test_convert_xls_does_not_modify_source(corpus_path: Path) -> None:
    """Test that the source .xls file is not modified by the conversion."""
    import hashlib

    before = hashlib.md5(corpus_path.read_bytes()).hexdigest()
    convert_xls_to_xlsx(corpus_path)
    after = hashlib.md5(corpus_path.read_bytes()).hexdigest()

    assert before == after, "Source .xls file was modified during conversion"
