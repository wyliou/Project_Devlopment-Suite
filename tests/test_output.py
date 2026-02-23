"""Tests for output.py — FR-029, FR-030: Template population and output file writing.

Seven test cases covering:
1. Output file is created and is a valid workbook.
2. Template rows 1–4 are preserved after write.
3. Column mapping correctness (spot checks on key columns).
4. total_gw and total_packets written only at row 5.
5. total_packets=None produces an empty cell at AK row 5.
6. ERR_051 raised when template path does not exist.
7. ERR_052 raised when output path is unwritable.
"""

import stat
import sys
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.models import AppConfig, InvoiceItem, PackingTotals
from autoconvert.output import write_template

# ---------------------------------------------------------------------------
# Helpers / factories
# ---------------------------------------------------------------------------

_INVOICE_DEFAULTS = {
    "po_no": "PO-001",
    "price": Decimal("1.23456"),
    "amount": Decimal("12.34"),
    "currency": "502",
    "coo": "116",
    "cod": "US",
    "brand": "BrandX",
    "brand_type": "1",
    "model_no": "MDL-001",
    "inv_no": "INV-2024-001",
    "serial": "1",
}


def _item(part_no: str = "PART-001", **overrides: object) -> InvoiceItem:
    """Create an InvoiceItem with sensible defaults.

    Args:
        part_no: Part number string.
        **overrides: Field overrides applied after defaults.

    Returns:
        InvoiceItem instance.
    """
    fields = {
        **_INVOICE_DEFAULTS,
        "part_no": part_no,
        "qty": Decimal("10.00000"),
        "allocated_weight": Decimal("2.50000"),
    }
    fields.update(overrides)
    return InvoiceItem(**fields)  # type: ignore[arg-type]


def _totals(
    total_gw: Decimal = Decimal("15.50000"),
    total_packets: int | None = 3,
) -> PackingTotals:
    """Create a PackingTotals instance.

    Args:
        total_gw: Total gross weight.
        total_packets: Total packet count; may be None.

    Returns:
        PackingTotals instance.
    """
    return PackingTotals(
        total_nw=Decimal("14.00000"),
        total_nw_precision=5,
        total_gw=total_gw,
        total_gw_precision=5,
        total_packets=total_packets,
    )


def _make_config(project_root: Path) -> AppConfig:
    """Build a minimal AppConfig pointing to the real template.

    Args:
        project_root: Absolute path to the project root.

    Returns:
        AppConfig with template_path set.
    """
    import re

    return AppConfig(
        invoice_sheet_patterns=[re.compile(r"invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"packing", re.IGNORECASE)],
        invoice_columns={},
        packing_columns={},
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=project_root / "config" / "output_template.xlsx",
    )


# ---------------------------------------------------------------------------
# Test 1 — Output file is created and valid
# ---------------------------------------------------------------------------


def test_write_template_creates_output_file(
    project_root: Path, tmp_output: Path
) -> None:
    """Call with 2 invoice items; verify output file exists and is openable."""
    config = _make_config(project_root)
    output_path = tmp_output / "test_creates_template.xlsx"
    items = [_item("PART-001"), _item("PART-002")]

    write_template(items, _totals(), config, output_path)

    assert output_path.exists(), "Output file was not created"
    wb = openpyxl.load_workbook(output_path)
    assert wb is not None
    assert "工作表1" in wb.sheetnames


# ---------------------------------------------------------------------------
# Test 2 — Template rows 1–4 are preserved
# ---------------------------------------------------------------------------


def test_write_template_rows_1_to_4_preserved(
    project_root: Path, tmp_output: Path
) -> None:
    """Rows 1–4 content must be identical before and after write_template()."""
    config = _make_config(project_root)

    # Capture template rows 1–4 before calling write_template
    template_wb = openpyxl.load_workbook(config.template_path)
    template_ws = template_wb["工作表1"]
    before: dict[tuple[int, int], object] = {}
    for row in range(1, 5):
        for col in range(1, 41):
            before[(row, col)] = template_ws.cell(row=row, column=col).value

    # Write the output file
    output_path = tmp_output / "test_rows_preserved_template.xlsx"
    items = [_item()]
    write_template(items, _totals(), config, output_path)

    # Reload the output and compare rows 1–4
    out_wb = openpyxl.load_workbook(output_path)
    out_ws = out_wb["工作表1"]
    for (row, col), expected_value in before.items():
        actual_value = out_ws.cell(row=row, column=col).value
        assert actual_value == expected_value, (
            f"Row {row}, Col {col}: expected {expected_value!r}, got {actual_value!r}"
        )

    # Row 5 col A should have first item's part_no (data was written)
    assert out_ws.cell(row=5, column=1).value == "PART-001"


# ---------------------------------------------------------------------------
# Test 3 — Column mapping correctness
# ---------------------------------------------------------------------------


def test_write_template_column_mapping_correct(
    project_root: Path, tmp_output: Path
) -> None:
    """Key columns for row 5 carry the expected values from the first item."""
    config = _make_config(project_root)
    output_path = tmp_output / "test_col_mapping_template.xlsx"
    pt = _totals(total_gw=Decimal("20.12345"), total_packets=5)
    item = _item(
        part_no="PN-XYZ",
        qty=Decimal("7.00000"),
        currency="502",
        coo="116",
        allocated_weight=Decimal("3.50000"),
        inv_no="INV-999",
        serial="2",
        brand="MyBrand",
        brand_type="0",
        model_no="MODEL-A",
    )

    write_template([item], pt, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["工作表1"]
    row = 5

    assert ws.cell(row=row, column=1).value == "PN-XYZ", "col A: part_no"
    assert ws.cell(row=row, column=3).value == "3", "col C: fixed '3'"
    assert ws.cell(row=row, column=4).value == "502", "col D: currency"
    assert ws.cell(row=row, column=5).value == Decimal("7.00000"), "col E: qty"
    # openpyxl reads Decimal values back as float; compare via Decimal(str(...))
    assert Decimal(str(ws.cell(row=row, column=13).value)) == Decimal("3.50000"), (
        "col M: allocated_weight"
    )
    assert Decimal(str(ws.cell(row=row, column=16).value)) == Decimal("20.12345"), (
        "col P: total_gw"
    )
    assert ws.cell(row=row, column=18).value == "32052", "col R: fixed '32052'"
    assert ws.cell(row=row, column=19).value == "320506", "col S: fixed '320506'"
    assert ws.cell(row=row, column=20).value == "142", "col T: fixed '142'"
    assert ws.cell(row=row, column=40).value == "MODEL-A", "col AN: model_no"


# ---------------------------------------------------------------------------
# Test 4 — total_gw and total_packets are row 5 ONLY
# ---------------------------------------------------------------------------


def test_write_template_total_gw_and_packets_row5_only(
    project_root: Path, tmp_output: Path
) -> None:
    """total_gw at col P and total_packets at col AK are written only for row 5."""
    config = _make_config(project_root)
    output_path = tmp_output / "test_row5_only_template.xlsx"
    items = [_item("PART-001"), _item("PART-002")]
    pt = _totals(total_gw=Decimal("30.00000"), total_packets=10)

    write_template(items, pt, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["工作表1"]

    # Row 5 (item index 0) — must have values
    # openpyxl reads Decimal back as float; use Decimal(str(...)) for comparison
    assert Decimal(str(ws.cell(row=5, column=16).value)) == Decimal("30.00000"), (
        "col P row 5 must have total_gw"
    )
    assert ws.cell(row=5, column=37).value == 10, "col AK row 5 must have total_packets"

    # Row 6 (item index 1) — must be empty
    assert ws.cell(row=6, column=16).value is None, "col P row 6 must be empty"
    assert ws.cell(row=6, column=37).value is None, "col AK row 6 must be empty"


# ---------------------------------------------------------------------------
# Test 5 — total_packets=None writes empty cell at AK row 5
# ---------------------------------------------------------------------------


def test_write_template_total_packets_none_writes_empty(
    project_root: Path, tmp_output: Path
) -> None:
    """When total_packets is None, col AK at row 5 stays empty (no error)."""
    config = _make_config(project_root)
    output_path = tmp_output / "test_packets_none_template.xlsx"
    pt = _totals(total_packets=None)

    # Must not raise
    write_template([_item()], pt, config, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["工作表1"]
    assert ws.cell(row=5, column=37).value is None, (
        "col AK row 5 must be empty when total_packets is None"
    )
    # total_gw should still be written
    assert Decimal(str(ws.cell(row=5, column=16).value)) == Decimal("15.50000")


# ---------------------------------------------------------------------------
# Test 6 — ERR_051 on bad template path
# ---------------------------------------------------------------------------


def test_write_template_err051_on_bad_template_path(
    project_root: Path, tmp_output: Path
) -> None:
    """ProcessingError(ERR_051) raised when template file does not exist."""
    import re

    bad_config = AppConfig(
        invoice_sheet_patterns=[re.compile(r"invoice")],
        packing_sheet_patterns=[re.compile(r"packing")],
        invoice_columns={},
        packing_columns={},
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=Path("/nonexistent/path/output_template.xlsx"),
    )
    output_path = tmp_output / "should_not_exist.xlsx"

    with pytest.raises(ProcessingError) as exc_info:
        write_template([_item()], _totals(), bad_config, output_path)

    assert exc_info.value.code == ErrorCode.ERR_051


# ---------------------------------------------------------------------------
# Test 7 — ERR_052 on unwritable output path
# ---------------------------------------------------------------------------


def test_write_template_err052_on_unwritable_path(
    project_root: Path, tmp_output: Path
) -> None:
    """ProcessingError(ERR_052) raised when output_path is in a read-only directory."""
    config = _make_config(project_root)

    # Create a read-only subdirectory
    read_only_dir = tmp_output / "read_only"
    read_only_dir.mkdir()

    if sys.platform == "win32":
        # On Windows, use icacls to remove write permission from the directory
        import subprocess

        subprocess.run(
            ["icacls", str(read_only_dir), "/deny", "Everyone:(W)"],
            check=True,
            capture_output=True,
        )
    else:
        read_only_dir.chmod(stat.S_IRUSR | stat.S_IXUSR)

    output_path = read_only_dir / "output_template.xlsx"

    try:
        with pytest.raises(ProcessingError) as exc_info:
            write_template([_item()], _totals(), config, output_path)
        assert exc_info.value.code == ErrorCode.ERR_052
    finally:
        # Restore permissions so tmp_path cleanup can remove the directory
        if sys.platform == "win32":
            subprocess.run(
                ["icacls", str(read_only_dir), "/remove:d", "Everyone"],
                check=False,
                capture_output=True,
            )
        else:
            read_only_dir.chmod(
                stat.S_IRUSR | stat.S_IWUSR | stat.S_IXUSR
            )
