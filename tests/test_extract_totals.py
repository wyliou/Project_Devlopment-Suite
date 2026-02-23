"""Tests for extract_totals — FR-014, FR-015, FR-016, FR-017."""

import logging
from decimal import Decimal

import pytest
from openpyxl import Workbook

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.extract_totals import detect_total_row, extract_totals
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_sheet_with_data(
    rows: dict[tuple[int, int], object],
    *,
    number_formats: dict[tuple[int, int], str] | None = None,
    merges: list[str] | None = None,
) -> tuple:
    """Create an in-memory openpyxl worksheet with specified cell data.

    Args:
        rows: Mapping of (row, col) -> value for cells to populate.
        number_formats: Mapping of (row, col) -> format string.
        merges: List of merge range strings (e.g. "A1:A3") to apply BEFORE
            creating the MergeTracker.

    Returns:
        Tuple of (worksheet, MergeTracker).
    """
    wb = Workbook()
    ws = wb.active

    for (r, c), val in rows.items():
        ws.cell(row=r, column=c, value=val)

    if number_formats:
        for (r, c), fmt in number_formats.items():
            ws.cell(row=r, column=c).number_format = fmt

    if merges:
        for merge_range in merges:
            ws.merge_cells(merge_range)

    tracker = MergeTracker(ws)
    return ws, tracker


def _make_column_map(
    *,
    part_no: int = 1,
    nw: int = 5,
    gw: int = 6,
    sheet_type: str = "packing",
    header_row: int = 1,
) -> ColumnMapping:
    """Create a ColumnMapping with configurable field indices.

    Args:
        part_no: 1-based column index for part_no.
        nw: 1-based column index for nw.
        gw: 1-based column index for gw.
        sheet_type: Sheet type string.
        header_row: Header row number.

    Returns:
        ColumnMapping instance.
    """
    return ColumnMapping(
        sheet_type=sheet_type,
        field_map={"part_no": part_no, "nw": nw, "gw": gw},
        header_row=header_row,
        effective_header_row=header_row,
    )


# ---------------------------------------------------------------------------
# detect_total_row() tests — FR-014
# ---------------------------------------------------------------------------


class TestDetectTotalRow:
    """Tests for detect_total_row()."""

    def test_detect_total_row_strategy1_keyword_total(self) -> None:
        """Test Strategy 1: row with 'TOTAL' in column A is detected as total row."""
        last_data_row = 10
        ws, tracker = _make_sheet_with_data(
            {
                # Row 13 (last_data_row + 3) has "TOTAL" in column A
                (13, 1): "TOTAL",
                (13, 5): 100.0,
                (13, 6): 120.0,
            }
        )
        col_map = _make_column_map()

        result = detect_total_row(ws, last_data_row, col_map, tracker)

        assert result == 13

    def test_detect_total_row_strategy1_keyword_chinese_合计(self) -> None:
        """Test Strategy 1: row with '合计' in column F is detected correctly."""
        last_data_row = 10
        ws, tracker = _make_sheet_with_data(
            {
                # Row 11 (last_data_row + 1) has "合计" in column F
                (11, 6): "合计",
                (11, 5): 100.0,
            }
        )
        col_map = _make_column_map()

        result = detect_total_row(ws, last_data_row, col_map, tracker)

        assert result == 11

    def test_detect_total_row_strategy2_implicit_when_strategy1_fails(self) -> None:
        """Test Strategy 2: no keyword rows; implicit detection via empty part_no + NW/GW > 0."""
        last_data_row = 10
        ws, tracker = _make_sheet_with_data(
            {
                # Row 12 (last_data_row + 2): empty part_no, NW=15.5, GW=18.0
                (12, 1): None,  # part_no empty
                (12, 5): 15.5,  # NW > 0
                (12, 6): 18.0,  # GW > 0
            }
        )
        col_map = _make_column_map()

        result = detect_total_row(ws, last_data_row, col_map, tracker)

        assert result == 12

    def test_detect_total_row_strategy2_excludes_merge_continuation(self) -> None:
        """Test Strategy 2: row with empty part_no due to vertical merge is excluded.

        MergeTracker shows the row is in a merge, so it should NOT be detected.
        Strategy 2 continues search to find the actual total row.
        """
        last_data_row = 10
        ws, tracker = _make_sheet_with_data(
            rows={
                # Row 12: part_no is in a merge (continuation row) — should be skipped
                # (Part_no col=1 is merged from row 9 to row 12)
                (9, 1): "PART-A",
                (12, 5): 15.5,
                (12, 6): 18.0,
                # Row 13: actual total row (empty part_no, not merged, NW/GW > 0)
                (13, 1): None,
                (13, 5): 30.0,
                (13, 6): 35.0,
            },
            merges=["A9:A12"],  # part_no merged from row 9 to 12
        )
        col_map = _make_column_map()

        result = detect_total_row(ws, last_data_row, col_map, tracker)

        assert result == 13

    def test_detect_total_row_both_strategies_fail_raises_err032(self) -> None:
        """Test that ERR_032 is raised when both strategies fail."""
        last_data_row = 10
        # No keyword rows and no implicit-pattern rows in range
        ws, tracker = _make_sheet_with_data(
            {
                # Rows 11-25: all have part_no values (non-empty), no keywords
                (11, 1): "PART-X",
                (11, 5): 10.0,
                (11, 6): 12.0,
            }
        )
        col_map = _make_column_map()

        with pytest.raises(ProcessingError) as exc_info:
            detect_total_row(ws, last_data_row, col_map, tracker)

        assert exc_info.value.code == ErrorCode.ERR_032


# ---------------------------------------------------------------------------
# extract_totals() tests — FR-015 (total_nw)
# ---------------------------------------------------------------------------


class TestExtractTotalsNW:
    """Tests for extract_totals() — FR-015 total_nw."""

    def test_extract_totals_nw_explicit_format_precision(self) -> None:
        """Test NW cell format '0.00000_'; value 15.5; total_nw = Decimal('15.50000'), precision = 5."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.5,  # NW
                (10, 6): 18.0,  # GW
            },
            number_formats={
                (10, 5): "0.00000_",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_nw == Decimal("15.50000")
        assert result.total_nw_precision == 5

    def test_extract_totals_nw_general_format_trailing_zero_normalization(self) -> None:
        """Test NW General format: value 15.5000000000001 (float artifact) normalizes to 2dp."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.5000000000001,  # NW (float artifact)
                (10, 6): 18.0,  # GW
            },
            number_formats={
                (10, 5): "General",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        # General format: round to 5dp -> "15.50000", strip trailing zeros -> "15.5"
        # But min precision is 2, so precision = 2
        assert result.total_nw == Decimal("15.50")
        assert result.total_nw_precision == 2

    def test_extract_totals_nw_with_unit_suffix(self) -> None:
        """Test NW cell value '15.5 KG'; unit stripped; total_nw = Decimal('15.50'), precision = 2."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): "15.5 KG",  # NW with unit suffix
                (10, 6): 18.0,  # GW
            },
            number_formats={
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_nw == Decimal("15.50")
        assert result.total_nw_precision == 2

    def test_extract_totals_nw_invalid_raises_err033(self) -> None:
        """Test ERR_033 raised when NW cell is empty."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): None,  # NW empty
                (10, 6): 18.0,  # GW
            },
        )
        col_map = _make_column_map()

        with pytest.raises(ProcessingError) as exc_info:
            extract_totals(ws, total_row, col_map)

        assert exc_info.value.code == ErrorCode.ERR_033


# ---------------------------------------------------------------------------
# extract_totals() tests — FR-016 (total_gw)
# ---------------------------------------------------------------------------


class TestExtractTotalsGW:
    """Tests for extract_totals() — FR-016 total_gw."""

    def test_extract_totals_gw_packaging_weight_addition(self) -> None:
        """Test GW with packaging weight: total_row=18.0, +1=2.5, +2=20.5; use +2 value."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): 18.0,  # GW at total_row
                (11, 6): 2.5,  # GW at +1 (pallet weight)
                (12, 6): 20.5,  # GW at +2 (final total)
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
                (12, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_gw == Decimal("20.50")
        assert result.total_gw_precision == 2

    def test_extract_totals_gw_no_additional_rows(self) -> None:
        """Test GW with no additional rows (+1 GW is empty); use total_row value."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): 18.0,  # GW at total_row
                # Row 11, col 6 is empty — no packaging weight
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_gw == Decimal("18.00")
        assert result.total_gw_precision == 2

    def test_extract_totals_gw_invalid_raises_err034(self) -> None:
        """Test ERR_034 raised when GW cell is empty or non-numeric."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): None,  # GW empty
            },
            number_formats={
                (10, 5): "0.00",
            },
        )
        col_map = _make_column_map()

        with pytest.raises(ProcessingError) as exc_info:
            extract_totals(ws, total_row, col_map)

        assert exc_info.value.code == ErrorCode.ERR_034


# ---------------------------------------------------------------------------
# extract_totals() tests — FR-017 (total_packets)
# ---------------------------------------------------------------------------


class TestExtractTotalsPackets:
    """Tests for extract_totals() — FR-017 total_packets."""

    def test_extract_totals_packets_priority1_件数_label(self) -> None:
        """Test Priority 1: 件数 label with adjacent right cell '55 件'."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): 18.0,  # GW
                # Priority 1: row total_row+1, label in col 2, value in col 3
                (11, 2): "件数",
                (11, 3): "55 件",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 55

    def test_extract_totals_packets_priority2_plt_g_number_before(self) -> None:
        """Test Priority 2: '7PLT.G' at total_row-1; extract 7."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): 18.0,  # GW
                # Priority 2: row total_row-1, "7PLT.G" in col 3
                (9, 3): "7PLT.G",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_priority3a_total_with_breakdown(self) -> None:
        """Test Priority 3a: '348（256胶框+92纸箱）'; extract 348."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,  # NW
                (10, 6): 18.0,  # GW
                (11, 2): "348（256胶框+92纸箱）",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 348

    def test_extract_totals_packets_priority3b_unit_suffix_托(self) -> None:
        """Test Priority 3b: '7托'; extract 7."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                (11, 2): "7托",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_priority3b_unit_suffix_ctns(self) -> None:
        """Test Priority 3b: '55 CTNS'; extract 55."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                (11, 2): "55 CTNS",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 55

    def test_extract_totals_packets_priority3c_embedded_chinese_共n托(self) -> None:
        """Test Priority 3c: '共7托' at total_row+2; extract 7."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                (12, 2): "共7托",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_priority3d_pallet_range(self) -> None:
        """Test Priority 3d: 'PLT#1(1~34)' at total_row+1; extract 34."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                (11, 2): "PLT#1(1~34)",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 34

    def test_extract_totals_packets_pallet_vs_box_priority(self) -> None:
        """Test pallet vs box: '共7托（172件）' extracts pallet count 7, not box count 172."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                (11, 2): "共7托（172件）",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        result = extract_totals(ws, total_row, col_map)

        assert result.total_packets == 7

    def test_extract_totals_packets_not_found_att002_warning(self, caplog: pytest.LogCaptureFixture) -> None:
        """Test ATT_002 warning when no priority pattern matches; total_packets = None."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                # No packet data in any search rows
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        with caplog.at_level(logging.DEBUG):
            result = extract_totals(ws, total_row, col_map)

        assert result.total_packets is None
        assert "not found" in caplog.text.lower()

    def test_extract_totals_packets_out_of_range_treated_as_not_found(self, caplog: pytest.LogCaptureFixture) -> None:
        """Test value > 1000 treated as not found; continues to next priority."""
        total_row = 10
        ws, _ = _make_sheet_with_data(
            rows={
                (10, 5): 15.0,
                (10, 6): 18.0,
                # All patterns yield value 1500 which is > 1000
                (11, 2): "1500托",
            },
            number_formats={
                (10, 5): "0.00",
                (10, 6): "0.00",
            },
        )
        col_map = _make_column_map()

        with caplog.at_level(logging.DEBUG):
            result = extract_totals(ws, total_row, col_map)

        assert result.total_packets is None
        assert "not found" in caplog.text.lower()
