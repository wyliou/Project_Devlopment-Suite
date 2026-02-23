"""Tests for extract_packing — FR-012, FR-013 acceptance criteria.

Every test creates an in-memory openpyxl Workbook with mock packing data,
initializes MergeTracker (to capture/unmerge), builds a ColumnMapping, then
exercises extract_packing_items() and validate_merged_weights().
"""

from decimal import Decimal

import pytest
from openpyxl import Workbook

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.extract_packing import extract_packing_items, validate_merged_weights
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Default column layout for tests:
# A=1: part_no, B=2: qty, C=3: nw, D=4: gw
_DEFAULT_FIELD_MAP = {"part_no": 1, "qty": 2, "nw": 3, "gw": 4}
_HEADER_ROW = 1


def _make_column_map(
    field_map: dict[str, int] | None = None,
    header_row: int = _HEADER_ROW,
    effective_header_row: int | None = None,
) -> ColumnMapping:
    """Create a ColumnMapping for tests.

    Args:
        field_map: Column index mapping. Defaults to A=part_no, B=qty, C=nw, D=gw.
        header_row: 1-based header row.
        effective_header_row: Defaults to header_row if not given.

    Returns:
        ColumnMapping instance.
    """
    if field_map is None:
        field_map = dict(_DEFAULT_FIELD_MAP)
    if effective_header_row is None:
        effective_header_row = header_row
    return ColumnMapping(
        sheet_type="packing",
        field_map=field_map,
        header_row=header_row,
        effective_header_row=effective_header_row,
    )


def _make_packing_sheet(
    rows: list[tuple],
    header_row: int = _HEADER_ROW,
    merges: list[str] | None = None,
) -> tuple:
    """Create an in-memory packing sheet with data rows and optional merges.

    Header is placed at header_row. Data rows start at header_row + 1.
    Each tuple in ``rows`` must have 4 elements: (part_no, qty, nw, gw).
    Use None for empty cells.

    Args:
        rows: List of (part_no, qty, nw, gw) tuples.
        header_row: 1-based header row number.
        merges: List of merge range strings (e.g., "A2:A4") to apply BEFORE
            MergeTracker initialization.

    Returns:
        Tuple of (worksheet, MergeTracker).
    """
    wb = Workbook()
    ws = wb.active

    # Write header
    ws.cell(row=header_row, column=1, value="Part No")
    ws.cell(row=header_row, column=2, value="QTY")
    ws.cell(row=header_row, column=3, value="N.W.")
    ws.cell(row=header_row, column=4, value="G.W.")

    # Write data rows
    data_start = header_row + 1
    for i, (part_no, qty, nw, gw) in enumerate(rows):
        r = data_start + i
        if part_no is not None:
            ws.cell(row=r, column=1, value=part_no)
        if qty is not None:
            ws.cell(row=r, column=2, value=qty)
        if nw is not None:
            ws.cell(row=r, column=3, value=nw)
        if gw is not None:
            ws.cell(row=r, column=4, value=gw)

    # Apply merges BEFORE MergeTracker
    if merges:
        for merge_range in merges:
            ws.merge_cells(merge_range)

    tracker = MergeTracker(ws)
    return ws, tracker


# ---------------------------------------------------------------------------
# extract_packing_items() tests
# ---------------------------------------------------------------------------


class TestExtractPackingItemsNormal:
    """Tests for normal extraction scenarios."""

    def test_extract_packing_items_normal_extraction(self):
        """Test normal extraction of three rows with distinct parts.

        Verify list length 3, correct Decimal values (nw at 5 decimals),
        last_data_row = row number of third item.
        """
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 20, 3.75, 4.5),
            ("P3", 5, 1.2, 1.8),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, last_data_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 3
        assert items[0].part_no == "P1"
        assert items[0].qty == Decimal("10")
        assert items[0].nw == Decimal("2.50000")
        assert items[0].is_first_row_of_merge is True

        assert items[1].part_no == "P2"
        assert items[1].nw == Decimal("3.75000")

        assert items[2].part_no == "P3"
        assert items[2].nw == Decimal("1.20000")

        # last_data_row is 1-based row of the third item (header=1, data starts at 2)
        assert last_data_row == 4

    def test_extract_packing_items_unit_suffix_stripped_from_nw(self):
        """Test that NW cell '2.5 KGS' is extracted as Decimal('2.50000')."""
        rows = [
            ("P1", 10, "2.5 KGS", 3.0),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 1
        assert items[0].nw == Decimal("2.50000")
        assert items[0].is_first_row_of_merge is True


class TestExtractPackingItemsContinuation:
    """Tests for implicit continuation handling."""

    def test_extract_packing_items_implicit_continuation_nw_zero(self):
        """Test implicit NW continuation: same part_no, empty NW.

        Row 1: part_no='P1', qty=10, nw=2.5.
        Row 2: part_no='P1', qty=5, nw=empty.
        Row 2 should be extracted as nw=0.00000, is_first_row_of_merge=False.
        """
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P1", 5, None, None),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].part_no == "P1"
        assert items[1].qty == Decimal("5")
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False


class TestExtractPackingItemsDittoMarks:
    """Tests for ditto mark handling in NW column."""

    def test_extract_packing_items_ditto_mark_sets_nw_zero(self):
        """Test ditto mark U+0022 (double quote) sets nw=0, no ERR_031."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 5, '"', 3.0),  # U+0022 double quote
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_ditto_mark_unicode_3003(self):
        """Test ditto mark U+3003 (Japanese ditto) sets nw=0."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 5, "\u3003", 3.0),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_ditto_mark_unicode_201c(self):
        """Test ditto mark U+201C (left double quote) sets nw=0."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 5, "\u201c", 3.0),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False

    def test_extract_packing_items_ditto_mark_unicode_201d(self):
        """Test ditto mark U+201D (right double quote) sets nw=0."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 5, "\u201d", 3.0),
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False


class TestExtractPackingItemsMergedPartNo:
    """Tests for merged part_no propagation."""

    def test_extract_packing_items_merged_part_no_propagated(self):
        """Test part_no merged vertically across rows 2-4 (header=1).

        Rows 3-4 should get propagated part_no. Stop condition 3 must NOT
        trigger on rows 3-4 even though part_no is empty before propagation.
        """
        # Set up: header at row 1, data starts at row 2
        # Part_no merged A2:A4 = "P1"
        # Each row has its own qty and nw
        wb = Workbook()
        ws = wb.active

        # Header
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")
        ws.cell(row=1, column=4, value="G.W.")

        # Data
        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=2.5)
        ws.cell(row=2, column=4, value=3.0)

        # Rows 3-4: part_no will be None after unmerge, but NW and GW filled
        ws.cell(row=3, column=2, value=20)
        ws.cell(row=3, column=3, value=3.0)
        ws.cell(row=3, column=4, value=4.0)

        ws.cell(row=4, column=2, value=15)
        ws.cell(row=4, column=3, value=1.5)
        ws.cell(row=4, column=4, value=2.0)

        # Merge part_no A2:A4
        ws.merge_cells("A2:A4")

        tracker = MergeTracker(ws)
        col_map = _make_column_map()

        items, last_data_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 3
        # All three should have part_no="P1" via propagation
        assert items[0].part_no == "P1"
        assert items[1].part_no == "P1"
        assert items[2].part_no == "P1"
        # Verify nw values
        assert items[0].nw == Decimal("2.50000")
        assert items[1].nw == Decimal("3.00000")
        assert items[2].nw == Decimal("1.50000")
        assert last_data_row == 4


class TestExtractPackingItemsStopConditions:
    """Tests for stop conditions."""

    def test_extract_packing_items_stop_at_total_keyword(self):
        """Test extraction stops when 'TOTAL' appears in column B (within A-J)."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 20, 3.0, 4.0),
            (None, "TOTAL", 5.5, 7.0),  # Stop keyword in col B
            ("P3", 5, 1.0, 1.5),  # Should NOT be extracted
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, last_data_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[0].part_no == "P1"
        assert items[1].part_no == "P2"
        assert last_data_row == 3

    def test_extract_packing_items_stop_at_blank_row(self):
        """Test extraction stops at truly blank row after first data row."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            (None, None, None, None),  # Blank row
            ("P3", 5, 1.0, 1.5),  # Should NOT be extracted
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, last_data_row = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 1
        assert items[0].part_no == "P1"
        assert last_data_row == 2

    def test_extract_packing_items_stop_check_before_blank_check(self):
        """Test stop keyword fires before blank check.

        Row has Chinese stop keyword in col A AND is otherwise blank-like.
        Stop should fire on keyword, not blank.
        """
        # Row with stop keyword in col A
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")
        ws.cell(row=1, column=4, value="G.W.")

        # Data row
        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=2.5)
        ws.cell(row=2, column=4, value=3.0)

        # Row with "合计" in col A and blank qty/nw
        ws.cell(row=3, column=1, value="合计")

        tracker = MergeTracker(ws)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        # Should stop at row 3 (keyword), extracting only 1 item
        assert len(items) == 1
        assert items[0].part_no == "P1"


class TestExtractPackingItemsFiltering:
    """Tests for row filtering rules."""

    def test_extract_packing_items_pallet_summary_row_skipped(self):
        """Test pallet summary row (PLT.) is skipped but extraction continues."""
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("PLT.", 1, 0.5, 0.8),  # Pallet row — skipped
            ("P2", 20, 3.0, 4.0),  # Should be extracted
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 2
        assert items[0].part_no == "P1"
        assert items[1].part_no == "P2"


class TestExtractPackingItemsMergedNW:
    """Tests for merged NW cell handling."""

    def test_extract_packing_items_merged_nw_anchor_retains_weight(self):
        """Test merged NW: anchor row retains weight, continuation rows get 0.

        Rows 2-4 share merged NW cell (C2:C4 = 3.0).
        Row 2 is anchor (nw=3.0), rows 3-4 get nw=0.00000.
        """
        wb = Workbook()
        ws = wb.active

        # Header
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")
        ws.cell(row=1, column=4, value="G.W.")

        # Data (same part_no for implicit continuation of QTY too)
        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=3.0)
        ws.cell(row=2, column=4, value=4.0)

        ws.cell(row=3, column=1, value="P1")
        ws.cell(row=3, column=2, value=5)
        ws.cell(row=3, column=4, value=4.0)

        ws.cell(row=4, column=1, value="P1")
        ws.cell(row=4, column=2, value=8)
        ws.cell(row=4, column=4, value=4.0)

        # Merge NW column C2:C4
        ws.merge_cells("C2:C4")

        tracker = MergeTracker(ws)
        col_map = _make_column_map()

        items, _ = extract_packing_items(ws, col_map, tracker)

        assert len(items) == 3
        # Anchor row retains weight
        assert items[0].nw == Decimal("3.00000")
        assert items[0].is_first_row_of_merge is True
        # Non-anchor rows get zero
        assert items[1].nw == Decimal("0.00000")
        assert items[1].is_first_row_of_merge is False
        assert items[2].nw == Decimal("0.00000")
        assert items[2].is_first_row_of_merge is False


class TestExtractPackingItemsErrors:
    """Tests for error raising."""

    def test_extract_packing_items_err030_raised_for_non_continuation_empty_required(
        self,
    ):
        """Test ERR_030 raised for genuinely empty part_no on non-continuation row.

        Row with empty NW, not a merge continuation, and different part_no from
        previous item.
        """
        rows = [
            ("P1", 10, 2.5, 3.0),
            ("P2", 5, None, None),  # P2 != P1, so NW is not continuation → ERR_030
        ]
        ws, tracker = _make_packing_sheet(rows)
        col_map = _make_column_map()

        with pytest.raises(ProcessingError) as exc_info:
            extract_packing_items(ws, col_map, tracker)

        assert exc_info.value.code == ErrorCode.ERR_030
        assert "nw" in exc_info.value.context.get("field_name", "")


# ---------------------------------------------------------------------------
# validate_merged_weights() tests
# ---------------------------------------------------------------------------


class TestValidateMergedWeightsSamePartNo:
    """Tests for validate_merged_weights with same part_no."""

    def test_validate_merged_weights_same_part_no_ok(self):
        """Test no error when rows sharing merged NW have same part_no."""
        wb = Workbook()
        ws = wb.active

        # Header at row 1
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")

        # Data: rows 2-3, same part_no
        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=5.0)

        ws.cell(row=3, column=1, value="P1")
        ws.cell(row=3, column=2, value=5)

        # Merge NW C2:C3
        ws.merge_cells("C2:C3")

        tracker = MergeTracker(ws)
        col_map = _make_column_map(field_map={"part_no": 1, "qty": 2, "nw": 3})

        packing_items = [
            # Items extracted from the sheet
        ]

        # Should not raise
        validate_merged_weights(packing_items, tracker, col_map)


class TestValidateMergedWeightsDifferentPartNo:
    """Tests for validate_merged_weights with different part_nos."""

    def test_validate_merged_weights_different_part_no_raises_err046(self):
        """Test ERR_046 raised when different part_nos share merged NW cell."""
        wb = Workbook()
        ws = wb.active

        # Header at row 1
        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")

        # Data: rows 2-3, different part_nos
        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=5.0)

        ws.cell(row=3, column=1, value="P2")
        ws.cell(row=3, column=2, value=5)

        # Merge NW C2:C3
        ws.merge_cells("C2:C3")

        tracker = MergeTracker(ws)
        col_map = _make_column_map(field_map={"part_no": 1, "qty": 2, "nw": 3})

        packing_items = []

        with pytest.raises(ProcessingError) as exc_info:
            validate_merged_weights(packing_items, tracker, col_map)

        assert exc_info.value.code == ErrorCode.ERR_046
        assert "P1" in str(exc_info.value.message)
        assert "P2" in str(exc_info.value.message)


class TestValidateMergedWeightsEdgeCases:
    """Edge case tests for validate_merged_weights."""

    def test_validate_merged_weights_header_area_merge_ignored(self):
        """Test that merges in header area (row <= header_row) are not checked."""
        wb = Workbook()
        ws = wb.active

        # Header at row 2
        ws.cell(row=2, column=1, value="Part No")
        ws.cell(row=2, column=2, value="QTY")
        ws.cell(row=2, column=3, value="N.W.")

        # Merge in header area (row 1, which is <= header_row=2)
        ws.cell(row=1, column=3, value="Weight Info")
        ws.merge_cells("C1:C2")

        # Data rows after header
        ws.cell(row=3, column=1, value="P1")
        ws.cell(row=3, column=2, value=10)
        ws.cell(row=3, column=3, value=5.0)

        tracker = MergeTracker(ws)
        col_map = _make_column_map(
            field_map={"part_no": 1, "qty": 2, "nw": 3},
            header_row=2,
        )

        # Should not raise — header area merge is ignored
        validate_merged_weights([], tracker, col_map)

    def test_validate_merged_weights_no_merges_no_error(self):
        """Test no error when sheet has no merged cells."""
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Part No")
        ws.cell(row=1, column=2, value="QTY")
        ws.cell(row=1, column=3, value="N.W.")

        ws.cell(row=2, column=1, value="P1")
        ws.cell(row=2, column=2, value=10)
        ws.cell(row=2, column=3, value=5.0)

        tracker = MergeTracker(ws)
        col_map = _make_column_map(field_map={"part_no": 1, "qty": 2, "nw": 3})

        # Should return None without error
        result = validate_merged_weights([], tracker, col_map)
        assert result is None
