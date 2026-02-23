"""Tests for sheet_detect — FR-004, FR-005, FR-006: Sheet detection by regex patterns.

Every test creates an in-memory openpyxl Workbook with specified sheet names
and uses an AppConfig with compiled patterns. Pattern matching is case-insensitive
and strips whitespace from sheet names.
"""

import re
from pathlib import Path

import pytest
from openpyxl import Workbook

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.models import AppConfig, FieldPattern
from autoconvert.sheet_detect import detect_sheets

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def invoice_packing_config() -> AppConfig:
    """Create an AppConfig with simple invoice/packing sheet patterns."""
    return AppConfig(
        invoice_sheet_patterns=[re.compile(r"^invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"^packing", re.IGNORECASE)],
        invoice_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        packing_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=Path("dummy.xlsx"),
    )


@pytest.fixture
def multi_pattern_config() -> AppConfig:
    """Create an AppConfig with multiple invoice and packing patterns.

    Invoice patterns: ^invoice, ^bill
    Packing patterns: ^packing, ^dn&pl$
    """
    return AppConfig(
        invoice_sheet_patterns=[
            re.compile(r"^invoice", re.IGNORECASE),
            re.compile(r"^bill", re.IGNORECASE),
        ],
        packing_sheet_patterns=[
            re.compile(r"^packing", re.IGNORECASE),
            re.compile(r"^dn&pl$", re.IGNORECASE),
        ],
        invoice_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        packing_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=Path("dummy.xlsx"),
    )


# ---------------------------------------------------------------------------
# Happy Path Tests
# ---------------------------------------------------------------------------


def test_detect_sheets_invoice_matched_by_pattern(invoice_packing_config: AppConfig):
    """Test detect_sheets when sheet name 'INVOICE' matches ^invoice pattern (case-insensitive).

    Returns correct pair of (invoice_sheet, packing_sheet).
    """
    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "INVOICE"

    wb.create_sheet("Packing Data")

    invoice, packing = detect_sheets(wb, invoice_packing_config)

    assert invoice.title == "INVOICE"
    assert packing.title == "Packing Data"


def test_detect_sheets_packing_matched_by_dn_pl(multi_pattern_config: AppConfig):
    """Test detect_sheets when sheet name 'DN&PL' matches ^dn&pl$ pattern.

    Returns correct packing sheet.
    """
    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "Invoice"

    wb.create_sheet("DN&PL")

    invoice, packing = detect_sheets(wb, multi_pattern_config)

    assert invoice.title == "Invoice"
    assert packing.title == "DN&PL"


def test_detect_sheets_multiple_matching_sheets_first_wins(
    invoice_packing_config: AppConfig,
):
    """Test detect_sheets when multiple sheets match — first matching sheet wins for each type.

    Has two invoice-like sheets and two packing-like sheets.
    Should return the first of each that matches.
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Invoice"

    wb.create_sheet("Invoice Data")
    wb.create_sheet("Packing")
    wb.create_sheet("Packing Details")

    invoice, packing = detect_sheets(wb, invoice_packing_config)

    # First invoice match is "Invoice"
    assert invoice.title == "Invoice"
    # First packing match is "Packing"
    assert packing.title == "Packing"


def test_detect_sheets_ignores_unrecognized_sheets(invoice_packing_config: AppConfig):
    """Test detect_sheets ignores unrecognized sheets like 'Purchase Contract', 'Lookup'.

    Correct pair still returned.
    """
    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "Invoice"

    wb.create_sheet("Packing")
    wb.create_sheet("Purchase Contract")
    wb.create_sheet("Lookup")

    invoice, packing = detect_sheets(wb, invoice_packing_config)

    assert invoice.title == "Invoice"
    assert packing.title == "Packing"


# ---------------------------------------------------------------------------
# Error Cases
# ---------------------------------------------------------------------------


def test_detect_sheets_invoice_not_found_raises_err012(
    invoice_packing_config: AppConfig,
):
    """Test detect_sheets raises ProcessingError ERR_012 when invoice sheet not found.

    Workbook has only "Packing" sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing"

    with pytest.raises(ProcessingError) as exc_info:
        detect_sheets(wb, invoice_packing_config)

    assert exc_info.value.code == ErrorCode.ERR_012
    assert "Invoice sheet not found" in exc_info.value.message


def test_detect_sheets_packing_not_found_raises_err013(
    invoice_packing_config: AppConfig,
):
    """Test detect_sheets raises ProcessingError ERR_013 when packing sheet not found.

    Workbook has only "Invoice" sheet.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    with pytest.raises(ProcessingError) as exc_info:
        detect_sheets(wb, invoice_packing_config)

    assert exc_info.value.code == ErrorCode.ERR_013
    assert "Packing sheet not found" in exc_info.value.message


def test_detect_sheets_both_missing_raises_err012_first(
    invoice_packing_config: AppConfig,
):
    """Test detect_sheets raises ERR_012 (invoice) before ERR_013 when both missing.

    Only "Other" sheet present.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Other"

    with pytest.raises(ProcessingError) as exc_info:
        detect_sheets(wb, invoice_packing_config)

    # Invoice check happens first, so ERR_012 is raised
    assert exc_info.value.code == ErrorCode.ERR_012


# ---------------------------------------------------------------------------
# Edge Cases: Whitespace Handling
# ---------------------------------------------------------------------------


def test_detect_sheets_strips_whitespace_from_sheet_name(
    invoice_packing_config: AppConfig,
):
    """Test detect_sheets strips leading/trailing whitespace from sheet names before matching.

    Sheet name is '  Invoice  ' (with spaces), should match ^invoice pattern.
    """
    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "  Invoice  "

    wb.create_sheet("  Packing  ")

    invoice, packing = detect_sheets(wb, invoice_packing_config)

    assert invoice.title == "  Invoice  "
    assert packing.title == "  Packing  "


# ---------------------------------------------------------------------------
# Edge Cases: Case-Insensitivity
# ---------------------------------------------------------------------------


def test_detect_sheets_case_insensitive_matching():
    """Test detect_sheets matches sheet names case-insensitively.

    Sheet names are mixed case: 'InVoIcE' and 'PaCkInG'.
    """
    config = AppConfig(
        invoice_sheet_patterns=[re.compile(r"^invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"^packing", re.IGNORECASE)],
        invoice_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        packing_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=Path("dummy.xlsx"),
    )

    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "InVoIcE"

    wb.create_sheet("PaCkInG")

    invoice, packing = detect_sheets(wb, config)

    assert invoice.title == "InVoIcE"
    assert packing.title == "PaCkInG"


# ---------------------------------------------------------------------------
# Edge Cases: Complex Patterns
# ---------------------------------------------------------------------------


def test_detect_sheets_pattern_with_anchors_and_special_chars(
    multi_pattern_config: AppConfig,
):
    """Test detect_sheets matches pattern ^dn&pl$ with special char & and anchors.

    Sheet 'DN&PL' should match ^dn&pl$ (exactly), not 'DN&PL Extra'.
    """
    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "Invoice"

    wb.create_sheet("DN&PL")
    wb.create_sheet("DN&PL Extra")  # Should not match because of $ anchor

    invoice, packing = detect_sheets(wb, multi_pattern_config)

    assert invoice.title == "Invoice"
    assert packing.title == "DN&PL"


def test_detect_sheets_multiple_patterns_fallback():
    """Test detect_sheets uses second pattern when first doesn't match.

    Config has ^invoice and ^bill patterns for invoices.
    Sheet is named 'Bill', should match second pattern.
    """
    config = AppConfig(
        invoice_sheet_patterns=[
            re.compile(r"^invoice", re.IGNORECASE),
            re.compile(r"^bill", re.IGNORECASE),
        ],
        packing_sheet_patterns=[re.compile(r"^packing", re.IGNORECASE)],
        invoice_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        packing_columns={
            "part_no": FieldPattern(patterns=["^part"], type="string", required=True),
        },
        inv_no_patterns=[],
        inv_no_label_patterns=[],
        inv_no_exclude_patterns=[],
        currency_lookup={},
        country_lookup={},
        template_path=Path("dummy.xlsx"),
    )

    wb = Workbook()
    ws_invoice = wb.active
    ws_invoice.title = "Bill"

    wb.create_sheet("Packing")

    invoice, packing = detect_sheets(wb, config)

    assert invoice.title == "Bill"
    assert packing.title == "Packing"
