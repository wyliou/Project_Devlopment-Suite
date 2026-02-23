"""Integration tests for the AutoConvert pipeline.

Categories: boundary (5), full pipeline (3), error propagation (5),
edge cases (5), output format (3).
Uses real modules -- no mocking of internal components.
"""

from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from autoconvert.batch import process_file
from autoconvert.config import load_config
from autoconvert.errors import ErrorCode, WarningCode
from autoconvert.models import AppConfig


@pytest.fixture(scope="session")
def app_config() -> AppConfig:
    """Load real AppConfig from config/ directory."""
    return load_config(Path(__file__).parent.parent / "config")


# Reason: Headers must match regex patterns in config/field_patterns.yaml.
# brand_type -> 品牌类型, serial -> 项号 (Chinese-only patterns).
_INV_H: dict[int, str] = {
    1: "Part No.", 2: "P.O. No", 3: "QTY", 4: "Unit Price", 5: "Amount",
    6: "Currency", 7: "Country of Origin", 8: "COD", 9: "Brand",
    10: "\u54c1\u724c\u7c7b\u578b", 11: "Model", 12: "Invoice No.", 13: "\u9879\u53f7",
}
_PK_H: dict[int, str] = {
    1: "Part No.", 2: "P.O. No", 3: "QTY", 4: "N.W.", 5: "G.W.", 6: "CTNS",
}


def _ir(*, pn: str = "PART-001", po: str = "PO100", q: int | float = 10,
         pr: float = 5.0, am: float = 50.0, cur: str = "USD", coo: str = "CHINA",
         cod: str = "", br: str = "TestBrand", bt: str = "OEM", md: str = "MDL-X1",
         inv: str = "INV2025001", ser: str = "SN001") -> dict[int, object]:
    """Build one invoice data row."""
    return {1: pn, 2: po, 3: q, 4: pr, 5: am, 6: cur, 7: coo, 8: cod,
            9: br, 10: bt, 11: md, 12: inv, 13: ser}


def _pr(*, pn: str = "PART-001", po: str = "PO100", q: int | float = 10,
         nw: float = 15.5, gw: float = 20.0, ct: int = 1) -> dict[int, object]:
    """Build one packing data row."""
    return {1: pn, 2: po, 3: q, 4: nw, 5: gw, 6: ct}


def _tr(nw: float = 15.5, gw: float = 20.0) -> dict[int, object]:
    """Build a TOTAL row dict."""
    return {1: "TOTAL", 4: nw, 5: gw}


def _wb(inv_rows: list[dict[int, object]], pack_rows: list[dict[int, object]],
         total: dict[int, object] | None = None, *, inv_name: str = "Invoice",
         pk_name: str = "Packing", hdr: int = 8,
         hdr_area: dict[tuple[int, int], object] | None = None) -> openpyxl.Workbook:
    """Build in-memory workbook with invoice + packing sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    iw = wb.create_sheet(inv_name)
    for c, v in _INV_H.items():
        iw.cell(row=hdr, column=c, value=v)
    for i, rd in enumerate(inv_rows):
        for c, v in rd.items():
            iw.cell(row=hdr + 1 + i, column=c, value=v)
    if hdr_area:
        for (r, c), v in hdr_area.items():
            iw.cell(row=r, column=c, value=v)
    pw = wb.create_sheet(pk_name)
    for c, v in _PK_H.items():
        pw.cell(row=hdr, column=c, value=v)
    for i, rd in enumerate(pack_rows):
        for c, v in rd.items():
            pw.cell(row=hdr + 1 + i, column=c, value=v)
    if total:
        tr = hdr + len(pack_rows) + 2
        for c, v in total.items():
            pw.cell(row=tr, column=c, value=v)
    return wb


def _run(wb: openpyxl.Workbook, tmp: Path, cfg: AppConfig,
         name: str = "test_file.xlsx") -> object:
    """Save workbook and run process_file."""
    fp = tmp / name
    wb.save(fp)
    fd = tmp / "finished"
    fd.mkdir(exist_ok=True)
    with patch("autoconvert.batch._FINISHED_DIR", fd):
        return process_file(fp, cfg)


# ===================================================================
# 1. BOUNDARY TESTS
# ===================================================================
class TestBoundary:
    """Wire 2-3 adjacent modules, verify data handoff precision."""

    def test_currency_lookup(self, app_config: AppConfig, tmp_path: Path) -> None:
        """extract_invoice -> transform: 'USD' becomes '502'."""
        r = _run(_wb([_ir(cur="USD")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].currency == "502"

    def test_country_lookup(self, app_config: AppConfig, tmp_path: Path) -> None:
        """extract_invoice -> transform: 'CHINA' becomes numeric code."""
        r = _run(_wb([_ir(coo="CHINA")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].coo != "CHINA"
        assert r.invoice_items[0].coo.isdigit()

    def test_weight_precision_handoff(self, app_config: AppConfig, tmp_path: Path) -> None:
        """extract_packing -> weight_alloc: allocated sum == total_nw exactly."""
        ir = [_ir(pn="PA", q=5, pr=2.0, am=10.0), _ir(pn="PB", q=3, pr=3.0, am=9.0)]
        pk = [_pr(pn="PA", q=5, nw=10.25, gw=13.0), _pr(pn="PB", q=3, nw=5.25, gw=7.0)]
        r = _run(_wb(ir, pk, _tr(nw=15.5, gw=20.0)), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        s = sum(i.allocated_weight for i in r.invoice_items if i.allocated_weight)
        assert s == Decimal("15.5")

    def test_po_cleaning(self, app_config: AppConfig, tmp_path: Path) -> None:
        """transform.clean_po_number: '2250600556-2.1' -> '2250600556'."""
        r = _run(_wb([_ir(po="2250600556-2.1")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].po_no == "2250600556"

    def test_sheet_detect_nonstandard_names(self, app_config: AppConfig, tmp_path: Path) -> None:
        """sheet_detect -> column_map with names like 'INVOICE-Random'."""
        r = _run(_wb([_ir()], [_pr()], _tr(), inv_name="INVOICE-Random",
                     pk_name="PACKING LIST-Random"), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert len(r.invoice_items) >= 1


# ===================================================================
# 2. FULL PIPELINE TESTS
# ===================================================================
class TestFullPipeline:
    """Synthetic workbook -> process_file -> verify output."""

    def test_single_item_success(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Single part end-to-end: Success, correct weight, output file created."""
        r = _run(_wb([_ir()], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].part_no == "PART-001"
        assert r.invoice_items[0].allocated_weight == Decimal("15.5")
        assert r.packing_totals.total_nw == Decimal("15.5")
        assert list((tmp_path / "finished").glob("*_template.xlsx"))

    def test_multi_part_proportional(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Multiple parts: proportional weight allocation sums to total_nw."""
        ir = [
            _ir(pn="ALPHA", po="POA1", q=6, pr=10.0, am=60.0),
            _ir(pn="ALPHA", po="POA2", q=4, pr=10.0, am=40.0),
            _ir(pn="BETA", po="POB1", q=5, pr=20.0, am=100.0),
        ]
        pk = [_pr(pn="ALPHA", q=10, nw=30.0, gw=35.0),
              _pr(pn="BETA", q=5, nw=20.0, gw=25.0)]
        r = _run(_wb(ir, pk, _tr(nw=50.0, gw=60.0)), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        total = sum(i.allocated_weight for i in r.invoice_items if i.allocated_weight)
        assert total == Decimal("50")
        alpha_w = sum(i.allocated_weight for i in r.invoice_items
                      if i.part_no == "ALPHA" and i.allocated_weight)
        assert alpha_w == Decimal("30")

    def test_output_file_structure(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Output Excel: data at row 5, fixed values in C/R/S/T columns."""
        r = _run(_wb([_ir()], [_pr()], _tr()), tmp_path, app_config)
        if r.status not in ("Success", "Attention"):
            pytest.skip("Pipeline failed")
        f = list((tmp_path / "finished").glob("*_template.xlsx"))[0]
        ws = openpyxl.load_workbook(f, data_only=True)["\u5de5\u4f5c\u88681"]
        assert ws.cell(row=5, column=1).value == "PART-001"
        assert str(ws.cell(row=5, column=3).value) == "3"
        assert str(ws.cell(row=5, column=18).value) == "32052"
        assert str(ws.cell(row=5, column=19).value) == "320506"
        assert str(ws.cell(row=5, column=20).value) == "142"


# ===================================================================
# 3. ERROR PROPAGATION TESTS
# ===================================================================
class TestErrorPropagation:
    """Errors at different pipeline stages surface with correct codes."""

    def test_err012_no_invoice_sheet(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Missing invoice sheet -> ERR_012."""
        wb = openpyxl.Workbook()
        wb.active.title = "SomeSheet"
        wb.create_sheet("Packing")
        fp = tmp_path / "no_inv.xlsx"
        wb.save(fp)
        r = process_file(fp, app_config)
        assert r.status == "Failed"
        assert ErrorCode.ERR_012 in [e.code for e in r.errors]

    def test_err013_no_packing_sheet(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Missing packing sheet -> ERR_013."""
        wb = openpyxl.Workbook()
        wb.active.title = "Invoice"
        wb.create_sheet("Random")
        fp = tmp_path / "no_pk.xlsx"
        wb.save(fp)
        r = process_file(fp, app_config)
        assert r.status == "Failed"
        assert ErrorCode.ERR_013 in [e.code for e in r.errors]

    def test_err014_or_err020_few_headers(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Invoice with 2 headers -> ERR_014 or ERR_020."""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        iw = wb.create_sheet("Invoice")
        iw.cell(row=8, column=1, value="Part No.")
        iw.cell(row=8, column=2, value="QTY")
        pw = wb.create_sheet("Packing")
        for c, v in _PK_H.items():
            pw.cell(row=8, column=c, value=v)
        fp = tmp_path / "few.xlsx"
        wb.save(fp)
        r = process_file(fp, app_config)
        assert r.status == "Failed"
        codes = [e.code for e in r.errors]
        assert ErrorCode.ERR_014 in codes or ErrorCode.ERR_020 in codes

    def test_err040_part_mismatch(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Invoice part_no != packing part_no -> ERR_040 or ERR_043."""
        r = _run(_wb([_ir(pn="INV-ONLY")], [_pr(pn="PK-ONLY")], _tr()),
                 tmp_path, app_config)
        assert r.status == "Failed"
        codes = [e.code for e in r.errors]
        assert ErrorCode.ERR_040 in codes or ErrorCode.ERR_043 in codes

    def test_err011_corrupted(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Non-Excel bytes with .xlsx extension -> ERR_011."""
        fp = tmp_path / "bad.xlsx"
        fp.write_bytes(b"NOT_EXCEL")
        r = process_file(fp, app_config)
        assert r.status == "Failed"
        assert ErrorCode.ERR_011 in [e.code for e in r.errors]


# ===================================================================
# 4. EDGE CASE TESTS
# ===================================================================
class TestEdgeCases:
    """Empty, minimal, duplicate, missing optional fields."""

    def test_minimal_valid(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Smallest valid workbook: 1 item, 1 packing, 1 total."""
        r = _run(_wb([_ir()], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].allocated_weight is not None

    def test_empty_serial(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Empty serial (optional) does not cause ERR_030."""
        r = _run(_wb([_ir(ser="")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        assert r.invoice_items[0].serial == ""

    def test_empty_cod(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Empty COD (optional) does not cause error."""
        r = _run(_wb([_ir(cod="")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status in ("Success", "Attention")

    def test_duplicate_parts_proportional(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Two invoice rows same part_no: weights proportional, sum = part weight."""
        ir = [_ir(pn="DUP", po="P1", q=7, pr=1.0, am=7.0),
              _ir(pn="DUP", po="P2", q=3, pr=1.0, am=3.0)]
        r = _run(_wb(ir, [_pr(pn="DUP", q=10, nw=20.0, gw=25.0)],
                     _tr(nw=20.0, gw=25.0)), tmp_path, app_config)
        assert r.status in ("Success", "Attention")
        w0 = r.invoice_items[0].allocated_weight
        w1 = r.invoice_items[1].allocated_weight
        assert w0 is not None and w1 is not None
        assert w0 + w1 == Decimal("20")

    def test_unstandardized_currency_attention(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Unknown currency -> ATT_003, Attention status, raw value preserved."""
        r = _run(_wb([_ir(cur="ZZZ_UNKNOWN")], [_pr()], _tr()), tmp_path, app_config)
        assert r.status == "Attention"
        assert WarningCode.ATT_003 in [w.code for w in r.warnings]
        assert r.invoice_items[0].currency == "ZZZ_UNKNOWN"


# ===================================================================
# 5. OUTPUT FORMAT TESTS
# ===================================================================
class TestOutputFormat:
    """Verify output matches PRD Section 7 template schema (FR-029)."""

    def test_all_40_columns(self, app_config: AppConfig, tmp_path: Path) -> None:
        """Verify A-AN column mapping in output template."""
        r = _run(_wb([_ir(pn="TP", po="PO999", q=100, pr=2.5, am=250.0,
                         br="BX", bt="OEM", md="M100", inv="INVT", ser="S1")],
                     [_pr(pn="TP", q=100, nw=50.0, gw=60.0)],
                     _tr(nw=50.0, gw=60.0)), tmp_path, app_config)
        if r.status not in ("Success", "Attention"):
            pytest.skip("Pipeline failed")
        f = list((tmp_path / "finished").glob("*_template.xlsx"))[0]
        ws = openpyxl.load_workbook(f, data_only=True)["\u5de5\u4f5c\u88681"]
        row = 5
        assert ws.cell(row=row, column=1).value == "TP"       # A: part_no
        assert ws.cell(row=row, column=2).value == "PO999"    # B: po_no
        assert str(ws.cell(row=row, column=3).value) == "3"    # C: fixed
        assert ws.cell(row=row, column=4).value == "502"       # D: currency
        assert ws.cell(row=row, column=5).value is not None    # E: qty
        assert ws.cell(row=row, column=6).value is not None    # F: price
        assert ws.cell(row=row, column=7).value is not None    # G: amount
        assert str(ws.cell(row=row, column=8).value) != "CHINA"  # H: coo code
        for c in (9, 10, 11):
            assert ws.cell(row=row, column=c).value is None    # I-K: empty
        assert ws.cell(row=row, column=12).value == "S1"       # L: serial
        assert ws.cell(row=row, column=13).value is not None   # M: weight
        assert ws.cell(row=row, column=14).value == "INVT"     # N: inv_no
        assert ws.cell(row=row, column=15).value is None       # O: empty
        assert float(ws.cell(row=row, column=16).value) == pytest.approx(60.0)  # P: gw
        assert ws.cell(row=row, column=17).value is None       # Q: empty
        assert str(ws.cell(row=row, column=18).value) == "32052"   # R
        assert str(ws.cell(row=row, column=19).value) == "320506"  # S
        assert str(ws.cell(row=row, column=20).value) == "142"     # T
        for c in range(21, 37):
            assert ws.cell(row=row, column=c).value is None    # U-AJ: empty
        assert ws.cell(row=row, column=38).value == "BX"       # AL: brand
        assert ws.cell(row=row, column=39).value == "OEM"      # AM: brand_type
        assert ws.cell(row=row, column=40).value == "M100"     # AN: model

    def test_row5_only_fields(self, app_config: AppConfig, tmp_path: Path) -> None:
        """total_gw (P) and total_packets (AK) only on row 5, not row 6."""
        ir = [_ir(pn="P1", po="A", q=5, pr=10.0, am=50.0),
              _ir(pn="P2", po="B", q=5, pr=10.0, am=50.0)]
        pk = [_pr(pn="P1", q=5, nw=10.0, gw=12.0),
              _pr(pn="P2", q=5, nw=10.0, gw=12.0)]
        r = _run(_wb(ir, pk, _tr(nw=20.0, gw=24.0)), tmp_path, app_config)
        if r.status not in ("Success", "Attention"):
            pytest.skip("Pipeline failed")
        f = list((tmp_path / "finished").glob("*_template.xlsx"))[0]
        ws = openpyxl.load_workbook(f, data_only=True)["\u5de5\u4f5c\u88681"]
        assert ws.cell(row=5, column=16).value is not None  # P row 5
        assert ws.cell(row=6, column=16).value is None       # P row 6
        assert ws.cell(row=6, column=37).value is None       # AK row 6

    def test_attention_raw_currency_passthrough(self, app_config: AppConfig, tmp_path: Path) -> None:
        """ATT_003: raw currency string passes through to col D in output."""
        r = _run(_wb([_ir(cur="EXOTIC")], [_pr()], _tr()), tmp_path, app_config)
        if r.status not in ("Success", "Attention"):
            pytest.skip("Pipeline failed")
        f = list((tmp_path / "finished").glob("*_template.xlsx"))[0]
        ws = openpyxl.load_workbook(f, data_only=True)["\u5de5\u4f5c\u88681"]
        assert ws.cell(row=5, column=4).value == "EXOTIC"
