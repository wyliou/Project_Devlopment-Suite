"""Tests for merge_tracker — FR-010 acceptance criteria.

Every test creates an in-memory openpyxl ``Workbook``, sets up merges and cell
values, then exercises ``MergeTracker`` methods.  No on-disk fixtures required.
"""

from decimal import Decimal

from openpyxl import Workbook

from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import MergeRange

# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------


def _make_sheet_with_vertical_merge():
    """Create a workbook with a vertical merge A5:A7 = 'PartA'.

    Returns:
        The active worksheet (already has the merge applied).
    """
    wb = Workbook()
    ws = wb.active
    ws.cell(row=5, column=1, value="PartA")
    ws.merge_cells("A5:A7")
    return ws


# ---------------------------------------------------------------------------
# __init__ — construction and unmerge
# ---------------------------------------------------------------------------


def test_merge_tracker_captures_ranges_before_unmerge():
    """Test __init__ captures 1 merge range and unmerges the sheet."""
    ws = _make_sheet_with_vertical_merge()

    tracker = MergeTracker(ws)

    # Exactly one range captured
    assert len(tracker._ranges) == 1
    mr = tracker._ranges[0]
    assert mr.min_row == 5
    assert mr.max_row == 7
    assert mr.min_col == 1
    assert mr.max_col == 1
    assert mr.value == "PartA"

    # Sheet is now unmerged
    assert len(list(ws.merged_cells.ranges)) == 0


def test_merge_tracker_non_anchor_cell_is_none_after_unmerge():
    """Test that non-anchor cells become None after unmerge."""
    ws = _make_sheet_with_vertical_merge()

    MergeTracker(ws)

    # Reason: openpyxl sets non-anchor cells to None after unmerge.
    assert ws.cell(row=6, column=1).value is None
    assert ws.cell(row=7, column=1).value is None
    # Anchor retains its value.
    assert ws.cell(row=5, column=1).value == "PartA"


def test_merge_tracker_horizontal_merge_captured():
    """Test __init__ captures a horizontal merge B3:C3."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=2, value="Brand")
    ws.merge_cells("B3:C3")

    tracker = MergeTracker(ws)

    assert len(tracker._ranges) == 1
    mr = tracker._ranges[0]
    assert mr.min_row == 3
    assert mr.max_row == 3
    assert mr.min_col == 2
    assert mr.max_col == 3
    assert mr.value == "Brand"


def test_merge_tracker_no_merges():
    """Test __init__ on a sheet with no merges — no exception, queries false."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=5, column=2, value="hello")

    tracker = MergeTracker(ws)

    assert len(tracker._ranges) == 0
    assert tracker.is_in_merge(5, 2) is False


# ---------------------------------------------------------------------------
# is_merge_anchor and is_in_merge
# ---------------------------------------------------------------------------


def test_is_merge_anchor_true_for_top_left():
    """Test is_merge_anchor returns True for the anchor cell."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.is_merge_anchor(5, 1) is True


def test_is_merge_anchor_false_for_non_anchor():
    """Test is_merge_anchor returns False for a non-anchor cell inside merge."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.is_merge_anchor(6, 1) is False


def test_is_in_merge_true_for_all_cells():
    """Test is_in_merge returns True for every cell in the merge range."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.is_in_merge(5, 1) is True
    assert tracker.is_in_merge(6, 1) is True
    assert tracker.is_in_merge(7, 1) is True


def test_is_in_merge_false_outside_range():
    """Test is_in_merge returns False for a cell outside the merge range."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.is_in_merge(8, 1) is False
    assert tracker.is_in_merge(4, 1) is False
    assert tracker.is_in_merge(5, 2) is False


# ---------------------------------------------------------------------------
# get_merge_range and get_anchor_value
# ---------------------------------------------------------------------------


def test_get_merge_range_returns_correct_range():
    """Test get_merge_range returns the correct MergeRange from a non-anchor."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    mr = tracker.get_merge_range(6, 1)
    assert mr is not None
    assert mr == MergeRange(min_row=5, max_row=7, min_col=1, max_col=1, value="PartA")


def test_get_merge_range_returns_none_outside():
    """Test get_merge_range returns None for a cell outside any merge."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.get_merge_range(8, 1) is None


def test_get_anchor_value_from_non_anchor_row():
    """Test get_anchor_value returns the anchor value from a non-anchor row."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.get_anchor_value(6, 1) == "PartA"


def test_get_anchor_value_from_anchor_row():
    """Test get_anchor_value returns the anchor value from the anchor row."""
    ws = _make_sheet_with_vertical_merge()
    tracker = MergeTracker(ws)

    assert tracker.get_anchor_value(5, 1) == "PartA"


# ---------------------------------------------------------------------------
# is_data_area_merge and get_string_value
# ---------------------------------------------------------------------------


def test_is_data_area_merge_true_when_starts_after_header():
    """Test is_data_area_merge returns True for merge starting after header."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=10, column=1, value="PartX")
    ws.merge_cells("A10:A12")

    tracker = MergeTracker(ws)

    assert tracker.is_data_area_merge(11, 1, header_row=5) is True


def test_is_data_area_merge_false_for_header_area():
    """Test is_data_area_merge returns False for merge spanning header area."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=1, value="Header")
    ws.merge_cells("A3:A5")

    tracker = MergeTracker(ws)

    # min_row=3 <= header_row=5, so this is header-area
    assert tracker.is_data_area_merge(4, 1, header_row=5) is False


def test_get_string_value_data_area_returns_anchor():
    """Test get_string_value propagates anchor value for data-area merges."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=10, column=1, value="PartX")
    ws.merge_cells("A10:A12")

    tracker = MergeTracker(ws)

    # Non-anchor row in data area should get the anchor value
    assert tracker.get_string_value(11, 1, header_row=5) == "PartX"
    assert tracker.get_string_value(12, 1, header_row=5) == "PartX"
    # Anchor row itself
    assert tracker.get_string_value(10, 1, header_row=5) == "PartX"


def test_get_string_value_non_merged_returns_cell():
    """Test get_string_value returns sheet cell value for non-merged cells."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=8, column=2, value="USD")

    tracker = MergeTracker(ws)

    assert tracker.get_string_value(8, 2, header_row=5) == "USD"


# ---------------------------------------------------------------------------
# get_weight_value
# ---------------------------------------------------------------------------


def test_get_weight_value_anchor_returns_value():
    """Test get_weight_value returns the captured value for the anchor row."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=10, column=1, value=2.5)
    ws.merge_cells("A10:A12")

    tracker = MergeTracker(ws)

    result = tracker.get_weight_value(10, 1, header_row=5)
    assert result == 2.5


def test_get_weight_value_non_anchor_returns_zero():
    """Test get_weight_value returns Decimal('0.0') for non-anchor rows."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=10, column=1, value=2.5)
    ws.merge_cells("A10:A12")

    tracker = MergeTracker(ws)

    assert tracker.get_weight_value(11, 1, header_row=5) == Decimal("0.0")


def test_get_weight_value_non_anchor_returns_zero_for_third_row():
    """Test get_weight_value returns Decimal('0.0') for the third row of merge."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=10, column=1, value=2.5)
    ws.merge_cells("A10:A12")

    tracker = MergeTracker(ws)

    assert tracker.get_weight_value(12, 1, header_row=5) == Decimal("0.0")


def test_get_weight_value_header_area_merge_returns_zero():
    """Test get_weight_value returns Decimal('0.0') for header-area merges."""
    wb = Workbook()
    ws = wb.active
    ws.cell(row=3, column=1, value=100)
    ws.merge_cells("A3:A5")

    tracker = MergeTracker(ws)

    # min_row=3 <= header_row=5, so this is header-area — returns zero
    assert tracker.get_weight_value(4, 1, header_row=5) == Decimal("0.0")
    # Even the anchor row of a header-area merge returns zero
    assert tracker.get_weight_value(3, 1, header_row=5) == Decimal("0.0")
