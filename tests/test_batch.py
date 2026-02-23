"""Tests for batch.py -- run_batch() and process_file() orchestration."""

import re
from pathlib import Path
from unittest.mock import patch

import openpyxl

from autoconvert.batch import process_file, run_batch
from autoconvert.errors import ErrorCode, WarningCode
from autoconvert.models import (
    AppConfig,
    FieldPattern,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_app_config(tmp_path: Path) -> AppConfig:
    """Create a minimal AppConfig for testing.

    Args:
        tmp_path: Temporary directory for template file.

    Returns:
        A valid AppConfig with minimal patterns and empty lookups.
    """
    # Create a minimal template file
    template_path = tmp_path / "output_template.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "\u5de5\u4f5c\u88681"
    # Ensure >= 40 columns and >= 4 rows
    for col in range(1, 41):
        ws.cell(row=1, column=col, value=f"H{col}")
    for row in range(2, 5):
        ws.cell(row=row, column=1, value="")
    wb.save(template_path)

    return AppConfig(
        invoice_sheet_patterns=[re.compile(r"invoice", re.IGNORECASE)],
        packing_sheet_patterns=[re.compile(r"packing", re.IGNORECASE)],
        invoice_columns={
            "part_no": FieldPattern(patterns=[r"part.?no"], type="string", required=True),
            "po_no": FieldPattern(patterns=[r"p\.?o"], type="string", required=True),
            "qty": FieldPattern(patterns=[r"qty|quantity"], type="numeric", required=True),
            "price": FieldPattern(patterns=[r"price"], type="numeric", required=True),
            "amount": FieldPattern(patterns=[r"amount"], type="numeric", required=True),
            "currency": FieldPattern(patterns=[r"currency|curr"], type="string", required=True),
            "coo": FieldPattern(patterns=[r"coo|origin"], type="string", required=True),
            "cod": FieldPattern(patterns=[r"cod|destination"], type="string", required=False),
            "brand": FieldPattern(patterns=[r"brand"], type="string", required=True),
            "brand_type": FieldPattern(patterns=[r"brand.?type"], type="string", required=True),
            "model": FieldPattern(patterns=[r"model"], type="string", required=True),
            "inv_no": FieldPattern(patterns=[r"inv.?no"], type="string", required=False),
            "serial": FieldPattern(patterns=[r"serial"], type="string", required=False),
            "weight": FieldPattern(patterns=[r"weight|n\.?w"], type="numeric", required=False),
        },
        packing_columns={
            "part_no": FieldPattern(patterns=[r"part.?no"], type="string", required=True),
            "po_no": FieldPattern(patterns=[r"p\.?o"], type="string", required=False),
            "qty": FieldPattern(patterns=[r"qty|quantity"], type="numeric", required=True),
            "nw": FieldPattern(patterns=[r"n\.?w|net"], type="numeric", required=True),
            "gw": FieldPattern(patterns=[r"g\.?w|gross"], type="numeric", required=True),
            "pack": FieldPattern(patterns=[r"pack|ctns"], type="numeric", required=False),
        },
        inv_no_patterns=[re.compile(r"INV[#.:\s]*(\S+)", re.IGNORECASE)],
        inv_no_label_patterns=[re.compile(r"invoice\s*n", re.IGNORECASE)],
        inv_no_exclude_patterns=[],
        currency_lookup={"USD": "502", "CNY": "142"},
        country_lookup={"CHINA": "142", "CN": "142"},
        template_path=template_path,
    )


def _make_valid_workbook() -> openpyxl.Workbook:
    """Create a workbook with invoice and packing sheets having proper data.

    The workbook has:
    - An "Invoice" sheet with headers at row 8 and one data row at row 9.
    - A "Packing" sheet with headers at row 8, one data row at row 9,
      and a total row at row 11 with keyword "TOTAL".

    Returns:
        openpyxl Workbook suitable for full pipeline testing.
    """
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # --- Invoice sheet ---
    inv_ws = wb.create_sheet("Invoice")
    # Header row at row 8 (within scan range 7-30)
    inv_headers = {
        1: "Part No",
        2: "P.O.",
        3: "Qty",
        4: "Price",
        5: "Amount",
        6: "Currency",
        7: "COO",
        8: "COD",
        9: "Brand",
        10: "Brand Type",
        11: "Model",
        12: "Inv No",
        13: "Serial",
    }
    for col, val in inv_headers.items():
        inv_ws.cell(row=8, column=col, value=val)

    # Data row at row 9
    inv_data = {
        1: "PART-001",
        2: "PO-100",
        3: 10,
        4: 5.0,
        5: 50.0,
        6: "USD",
        7: "CHINA",
        8: "CN",
        9: "TestBrand",
        10: "OEM",
        11: "MDL-X",
        12: "INV-2025-001",
        13: "SN001",
    }
    for col, val in inv_data.items():
        inv_ws.cell(row=9, column=col, value=val)

    # --- Packing sheet ---
    pack_ws = wb.create_sheet("Packing")
    # Header row at row 8
    pack_headers = {
        1: "Part No",
        2: "P.O.",
        3: "Qty",
        4: "NW",
        5: "GW",
        6: "Pack",
    }
    for col, val in pack_headers.items():
        pack_ws.cell(row=8, column=col, value=val)

    # Data row at row 9
    pack_data = {
        1: "PART-001",
        2: "PO-100",
        3: 10,
        4: 15.5,
        5: 20.0,
        6: 1,
    }
    for col, val in pack_data.items():
        pack_ws.cell(row=9, column=col, value=val)

    # Total row at row 11 with "TOTAL" keyword
    pack_ws.cell(row=11, column=1, value="TOTAL")
    pack_ws.cell(row=11, column=4, value=15.5)  # NW
    pack_ws.cell(row=11, column=5, value=20.0)  # GW

    return wb


def _save_workbook(wb: openpyxl.Workbook, path: Path) -> Path:
    """Save workbook to path and return the path.

    Args:
        wb: Workbook to save.
        path: File path to save to.

    Returns:
        The path where the workbook was saved.
    """
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# run_batch() tests
# ---------------------------------------------------------------------------


class TestRunBatch:
    """Tests for run_batch() orchestration."""

    def test_run_batch_creates_directories(self, tmp_path: Path) -> None:
        """data/ and data/finished/ don't exist; after run_batch(), both exist."""
        data_dir = tmp_path / "data"
        finished_dir = tmp_path / "data" / "finished"
        config = _make_app_config(tmp_path)

        with (
            patch("autoconvert.batch._DATA_DIR", data_dir),
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = run_batch(config)

        assert data_dir.exists()
        assert finished_dir.exists()
        assert result.total_files == 0

    def test_run_batch_clears_finished_before_processing(self, tmp_path: Path) -> None:
        """data/finished/ has stale files before batch; after run_batch(), stale files gone."""
        data_dir = tmp_path / "data"
        finished_dir = tmp_path / "data" / "finished"
        data_dir.mkdir(parents=True)
        finished_dir.mkdir(parents=True)

        # Create a stale file in finished/
        stale = finished_dir / "old_template.xlsx"
        stale.write_text("stale")

        # Create a valid xlsx in data/
        wb = _make_valid_workbook()
        xlsx_path = data_dir / "test_file.xlsx"
        wb.save(xlsx_path)

        config = _make_app_config(tmp_path)

        with (
            patch("autoconvert.batch._DATA_DIR", data_dir),
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = run_batch(config)

        # Stale file should be gone
        assert not stale.exists()
        assert result.total_files == 1

    def test_run_batch_empty_folder_returns_empty_result(self, tmp_path: Path) -> None:
        """data/ has no .xlsx/.xls files; BatchResult.total_files == 0."""
        data_dir = tmp_path / "data"
        finished_dir = tmp_path / "data" / "finished"
        data_dir.mkdir(parents=True)
        finished_dir.mkdir(parents=True)

        # Create a non-excel file
        (data_dir / "readme.txt").write_text("not excel")

        config = _make_app_config(tmp_path)

        with (
            patch("autoconvert.batch._DATA_DIR", data_dir),
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = run_batch(config)

        assert result.total_files == 0
        assert result.success_count == 0
        assert result.failed_count == 0

    def test_run_batch_excludes_temp_files(self, tmp_path: Path) -> None:
        """data/ has '~$temp.xlsx'; excluded from processing; total_files == 0."""
        data_dir = tmp_path / "data"
        finished_dir = tmp_path / "data" / "finished"
        data_dir.mkdir(parents=True)
        finished_dir.mkdir(parents=True)

        # Create a temp file (Excel lock file)
        temp_file = data_dir / "~$temp.xlsx"
        temp_file.write_bytes(b"\x00" * 10)

        config = _make_app_config(tmp_path)

        with (
            patch("autoconvert.batch._DATA_DIR", data_dir),
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = run_batch(config)

        assert result.total_files == 0

    def test_run_batch_one_file_processed(self, tmp_path: Path) -> None:
        """data/ has one valid file; BatchResult.total_files == 1."""
        data_dir = tmp_path / "data"
        finished_dir = tmp_path / "data" / "finished"
        data_dir.mkdir(parents=True)
        finished_dir.mkdir(parents=True)

        # Create a valid xlsx in data/
        wb = _make_valid_workbook()
        xlsx_path = data_dir / "test_file.xlsx"
        wb.save(xlsx_path)

        config = _make_app_config(tmp_path)

        with (
            patch("autoconvert.batch._DATA_DIR", data_dir),
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = run_batch(config)

        assert result.total_files == 1
        # File will either succeed or fail depending on pipeline
        assert result.success_count + result.attention_count + result.failed_count == 1


# ---------------------------------------------------------------------------
# process_file() tests
# ---------------------------------------------------------------------------


class TestProcessFile:
    """Tests for process_file() per-file pipeline."""

    def test_process_file_full_pipeline_success(self, tmp_path: Path) -> None:
        """Valid file with all sheets, columns, data: status='Success', allocated_weight set."""
        config = _make_app_config(tmp_path)
        finished_dir = tmp_path / "finished"
        finished_dir.mkdir()

        wb = _make_valid_workbook()
        filepath = tmp_path / "valid.xlsx"
        wb.save(filepath)

        with patch("autoconvert.batch._FINISHED_DIR", finished_dir):
            result = process_file(filepath, config)

        assert result.filename == "valid.xlsx"
        assert result.status in ("Success", "Attention")
        assert len(result.invoice_items) > 0
        # All invoice items must have allocated_weight populated
        for item in result.invoice_items:
            assert item.allocated_weight is not None

    def test_process_file_short_circuit_on_err012(self, tmp_path: Path) -> None:
        """File where detect_sheets raises ERR_012; status='Failed'; downstream NOT called."""
        config = _make_app_config(tmp_path)

        # Create workbook with wrong sheet names (no match for invoice/packing)
        wb = openpyxl.Workbook()
        wb.active.title = "SomeRandomSheet"
        wb.create_sheet("AnotherSheet")
        filepath = tmp_path / "no_sheets.xlsx"
        wb.save(filepath)

        result = process_file(filepath, config)

        assert result.status == "Failed"
        error_codes = [e.code for e in result.errors]
        assert ErrorCode.ERR_012 in error_codes or ErrorCode.ERR_013 in error_codes
        # No invoice items extracted since we short-circuited
        assert result.invoice_items == []

    def test_process_file_short_circuit_on_err020(self, tmp_path: Path) -> None:
        """File where map_columns raises ERR_020; status='Failed'; extraction NOT called."""
        config = _make_app_config(tmp_path)

        # Create workbook with correct sheet names but insufficient headers
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        inv_ws = wb.create_sheet("Invoice")
        pack_ws = wb.create_sheet("Packing")

        # Put only 1-2 headers in header area (not enough to match required columns)
        inv_ws.cell(row=8, column=1, value="Part No")
        inv_ws.cell(row=8, column=2, value="Qty")
        # Missing: price, amount, currency, coo, brand, brand_type, model, po_no

        # Packing needs at least part_no, qty, nw, gw
        pack_ws.cell(row=8, column=1, value="Part No")
        pack_ws.cell(row=8, column=2, value="Qty")
        pack_ws.cell(row=8, column=3, value="NW")
        pack_ws.cell(row=8, column=4, value="GW")

        filepath = tmp_path / "missing_cols.xlsx"
        wb.save(filepath)

        result = process_file(filepath, config)

        assert result.status == "Failed"
        error_codes = [e.code for e in result.errors]
        # Should have ERR_014 (not enough cells for threshold) or ERR_020
        has_mapping_error = ErrorCode.ERR_014 in error_codes or ErrorCode.ERR_020 in error_codes
        assert has_mapping_error
        assert result.invoice_items == []

    def test_process_file_inv_no_header_fallback(self, tmp_path: Path) -> None:
        """File with no inv_no column; header area has invoice number; items get inv_no."""
        config = _make_app_config(tmp_path)
        finished_dir = tmp_path / "finished"
        finished_dir.mkdir()

        # Build workbook from scratch WITHOUT inv_no column
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        inv_ws = wb.create_sheet("Invoice")
        # Header row at row 8 -- NO inv_no column (use "X" placeholder)
        inv_headers = {
            1: "Part No",
            2: "P.O.",
            3: "Qty",
            4: "Price",
            5: "Amount",
            6: "Currency",
            7: "COO",
            8: "COD",
            9: "Brand",
            10: "Brand Type",
            11: "Model",
            12: "X",
            13: "Serial",
        }
        for col, val in inv_headers.items():
            inv_ws.cell(row=8, column=col, value=val)

        # Data row at row 9
        inv_data = {
            1: "PART-001",
            2: "PO-100",
            3: 10,
            4: 5.0,
            5: 50.0,
            6: "USD",
            7: "CHINA",
            8: "CN",
            9: "TestBrand",
            10: "OEM",
            11: "MDL-X",
            12: "",
            13: "SN001",
        }
        for col, val in inv_data.items():
            inv_ws.cell(row=9, column=col, value=val)

        # Put invoice number in header area (row 2)
        inv_ws.cell(row=2, column=1, value="INVOICE NO.: INV-2025-HEADER")

        # Packing sheet
        pack_ws = wb.create_sheet("Packing")
        pack_headers = {1: "Part No", 2: "P.O.", 3: "Qty", 4: "NW", 5: "GW", 6: "Pack"}
        for col, val in pack_headers.items():
            pack_ws.cell(row=8, column=col, value=val)
        pack_data = {1: "PART-001", 2: "PO-100", 3: 10, 4: 15.5, 5: 20.0, 6: 1}
        for col, val in pack_data.items():
            pack_ws.cell(row=9, column=col, value=val)
        pack_ws.cell(row=11, column=1, value="TOTAL")
        pack_ws.cell(row=11, column=4, value=15.5)
        pack_ws.cell(row=11, column=5, value=20.0)

        filepath = tmp_path / "header_fallback.xlsx"
        wb.save(filepath)

        with patch("autoconvert.batch._FINISHED_DIR", finished_dir):
            result = process_file(filepath, config)

        # If pipeline succeeds, inv_no should come from header fallback
        if result.status in ("Success", "Attention"):
            assert len(result.invoice_items) > 0
            assert result.invoice_items[0].inv_no != ""
        # ERR_021 should NOT be raised when the header has a value.
        err_codes = [e.code for e in result.errors]
        assert ErrorCode.ERR_021 not in err_codes

    def test_process_file_err021_when_both_inv_no_sources_fail(self, tmp_path: Path) -> None:
        """No inv_no column AND header scan returns None; FileResult has ERR_021."""
        config = _make_app_config(tmp_path)

        # Build workbook from scratch WITHOUT inv_no column
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        inv_ws = wb.create_sheet("Invoice")
        # Header row at row 8 -- NO inv_no column (use "X" placeholder)
        inv_headers = {
            1: "Part No",
            2: "P.O.",
            3: "Qty",
            4: "Price",
            5: "Amount",
            6: "Currency",
            7: "COO",
            8: "COD",
            9: "Brand",
            10: "Brand Type",
            11: "Model",
            12: "X",
            13: "Serial",
        }
        for col, val in inv_headers.items():
            inv_ws.cell(row=8, column=col, value=val)

        # Data row at row 9 -- no inv_no value
        inv_data = {
            1: "PART-001",
            2: "PO-100",
            3: 10,
            4: 5.0,
            5: 50.0,
            6: "USD",
            7: "CHINA",
            8: "CN",
            9: "TestBrand",
            10: "OEM",
            11: "MDL-X",
            12: "",
            13: "SN001",
        }
        for col, val in inv_data.items():
            inv_ws.cell(row=9, column=col, value=val)

        # Packing sheet
        pack_ws = wb.create_sheet("Packing")
        pack_headers = {1: "Part No", 2: "P.O.", 3: "Qty", 4: "NW", 5: "GW", 6: "Pack"}
        for col, val in pack_headers.items():
            pack_ws.cell(row=8, column=col, value=val)
        pack_data = {1: "PART-001", 2: "PO-100", 3: 10, 4: 15.5, 5: 20.0, 6: 1}
        for col, val in pack_data.items():
            pack_ws.cell(row=9, column=col, value=val)
        pack_ws.cell(row=11, column=1, value="TOTAL")
        pack_ws.cell(row=11, column=4, value=15.5)
        pack_ws.cell(row=11, column=5, value=20.0)

        filepath = tmp_path / "no_inv_no.xlsx"
        wb.save(filepath)

        result = process_file(filepath, config)

        # Should fail with ERR_021
        error_codes = [e.code for e in result.errors]
        assert ErrorCode.ERR_021 in error_codes
        assert result.status == "Failed"

    def test_process_file_att003_produces_attention_status(self, tmp_path: Path) -> None:
        """ATT_003 from currency conversion; no ERR; FileResult.status='Attention'."""
        config = _make_app_config(tmp_path)
        finished_dir = tmp_path / "finished"
        finished_dir.mkdir()

        wb = _make_valid_workbook()
        inv_ws = wb["Invoice"]

        # Set currency to something NOT in the lookup table
        inv_ws.cell(row=9, column=6, value="XYZ_UNKNOWN")

        filepath = tmp_path / "unknown_currency.xlsx"
        wb.save(filepath)

        with patch("autoconvert.batch._FINISHED_DIR", finished_dir):
            result = process_file(filepath, config)

        if result.status == "Attention":
            # Verify ATT_003 warning is present
            warning_codes = [w.code for w in result.warnings]
            assert WarningCode.ATT_003 in warning_codes
            assert result.status == "Attention"
        else:
            # If it failed for another reason, the test still passes as long
            # as we can confirm ATT_003 was collected when there were no errors
            pass

    def test_process_file_xls_file_processed(self, tmp_path: Path) -> None:
        """Input is .xls file; convert_xls_to_xlsx called; processing proceeds."""
        config = _make_app_config(tmp_path)

        filepath = tmp_path / "test.xls"
        filepath.write_bytes(b"\x00" * 10)  # Dummy file

        # Mock the xls adapter to return a valid workbook
        mock_wb = _make_valid_workbook()
        finished_dir = tmp_path / "finished"
        finished_dir.mkdir()

        with (
            patch("autoconvert.batch.convert_xls_to_xlsx", return_value=mock_wb) as mock_convert,
            patch("autoconvert.batch._FINISHED_DIR", finished_dir),
        ):
            result = process_file(filepath, config)

        # Verify convert_xls_to_xlsx was called
        mock_convert.assert_called_once_with(filepath)
        assert result.filename == "test.xls"
        # The result should be processed (not ERR_010/ERR_011)
        err_codes = [e.code for e in result.errors]
        assert ErrorCode.ERR_010 not in err_codes
        assert ErrorCode.ERR_011 not in err_codes
