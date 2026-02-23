"""Tests for extract_invoice --- FR-011 acceptance criteria.

Every test creates an in-memory openpyxl ``Workbook``, sets up headers and data
rows, then exercises ``extract_invoice_items()``.  No on-disk fixtures required.
"""

from decimal import Decimal

import pytest
from openpyxl import Workbook

from autoconvert.errors import ErrorCode, ProcessingError
from autoconvert.extract_invoice import extract_invoice_items
from autoconvert.merge_tracker import MergeTracker
from autoconvert.models import ColumnMapping

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Standard field_map with all 13 fields, matching typical column layout.
# Columns: A=part_no(1), B=po_no(2), C=qty(3), D=price(4), E=amount(5),
#          F=currency(6), G=coo(7), H=brand(8), I=brand_type(9), J=model(10),
#          K=cod(11), L=inv_no(12), M=serial(13)
FULL_FIELD_MAP: dict[str, int] = {
    "part_no": 1,
    "po_no": 2,
    "qty": 3,
    "price": 4,
    "amount": 5,
    "currency": 6,
    "coo": 7,
    "brand": 8,
    "brand_type": 9,
    "model": 10,
    "cod": 11,
    "inv_no": 12,
    "serial": 13,
}


def _make_column_map(
    field_map: dict[str, int] | None = None,
    header_row: int = 1,
    effective_header_row: int = 1,
) -> ColumnMapping:
    """Create a ColumnMapping for testing.

    Args:
        field_map: Column index mapping. Defaults to FULL_FIELD_MAP.
        header_row: Header row number.
        effective_header_row: Effective header row number.

    Returns:
        A ColumnMapping instance.
    """
    return ColumnMapping(
        sheet_type="invoice",
        field_map=field_map or FULL_FIELD_MAP,
        header_row=header_row,
        effective_header_row=effective_header_row,
    )


def _build_invoice_sheet(
    header_row: int,
    data_rows: list[list],
    data_start_row: int | None = None,
) -> Workbook:
    """Create an in-memory workbook with an invoice sheet.

    Populates row ``header_row`` with header labels matching FULL_FIELD_MAP,
    then fills data rows starting at ``data_start_row`` (defaults to
    ``header_row + 1``).

    Args:
        header_row: 1-based row for headers.
        data_rows: List of lists, each containing values for columns A-M.
        data_start_row: 1-based row for first data row. Defaults to header_row + 1.

    Returns:
        The created Workbook (active sheet contains the data).
    """
    wb = Workbook()
    ws = wb.active

    # Write header labels.
    headers = [
        "PART NO.",
        "PO NO.",
        "QTY",
        "PRICE",
        "AMOUNT",
        "CURRENCY",
        "COO",
        "BRAND",
        "BRAND TYPE",
        "MODEL",
        "COD",
        "INV NO.",
        "SERIAL",
    ]
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col_idx, value=hdr)

    # Write data rows.
    start = data_start_row if data_start_row is not None else header_row + 1
    for row_offset, row_data in enumerate(data_rows):
        for col_idx, val in enumerate(row_data, start=1):
            ws.cell(row=start + row_offset, column=col_idx, value=val)

    return wb


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_extract_invoice_items_happy_path_13_fields():
    """Test extract_invoice_items with all 13 fields populated; verify
    InvoiceItem has correct values for all fields and allocated_weight is None.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            [
                "ABC-001",  # part_no
                "PO-100",  # po_no
                100,  # qty
                1.23456,  # price
                123.46,  # amount
                "USD",  # currency
                "TW",  # coo
                "BrandX",  # brand
                "OEM",  # brand_type
                "MOD-1",  # model
                "CN",  # cod
                "INV-2025-001",  # inv_no
                "SN-001",  # serial
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    item = items[0]
    assert item.part_no == "ABC-001"
    assert item.po_no == "PO-100"
    # qty: General format integer -> 0 decimals
    assert item.qty == Decimal("100")
    # price: 5 decimals fixed
    assert item.price == Decimal("1.23456")
    # amount: 2 decimals fixed
    assert item.amount == Decimal("123.46")
    assert item.currency == "USD"
    assert item.coo == "TW"
    assert item.brand == "BrandX"
    assert item.brand_type == "OEM"
    assert item.model_no == "MOD-1"
    assert item.cod == "CN"
    assert item.inv_no == "INV-2025-001"
    assert item.serial == "SN-001"
    assert item.allocated_weight is None


def test_extract_invoice_items_coo_cod_fallback():
    """Test COO/COD fallback: COO column empty, COD has 'CN'; extracted
    InvoiceItem.coo == 'CN'; no ERR_030 raised.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            [
                "PART-X",  # part_no
                "PO-200",  # po_no
                50,  # qty
                2.00000,  # price
                100.00,  # amount
                "USD",  # currency
                "",  # coo -- EMPTY
                "TestBrand",  # brand
                "ODM",  # brand_type
                "MOD-2",  # model
                "CN",  # cod -- has value
                "INV-002",  # inv_no
                "SN-002",  # serial
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].coo == "CN"
    assert items[0].cod == "CN"


def test_extract_invoice_items_stop_at_empty_part_no_qty_zero():
    """Test stop condition: after first data row, row has empty part_no and
    qty=0; extraction stops; only rows before stop included.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            # Row 2: valid data
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", "I1", "S1"],
            # Row 3: empty part_no + qty=0 -> stop
            ["", "PO2", 0, 0.00000, 0.00, "USD", "TW", "B2", "BT2", "M2", "CN", "I2", "S2"],
            # Row 4: should NOT be reached
            ["P3", "PO3", 5, 2.00000, 10.00, "USD", "TW", "B3", "BT3", "M3", "CN", "I3", "S3"],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].part_no == "P1"


def test_extract_invoice_items_stop_at_total_keyword_in_part_no():
    """Test stop condition: row where part_no cell contains 'TOTAL'; extraction
    stops before that row.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", "I1", "S1"],
            # Row 3: part_no contains "TOTAL"
            ["TOTAL", "", 0, 0.00000, 0.00, "", "", "", "", "", "", "", ""],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].part_no == "P1"


def test_extract_invoice_items_stop_at_footer_keyword():
    """Test stop condition: row where part_no contains footer keyword."""
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", "I1", "S1"],
            # Row 3: footer keyword in Chinese
            [
                "XX有限公司",
                "",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].part_no == "P1"


def test_extract_invoice_items_stop_at_any_cell_aj_total():
    """Test stop condition: row where column F contains a stop keyword like
    '合计'; extraction stops.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", "I1", "S1"],
            # Row 3: column F (currency=6) has stop keyword
            ["", "", None, None, None, "合计", None, None, None, None, None, None, None],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].part_no == "P1"


def test_extract_invoice_items_header_continuation_skipped():
    """Test header continuation filtering: row where part_no == 'PART NO.'
    is skipped; the next data row is extracted normally.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            # Row 2: header continuation row
            [
                "PART NO.",
                "PO NO.",
                "QTY",
                "PRICE",
                "AMOUNT",
                "CURRENCY",
                "COO",
                "BRAND",
                "BRAND TYPE",
                "MODEL",
                "COD",
                "INV NO.",
                "SERIAL",
            ],
            # Row 3: real data
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", "I1", "S1"],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].part_no == "P1"


def test_extract_invoice_items_unit_suffix_stripped():
    """Test unit suffix stripping: qty cell value '100 PCS' is parsed as
    Decimal('100') with cell precision; no ERR_031.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            [
                "P1",
                "PO1",
                "100 PCS",  # qty with unit suffix
                1.00000,
                100.00,
                "USD",
                "TW",
                "B1",
                "BT1",
                "M1",
                "CN",
                "I1",
                "S1",
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    assert items[0].qty == Decimal("100")


def test_extract_invoice_items_err030_on_missing_required_field():
    """Test ERR_030: row with empty brand field raises ProcessingError
    with ERR_030, field 'brand', and row number in context.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            [
                "P1",
                "PO1",
                10,
                1.00000,
                10.00,
                "USD",
                "TW",
                "",  # brand -- EMPTY
                "BT1",
                "M1",
                "CN",
                "I1",
                "S1",
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    with pytest.raises(ProcessingError) as exc_info:
        extract_invoice_items(ws, col_map, mt, inv_no=None)

    err = exc_info.value
    assert err.code == ErrorCode.ERR_030
    assert err.context["field_name"] == "brand"
    assert err.context["row"] == 2


def test_extract_invoice_items_err031_on_non_numeric_qty():
    """Test ERR_031: row with 'N/A' in qty column raises ProcessingError
    with ERR_031, row and column in context.
    """
    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            [
                "P1",
                "PO1",
                "N/A",  # qty -- non-numeric
                1.00000,
                10.00,
                "USD",
                "TW",
                "B1",
                "BT1",
                "M1",
                "CN",
                "I1",
                "S1",
            ],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    with pytest.raises(ProcessingError) as exc_info:
        extract_invoice_items(ws, col_map, mt, inv_no=None)

    err = exc_info.value
    assert err.code == ErrorCode.ERR_031
    assert err.context["row"] == 2
    assert err.context["column"] == 3  # qty column


def test_extract_invoice_items_horizontal_merge_brand_brand_type():
    """Test horizontal merge propagation: brand and brand_type columns are
    horizontally merged (anchor at brand col, brand_type col = None after
    unmerge); MergeTracker propagates the anchor value to both fields.
    """
    wb = Workbook()
    ws = wb.active

    # Header row 1.
    headers = [
        "PART NO.",
        "PO NO.",
        "QTY",
        "PRICE",
        "AMOUNT",
        "CURRENCY",
        "COO",
        "BRAND",
        "BRAND TYPE",
        "MODEL",
        "COD",
        "INV NO.",
        "SERIAL",
    ]
    for col_idx, hdr in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=hdr)

    # Data row 2: fill all fields.
    data = [
        "P1",  # A2 part_no
        "PO1",  # B2 po_no
        10,  # C2 qty
        1.00000,  # D2 price
        10.00,  # E2 amount
        "USD",  # F2 currency
        "TW",  # G2 coo
        "无品牌",  # H2 brand (anchor of horizontal merge)
        None,  # I2 brand_type (will be None after unmerge)
        "M1",  # J2 model
        "CN",  # K2 cod
        "I1",  # L2 inv_no
        "S1",  # M2 serial
    ]
    for col_idx, val in enumerate(data, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Create horizontal merge H2:I2 (brand + brand_type).
    ws.merge_cells("H2:I2")

    # Now create MergeTracker which captures and unmerges.
    mt = MergeTracker(ws)
    col_map = _make_column_map()

    items = extract_invoice_items(ws, col_map, mt, inv_no=None)

    assert len(items) == 1
    # Both brand and brand_type should have the anchor value.
    assert items[0].brand == "无品牌"
    assert items[0].brand_type == "无品牌"


def test_extract_invoice_items_inv_no_param_used_when_no_column():
    """Test inv_no parameter fallback: no inv_no column in field_map; all
    extracted items have inv_no == parameter value 'INV-2025-001'.
    """
    # Create field_map without inv_no column.
    field_map_no_inv = {k: v for k, v in FULL_FIELD_MAP.items() if k != "inv_no"}

    wb = _build_invoice_sheet(
        header_row=1,
        data_rows=[
            ["P1", "PO1", 10, 1.00000, 10.00, "USD", "TW", "B1", "BT1", "M1", "CN", None, "S1"],
            ["P2", "PO2", 20, 2.00000, 40.00, "USD", "TW", "B2", "BT2", "M2", "CN", None, "S2"],
        ],
    )
    ws = wb.active
    mt = MergeTracker(ws)
    col_map = _make_column_map(field_map=field_map_no_inv)

    items = extract_invoice_items(ws, col_map, mt, inv_no="INV-2025-001")

    assert len(items) == 2
    assert items[0].inv_no == "INV-2025-001"
    assert items[1].inv_no == "INV-2025-001"
