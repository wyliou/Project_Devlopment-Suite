"""Tests for config.py — FR-002: Load and Validate All Configuration Files."""

import re
from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from autoconvert.config import load_config
from autoconvert.errors import ConfigError, ErrorCode

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _create_minimal_yaml(config_dir: Path) -> Path:
    """Create a minimal valid field_patterns.yaml in config_dir.

    Returns:
        Path to the created YAML file.
    """
    yaml_path = config_dir / "field_patterns.yaml"
    data = {
        "invoice_sheet": {"patterns": ["^invoice"]},
        "packing_sheet": {"patterns": ["^packing"]},
        "invoice_columns": {
            "part_no": {"patterns": ["^part"], "type": "string", "required": True},
            "po_no": {"patterns": ["^po"], "type": "string", "required": True},
            "qty": {"patterns": ["^qty"], "type": "numeric", "required": True},
            "price": {"patterns": ["^price"], "type": "numeric", "required": True},
            "amount": {"patterns": ["^amount"], "type": "currency", "required": True},
            "currency": {"patterns": ["^currency"], "type": "string", "required": True},
            "coo": {"patterns": ["^coo"], "type": "string", "required": True},
            "brand": {"patterns": ["^brand"], "type": "string", "required": True},
            "brand_type": {"patterns": ["^brand_type"], "type": "string", "required": True},
            "model": {"patterns": ["^model"], "type": "string", "required": True},
            "cod": {"patterns": ["^cod"], "type": "string", "required": False},
            "weight": {"patterns": ["^weight"], "type": "numeric", "required": False},
            "inv_no": {"patterns": ["^inv_no"], "type": "string", "required": False},
            "serial": {"patterns": ["^serial"], "type": "string", "required": False},
        },
        "packing_columns": {
            "part_no": {"patterns": ["^part"], "type": "string", "required": True},
            "po_no": {"patterns": ["^po"], "type": "string", "required": False},
            "qty": {"patterns": ["^qty"], "type": "numeric", "required": True},
            "nw": {"patterns": ["^nw"], "type": "numeric", "required": True},
            "gw": {"patterns": ["^gw"], "type": "numeric", "required": True},
            "pack": {"patterns": ["^pack"], "type": "numeric", "required": False},
        },
        "inv_no_cell": {
            "patterns": [r"INVOICE\s*NO\.?\s*[:：]\s*(\S+)"],
            "label_patterns": [r"^INV\.?\s*NO\.?\s*[:：]?$"],
            "exclude_patterns": [r"(?i)^invoice\s*no\.?[:：]?"],
        },
    }
    yaml_path.write_text(yaml.dump(data, allow_unicode=True), encoding="utf-8")
    return yaml_path


def _create_currency_xlsx(
    config_dir: Path,
    *,
    sheet_name: str = "Currency_Rules",
    rows: list[list[object]] | None = None,
) -> Path:
    """Create a currency_rules.xlsx in config_dir.

    Args:
        config_dir: Target directory.
        sheet_name: Sheet name to use.
        rows: Data rows (each row is [Source_Value, Target_Code]).
            Defaults to a minimal valid dataset.

    Returns:
        Path to the created file.
    """
    if rows is None:
        rows = [
            ["USD", 502],
            ["\u7f8e\u5143", 502],  # 美元
        ]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    ws.append(["Source_Value", "Target_Code"])
    for row in rows:
        ws.append(row)

    path = config_dir / "currency_rules.xlsx"
    wb.save(path)
    wb.close()
    return path


def _create_country_xlsx(
    config_dir: Path,
    *,
    sheet_name: str = "Country_Rules",
    rows: list[list[object]] | None = None,
) -> Path:
    """Create a country_rules.xlsx in config_dir.

    Args:
        config_dir: Target directory.
        sheet_name: Sheet name to use.
        rows: Data rows (each row is [Source_Value, Target_Code]).
            Defaults to a minimal valid dataset including string Target_Code entries.

    Returns:
        Path to the created file.
    """
    if rows is None:
        rows = [
            ["TAIWAN,CHINA", 143],
            ["JAPAN", 116],
            ["JP", "116"],  # String Target_Code
            ["Taiwan, China", 143],  # NOTE: different raw but will conflict - use unique
            ["CHINA", 142],
            ["CN", 142],
        ]
        # Reason: Avoid duplicate "Taiwan, China" vs "TAIWAN,CHINA" by using a
        # distinct set that includes string Target_Code values for testing.
        rows = [
            ["TAIWAN", 143],
            ["JAPAN", 116],
            ["JP", "116"],  # String Target_Code
            ["PH", "129"],  # String Target_Code
            ["CHINA", 142],
            ["CN", 142],
            ["Taiwan, China", 143],  # Has comma-space for normalization test
        ]

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name
    ws.append(["Source_Value", "Target_Code"])
    for row in rows:
        ws.append(row)

    path = config_dir / "country_rules.xlsx"
    wb.save(path)
    wb.close()
    return path


def _create_template_xlsx(
    config_dir: Path,
    *,
    sheet_name: str = "\u5de5\u4f5c\u88681",  # 工作表1
    num_columns: int = 40,
    num_rows: int = 4,
) -> Path:
    """Create a minimal output_template.xlsx in config_dir.

    Args:
        config_dir: Target directory.
        sheet_name: Sheet name to create.
        num_columns: Number of columns to populate in row 1.
        num_rows: Number of rows to populate.

    Returns:
        Path to the created file.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet_name

    for r in range(1, num_rows + 1):
        for c in range(1, num_columns + 1):
            ws.cell(row=r, column=c, value=f"H{r}C{c}")

    path = config_dir / "output_template.xlsx"
    wb.save(path)
    wb.close()
    return path


def _create_all_config_files(config_dir: Path) -> None:
    """Create all 4 required config files in config_dir."""
    _create_minimal_yaml(config_dir)
    _create_currency_xlsx(config_dir)
    _create_country_xlsx(config_dir)
    _create_template_xlsx(config_dir)


# ---------------------------------------------------------------------------
# Happy Path Tests
# ---------------------------------------------------------------------------


class TestLoadConfigSuccess:
    """Tests for load_config happy path scenarios."""

    def test_load_config_success(self, tmp_path: Path) -> None:
        """Test load_config with all valid config files returns AppConfig."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_all_config_files(config_dir)

        result = load_config(config_dir)

        # Verify returned AppConfig structure
        assert isinstance(result.invoice_sheet_patterns, list)
        assert len(result.invoice_sheet_patterns) > 0
        assert isinstance(result.invoice_sheet_patterns[0], re.Pattern)

        assert isinstance(result.packing_sheet_patterns, list)
        assert len(result.packing_sheet_patterns) > 0

        assert isinstance(result.currency_lookup, dict)
        assert len(result.currency_lookup) > 0
        # Keys should be normalized (uppercase)
        for key in result.currency_lookup:
            assert key == key.upper() or not key.isascii()

        assert isinstance(result.template_path, Path)
        assert result.template_path.exists()

        # Verify inv_no patterns
        assert isinstance(result.inv_no_patterns, list)
        assert len(result.inv_no_patterns) > 0
        assert isinstance(result.inv_no_label_patterns, list)
        assert isinstance(result.inv_no_exclude_patterns, list)

        # Verify field patterns
        assert "part_no" in result.invoice_columns
        assert "qty" in result.packing_columns

    def test_load_config_currency_lookup_normalized(self, tmp_path: Path) -> None:
        """Test currency_lookup has normalized keys and string Target_Code values.

        currency_rules.xlsx has Source_Value="美元", Target_Code=502 (int);
        verify app_config.currency_lookup["美元"] == "502" (string).
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(
            config_dir,
            rows=[
                ["USD", 502],
                ["\u7f8e\u5143", 502],  # 美元
            ],
        )
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        result = load_config(config_dir)

        # 美元 normalized key (normalize_lookup_key: strip, upper, collapse ", ")
        assert result.currency_lookup["\u7f8e\u5143"] == "502"
        assert result.currency_lookup["USD"] == "502"
        # Verify type is str, not int
        assert isinstance(result.currency_lookup["USD"], str)

    def test_load_config_country_lookup_target_code_str(self, tmp_path: Path) -> None:
        """Test ALL Target_Code values in country_lookup are str type.

        country_rules.xlsx has 3 entries with string Target_Code and others with int;
        verify ALL values are str.
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(
            config_dir,
            rows=[
                ["TAIWAN", 143],  # int Target_Code
                ["JAPAN", 116],  # int Target_Code
                ["JP", "116"],  # string Target_Code
                ["PH", "129"],  # string Target_Code
                ["INDONESIA", "112"],  # string Target_Code
                ["CHINA", 142],  # int Target_Code
            ],
        )
        _create_template_xlsx(config_dir)

        result = load_config(config_dir)

        # ALL values must be str type
        for key, value in result.country_lookup.items():
            assert isinstance(value, str), (
                f"country_lookup[{key!r}] = {value!r} is {type(value).__name__}, expected str"
            )

        # Verify specific values
        assert result.country_lookup["TAIWAN"] == "143"
        assert result.country_lookup["JP"] == "116"
        assert result.country_lookup["PH"] == "129"
        assert result.country_lookup["INDONESIA"] == "112"

    def test_load_config_country_lookup_comma_space_key(self, tmp_path: Path) -> None:
        """Test country_lookup normalizes 'Taiwan, China' to 'TAIWAN,CHINA'.

        normalize_lookup_key() collapses comma-space to comma and uppercases.
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(
            config_dir,
            rows=[
                ["Taiwan, China", 143],
                ["JAPAN", 116],
            ],
        )
        _create_template_xlsx(config_dir)

        result = load_config(config_dir)

        # "Taiwan, China" -> normalize_lookup_key -> "TAIWAN,CHINA"
        assert "TAIWAN,CHINA" in result.country_lookup
        assert result.country_lookup["TAIWAN,CHINA"] == "143"

    def test_load_config_field_patterns_compiled(self, tmp_path: Path) -> None:
        """Test FieldPattern objects store raw patterns and sheet patterns are compiled.

        YAML has regex patterns for invoice_columns.part_no; verify
        app_config.invoice_columns["part_no"].patterns is a list[str] and
        app_config.invoice_sheet_patterns is a list[re.Pattern].
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_all_config_files(config_dir)

        result = load_config(config_dir)

        # FieldPattern.patterns stores raw strings
        part_no_fp = result.invoice_columns["part_no"]
        assert isinstance(part_no_fp.patterns, list)
        assert len(part_no_fp.patterns) > 0
        assert isinstance(part_no_fp.patterns[0], str)

        # Sheet patterns are compiled re.Pattern objects
        assert isinstance(result.invoice_sheet_patterns, list)
        assert len(result.invoice_sheet_patterns) > 0
        assert isinstance(result.invoice_sheet_patterns[0], re.Pattern)

        # Verify packing columns too
        assert "nw" in result.packing_columns
        assert result.packing_columns["nw"].required is True
        assert result.packing_columns["pack"].required is False


# ---------------------------------------------------------------------------
# ERR_001 Tests — Missing Files
# ---------------------------------------------------------------------------


class TestLoadConfigErr001:
    """Tests for ERR_001: missing configuration files."""

    def test_load_config_missing_yaml_raises_err001(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_001 when field_patterns.yaml is missing."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        # Do NOT create YAML file
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "field_patterns.yaml" in exc_info.value.path

    def test_load_config_missing_currency_xlsx_raises_err001(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_001 when currency_rules.xlsx is missing."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        # Do NOT create currency file
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "currency_rules.xlsx" in exc_info.value.path

    def test_load_config_missing_country_xlsx_raises_err001(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_001 when country_rules.xlsx is missing."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        # Do NOT create country file
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "country_rules.xlsx" in exc_info.value.path

    def test_load_config_missing_template_raises_err001(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_001 when output_template.xlsx is missing."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        # Do NOT create template file

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_001
        assert "output_template.xlsx" in exc_info.value.path


# ---------------------------------------------------------------------------
# ERR_002 Tests — Invalid Regex
# ---------------------------------------------------------------------------


class TestLoadConfigErr002:
    """Tests for ERR_002: invalid regex pattern compilation."""

    def test_load_config_invalid_regex_raises_err002(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_002 when YAML contains an invalid regex.

        YAML contains invoice_columns.part_no.patterns: ["[invalid"]; verify
        ConfigError with code == ErrorCode.ERR_002; message includes
        pattern_name and error description.
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        # Create YAML with an invalid regex in invoice_columns.part_no
        yaml_path = config_dir / "field_patterns.yaml"
        data = {
            "invoice_sheet": {"patterns": ["^invoice"]},
            "packing_sheet": {"patterns": ["^packing"]},
            "invoice_columns": {
                "part_no": {"patterns": ["[invalid"], "type": "string", "required": True},
                "po_no": {"patterns": ["^po"], "type": "string", "required": True},
                "qty": {"patterns": ["^qty"], "type": "numeric", "required": True},
                "price": {"patterns": ["^price"], "type": "numeric", "required": True},
                "amount": {"patterns": ["^amount"], "type": "currency", "required": True},
                "currency": {"patterns": ["^currency"], "type": "string", "required": True},
                "coo": {"patterns": ["^coo"], "type": "string", "required": True},
                "brand": {"patterns": ["^brand"], "type": "string", "required": True},
                "brand_type": {"patterns": ["^brand_type"], "type": "string", "required": True},
                "model": {"patterns": ["^model"], "type": "string", "required": True},
                "cod": {"patterns": ["^cod"], "type": "string", "required": False},
                "weight": {"patterns": ["^weight"], "type": "numeric", "required": False},
                "inv_no": {"patterns": ["^inv_no"], "type": "string", "required": False},
                "serial": {"patterns": ["^serial"], "type": "string", "required": False},
            },
            "packing_columns": {
                "part_no": {"patterns": ["^part"], "type": "string", "required": True},
                "po_no": {"patterns": ["^po"], "type": "string", "required": False},
                "qty": {"patterns": ["^qty"], "type": "numeric", "required": True},
                "nw": {"patterns": ["^nw"], "type": "numeric", "required": True},
                "gw": {"patterns": ["^gw"], "type": "numeric", "required": True},
                "pack": {"patterns": ["^pack"], "type": "numeric", "required": False},
            },
            "inv_no_cell": {
                "patterns": [r"INVOICE\s*NO\.?\s*[:：]\s*(\S+)"],
                "label_patterns": [r"^INV\.?\s*NO\.?\s*[:：]?$"],
                "exclude_patterns": [r"(?i)^invoice\s*no\.?[:：]?"],
            },
        }
        yaml_path.write_text(yaml.dump(data, allow_unicode=True), encoding="utf-8")

        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_002
        # Verify message includes pattern identification and error description
        assert "part_no" in exc_info.value.message
        assert "pattern[0]" in exc_info.value.message


# ---------------------------------------------------------------------------
# ERR_003 Tests — Duplicate Source_Value
# ---------------------------------------------------------------------------


class TestLoadConfigErr003:
    """Tests for ERR_003: duplicate Source_Value entries."""

    def test_load_config_duplicate_source_value_raises_err003(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_003 when currency_rules.xlsx has duplicate Source_Value.

        Two rows with same Source_Value (e.g., "USD" twice); verify ConfigError
        with code == ErrorCode.ERR_003; message includes the duplicate value.
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(
            config_dir,
            rows=[
                ["USD", 502],
                ["USD", 503],  # Duplicate!
            ],
        )
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_003
        assert "USD" in exc_info.value.message

    def test_load_config_duplicate_source_value_case_insensitive(self, tmp_path: Path) -> None:
        """Test ERR_003 for case-insensitive duplicate detection in country_rules.

        country_rules.xlsx has "china" and "CHINA" as Source_Value entries;
        both normalize to "CHINA" via normalize_lookup_key(), triggering ERR_003.
        """
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(
            config_dir,
            rows=[
                ["china", 142],
                ["CHINA", 142],  # Same after normalization
            ],
        )
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_003
        assert "CHINA" in exc_info.value.message


# ---------------------------------------------------------------------------
# ERR_004 Tests — Malformed Configuration
# ---------------------------------------------------------------------------


class TestLoadConfigErr004:
    """Tests for ERR_004: malformed configuration files."""

    def test_load_config_malformed_yaml_missing_key_raises_err004(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_004 when YAML is missing the invoice_sheet key."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()

        yaml_path = config_dir / "field_patterns.yaml"
        # Intentionally omit "invoice_sheet"
        data = {
            "packing_sheet": {"patterns": ["^packing"]},
            "invoice_columns": {},
            "packing_columns": {},
            "inv_no_cell": {
                "patterns": [],
                "label_patterns": [],
                "exclude_patterns": [],
            },
        }
        yaml_path.write_text(yaml.dump(data, allow_unicode=True), encoding="utf-8")

        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_004
        assert "invoice_sheet" in exc_info.value.message

    def test_load_config_malformed_currency_sheet_raises_err004(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_004 when currency_rules.xlsx lacks Currency_Rules sheet."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        # Create currency xlsx with wrong sheet name
        _create_currency_xlsx(config_dir, sheet_name="WrongSheet")
        _create_country_xlsx(config_dir)
        _create_template_xlsx(config_dir)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_004
        assert "Currency_Rules" in exc_info.value.message


# ---------------------------------------------------------------------------
# ERR_005 Tests — Template Validation
# ---------------------------------------------------------------------------


class TestLoadConfigErr005:
    """Tests for ERR_005: output template structure validation."""

    def test_load_config_template_missing_sheet_raises_err005(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_005 when output_template.xlsx lacks required sheet."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        # Create template with wrong sheet name
        _create_template_xlsx(config_dir, sheet_name="WrongSheet")

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_005
        assert "\u5de5\u4f5c\u88681" in exc_info.value.message  # 工作表1

    def test_load_config_template_insufficient_columns_raises_err005(self, tmp_path: Path) -> None:
        """Test ConfigError ERR_005 when template has only 30 columns (need >= 40)."""
        config_dir = tmp_path / "config"
        config_dir.mkdir()
        _create_minimal_yaml(config_dir)
        _create_currency_xlsx(config_dir)
        _create_country_xlsx(config_dir)
        # Create template with only 30 columns
        _create_template_xlsx(config_dir, num_columns=30)

        with pytest.raises(ConfigError) as exc_info:
            load_config(config_dir)

        assert exc_info.value.code == ErrorCode.ERR_005
        assert "30" in exc_info.value.message


# ---------------------------------------------------------------------------
# Integration Test — Real Config Files
# ---------------------------------------------------------------------------


class TestLoadConfigRealFiles:
    """Integration tests using the actual project config files."""

    def test_load_real_config_files(self, config_dir: Path) -> None:
        """Test load_config with the real project config directory.

        Verifies the real config files load without error and produce
        the expected AppConfig structure.
        """
        result = load_config(config_dir)

        # Verify invoice sheet patterns (from real YAML)
        assert len(result.invoice_sheet_patterns) >= 3
        assert all(isinstance(p, re.Pattern) for p in result.invoice_sheet_patterns)

        # Verify packing sheet patterns
        assert len(result.packing_sheet_patterns) >= 4
        assert all(isinstance(p, re.Pattern) for p in result.packing_sheet_patterns)

        # Verify all 14 invoice columns present
        expected_invoice_fields = {
            "part_no",
            "po_no",
            "qty",
            "price",
            "amount",
            "currency",
            "coo",
            "brand",
            "brand_type",
            "model",
            "cod",
            "weight",
            "inv_no",
            "serial",
        }
        assert set(result.invoice_columns.keys()) == expected_invoice_fields

        # Verify all 6 packing columns present
        expected_packing_fields = {"part_no", "po_no", "qty", "nw", "gw", "pack"}
        assert set(result.packing_columns.keys()) == expected_packing_fields

        # Verify inv_no_cell patterns
        assert len(result.inv_no_patterns) >= 3
        assert len(result.inv_no_label_patterns) >= 4
        assert len(result.inv_no_exclude_patterns) >= 10

        # Verify currency lookup
        assert "USD" in result.currency_lookup
        assert result.currency_lookup["USD"] == "502"
        # 美元 normalized
        assert "\u7f8e\u5143" in result.currency_lookup
        assert result.currency_lookup["\u7f8e\u5143"] == "502"

        # Verify country lookup
        assert "JAPAN" in result.country_lookup
        assert result.country_lookup["JAPAN"] == "116"
        # String Target_Code entries are also str
        assert result.country_lookup["JP"] == "116"

        # Verify template path
        assert result.template_path.exists()
        assert result.template_path.name == "output_template.xlsx"

    def test_real_config_country_lookup_all_str(self, config_dir: Path) -> None:
        """Verify ALL Target_Code values in real country_lookup are str type."""
        result = load_config(config_dir)

        for key, value in result.country_lookup.items():
            assert isinstance(value, str), (
                f"country_lookup[{key!r}] = {value!r} is {type(value).__name__}, expected str"
            )

    def test_real_config_comma_space_normalization(self, config_dir: Path) -> None:
        """Verify comma-space normalization in real country_rules.xlsx.

        'Taiwan, China' should be normalized to 'TAIWAN,CHINA'.
        """
        result = load_config(config_dir)

        # The real file has entries with comma-space that get normalized
        assert "TAIWAN,CHINA" in result.country_lookup
        assert result.country_lookup["TAIWAN,CHINA"] == "143"

    def test_real_config_float_target_code_normalization(self, config_dir: Path) -> None:
        """Verify float Target_Code values are normalized (e.g., 502.0 -> '502')."""
        result = load_config(config_dir)

        # All currency Target_Code values should be clean integers as strings
        for key, value in result.currency_lookup.items():
            assert "." not in value, f"currency_lookup[{key!r}] = {value!r} has decimal point"
